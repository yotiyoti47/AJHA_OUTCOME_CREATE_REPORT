import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

def getSQL_CI_47_Door_to_balloon実施率(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_47" + _C + ".年度 ,  " + \
	            "CI_47" + _C + ".期,  " + \
                "CI_47" + _C + ".Door_to_balloon実施率, " + \
                "CI_47" + _C + ".分母のうち_入院日もしくは翌日に_経皮的冠動脈形成術もしくは経皮的冠動脈ステント留置術を実施した患者数, " + \
                "CI_47" + _C + "._18歳以上の急性心筋梗塞でPCIを受けた患者数 " + \
            "FROM " + \
	            "CI_47" + _C + " " + \
            "WHERE " + \
	            "NOT CI_47" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_47" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_47" + _C + ".年度, CI_47" + _C + ".期 " 

def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_47 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_47" in wb.sheetnames:
        wb.remove(wb["CI_47"])
    wb.create_sheet("CI_47")
    sheet = wb["CI_47"]    

    # 書込み列の位置
    colCnt = 1

    # 書込み列の繰り返し数
    roopColCnt = 12
    
    #Door_to_balloon実施率
    tempDoor_to_balloon実施率 = X_01.excuteSQL(getSQL_CI_47_Door_to_balloon実施率(HOSPITAL_ID, _C))

    # 1行目に病院名
    sheet.cell(1, 1 + (colCnt - 1) * roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, 1 + (colCnt - 1) * roopColCnt).value = "Door_to_balloon実施率"
    # 3行目に疾患名
    sheet.cell(3, 1 + (colCnt - 1) * roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, 1 + (colCnt - 1) * roopColCnt).value = ""

    # 5行名からヘッダを出力
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt).value = "年度"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 1).value = "期"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 2).value = "Door_to_balloon実施率"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 3).value = "分母のうち_入院日もしくは翌日に_経皮的冠動脈形成術もしくは経皮的冠動脈ステント留置術を実施した患者数"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 4).value = "_18歳以上の急性心筋梗塞でPCIを受けた患者数"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 5).value = "Door_to_balloon実施率_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 6).value = "分母のうち_入院日もしくは翌日に_経皮的冠動脈形成術もしくは経皮的冠動脈ステント留置術を実施した患者数_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 7).value = "_18歳以上の急性心筋梗塞でPCIを受けた患者数_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 8).value = "Door_to_balloon実施率_比較用"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 9).value = "分母のうち_入院日もしくは翌日に_経皮的冠動脈形成術もしくは経皮的冠動脈ステント留置術を実施した患者数_比較用"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 10).value = "_18歳以上の急性心筋梗塞でPCIを受けた患者数_比較用"

    if tempDoor_to_balloon実施率 is None:
        print("エラー: Door_to_balloon実施率の取得に失敗しました。")
        return -1

    # 6行目からデータ入力
    rowCnt = 6
    for tempRow in tempDoor_to_balloon実施率:

        #年度
        sheet.cell(rowCnt, 1 + (colCnt - 1) * roopColCnt).value = tempRow[0]
        #期
        sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 1).value = tempRow[1]
        
        #Door_to_balloon実施率
        if tempRow[2] == -1 or tempRow[2] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = tempRow[2]  / 100
        #分母のうち_入院日もしくは翌日に_経皮的冠動脈形成術もしくは経皮的冠動脈ステント留置術を実施した患者数
        if tempRow[3] == -1 or tempRow[3] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = tempRow[3] 
        #_18歳以上の急性心筋梗塞でPCIを受けた患者数
        if tempRow[4] == -1 or tempRow[4] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 4).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 4).value = tempRow[4] 

        #Door_to_balloon実施率_ラベル
        if tempRow[2] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "N/A"
        elif tempRow[2] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = tempRow[2] / 100
        #分母のうち_入院日もしくは翌日に_経皮的冠動脈形成術もしくは経皮的冠動脈ステント留置術を実施した患者数_ラベル
        if tempRow[3] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = "N/A"
        elif tempRow[3] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = tempRow[3]
        #_18歳以上の急性心筋梗塞でPCIを受けた患者数_ラベル
        if tempRow[4] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = "N/A"
        elif tempRow[4] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = tempRow[4]

        #Door_to_balloon実施率_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 8).value = tempRow[2] / 100
        #分母のうち_入院日もしくは翌日に_経皮的冠動脈形成術もしくは経皮的冠動脈ステント留置術を実施した患者数_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 9).value = tempRow[3]
        #_18歳以上の急性心筋梗塞でPCIを受けた患者数_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 10).value = tempRow[4]

        rowCnt+=1
    colCnt+=1
    #wb.save(wbPath)
    print(" " + HOSPITAL_NAME + " CI_47 終了")


