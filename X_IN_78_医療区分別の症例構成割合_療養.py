import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

# 2024年度にCI番号変更
# CI_125　→　CI_78
# ただし、DBのテーブルは「CI_125」のまま
def getSQL_CI_78_医療区分別の症例構成割合_療養(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_125" + _C + ".年度 ,  " + \
	            "CI_125" + _C + ".月, " + \
                "CI_125" + _C + ".療養病床_主目的, " + \
                "CI_125" + _C + ".病床数, " + \
                "CI_125" + _C + ".症例数, " + \
                "CI_125" + _C + ".延べ在院日数, " + \
                "CI_125" + _C + ".医療区分1, " + \
                "CI_125" + _C + ".医療区分2, " + \
                "CI_125" + _C + ".医療区分3 " + \
            "FROM " + \
	            "CI_125" + _C + " " + \
            "WHERE " + \
	            "CI_125" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_125" + _C + ".年度, CI_125" + _C + ".月 " 


def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_78 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_78" in wb.sheetnames:
        wb.remove(wb["CI_78"])
    wb.create_sheet("CI_78")
    sheet = wb["CI_78"]    

    #医療区分別の症例構成割合_療養
    temp医療区分別の症例構成割合_療養 = X_01.excuteSQL(getSQL_CI_78_医療区分別の症例構成割合_療養(HOSPITAL_ID, _C))

    if temp医療区分別の症例構成割合_療養 is None:
        print("エラー: 医療区分別の症例構成割合_療養の取得に失敗しました。")
        return -1

    keyList = ["年度平均", "四月", "五月", "六月", 
                "七月", "八月", "九月", "十月", "十一月", "十二月", "一月", 
                "二月", "三月" ]
    yearList = []
    dict_temp = {}

    for a in temp医療区分別の症例構成割合_療養:
        # yearListに年が存在していなければ追加
        if a[0] not in yearList:
            yearList.append(a[0])
        #年と月をキーに辞書に追加
        dict_temp[str(a[0]) + str(a[1])] = a

    
    # 書込み列の繰り返し数
    roopColCnt = 1
    
    # 1行目に病院名
    sheet.cell(1, roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, roopColCnt).value = "医療区分別の症例構成割合_療養"
    # 3行目に疾患名
    sheet.cell(3, roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, roopColCnt).value = ""

    # 5行名からヘッダを出力
    loopHeaderList = ["医療区分1", "医療区分2", "医療区分3", ]

    sheet.cell(5, roopColCnt).value = "病床数"
    tmpCnt = 1
    sheet.cell(5, roopColCnt + tmpCnt).value = "療養病床_主目的"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "症例数"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "年度"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "月"
    tmpCnt+=1

    #ヘッダ　グラフ作成用
    for a in loopHeaderList:
        sheet.cell(5, roopColCnt + tmpCnt).value = a
        tmpCnt+=1
    
    #ヘッダ　ラベル用
    for a in loopHeaderList:
        sheet.cell(5, roopColCnt + tmpCnt).value = a + "_ラベル"
        tmpCnt+=1
    
    #ヘッダ　比較用
    for a in loopHeaderList:
        sheet.cell(5, roopColCnt + tmpCnt).value = a + "_比較用"
        tmpCnt+=1


    # 6行目からデータ入力
    rowCnt = 6
    for year in yearList:
        for key in keyList:
            tempRow = dict_temp[str(year) + str(key)]
            #病床数
            sheet.cell(rowCnt, roopColCnt).value = tempRow[3]
            #療養病床_主目的
            sheet.cell(rowCnt, roopColCnt + 1).value = tempRow[2]
            #症例数
            if tempRow[1] == "年度平均":
                sheet.cell(rowCnt, roopColCnt + 2).value = tempRow[4]
            else:
                sheet.cell(rowCnt, roopColCnt + 2).value = ""
            #年度
            sheet.cell(rowCnt, roopColCnt + 3).value = tempRow[0]
            #月
            sheet.cell(rowCnt, roopColCnt + 4).value = tempRow[1]

            for num in range(len(loopHeaderList)):
                #グラフ作成用数値
                if tempRow[6 + num] == -1 or tempRow[6 + num] == -2:
                    sheet.cell(rowCnt, roopColCnt + 5 + num).value = 0
                else:
                    sheet.cell(rowCnt, roopColCnt + 5 + num).value = tempRow[6 + num] 

                #ラベル用数値
                if tempRow[6 + num] == -1:
                    sheet.cell(rowCnt, roopColCnt + 5 + num + len(loopHeaderList)).value = "N/A"
                elif tempRow[6 + num] == -2:
                    sheet.cell(rowCnt, roopColCnt + 5 + num + len(loopHeaderList)).value = "-"
                else:
                    sheet.cell(rowCnt, roopColCnt + 5 + num + len(loopHeaderList)).value = tempRow[6 + num]

                #比較用
                sheet.cell(rowCnt, roopColCnt + 5 + num + len(loopHeaderList) * 2).value = tempRow[6 + num]
            rowCnt += 1


    # 書込み列の繰り返し数
    roopColCnt = 16
    
    # 1行目に病院名
    sheet.cell(1, roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, roopColCnt).value = "延べ在院日数"
    # 3行目に疾患名
    sheet.cell(3, roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, roopColCnt).value = ""

    # 5行名からヘッダを出力
    loopHeaderList = ["延べ在院日数", ]

    sheet.cell(5, roopColCnt).value = "症例数"
    tmpCnt = 1
    sheet.cell(5, roopColCnt + tmpCnt).value = "年度"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "月"
    tmpCnt+=1

    #ヘッダ　グラフ作成用
    for a in loopHeaderList:
        sheet.cell(5, roopColCnt + tmpCnt).value = a
        tmpCnt+=1
    
    #ヘッダ　ラベル用
    for a in loopHeaderList:
        sheet.cell(5, roopColCnt + tmpCnt).value = a + "_ラベル"
        tmpCnt+=1
    
    #ヘッダ　比較用
    for a in loopHeaderList:
        sheet.cell(5, roopColCnt + tmpCnt).value = a + "_比較用"
        tmpCnt+=1

    # 6行目からデータ入力
    rowCnt = 6
    for year in yearList:
        for key in keyList:
            tempRow = dict_temp[str(year) + str(key)]
            #症例数
            if tempRow[1] == "年度平均":
                sheet.cell(rowCnt, roopColCnt).value = tempRow[4]
            else:
                sheet.cell(rowCnt, roopColCnt).value = ""
            #年度
            sheet.cell(rowCnt, roopColCnt + 1).value = tempRow[0]
            #月
            sheet.cell(rowCnt, roopColCnt + 2).value = tempRow[1]

            for num in range(len(loopHeaderList)):
                #グラフ作成用数値
                if tempRow[5 + num] == -1 or tempRow[5 + num] == -2:
                    sheet.cell(rowCnt, roopColCnt + 3 + num).value = 0
                else:
                    sheet.cell(rowCnt, roopColCnt + 3 + num).value = tempRow[5 + num] 

                #ラベル用数値
                if tempRow[5 + num] == -1:
                    sheet.cell(rowCnt, roopColCnt + 3 + num + len(loopHeaderList)).value = "N/A"
                elif tempRow[5 + num] == -2:
                    sheet.cell(rowCnt, roopColCnt + 3 + num + len(loopHeaderList)).value = "-"
                else:
                    sheet.cell(rowCnt, roopColCnt + 3 + num + len(loopHeaderList)).value = tempRow[5 + num]

                #比較用
                sheet.cell(rowCnt, roopColCnt + 3 + num + len(loopHeaderList) * 2).value = tempRow[5 + num]
            rowCnt += 1

    print(" " + HOSPITAL_NAME + " CI_78 終了")

