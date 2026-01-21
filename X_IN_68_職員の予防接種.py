import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

# 2024年度にCI番号変更
# CI_94　→　CI_68
# ただし、DBのテーブルは「CI_94」のまま
def getSQL_CI_68_職員の予防接種(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_94" + _C + ".年度 ,  " + \
	            "CI_94" + _C + ".割合, " + \
                "CI_94" + _C + ".分子, " + \
                "CI_94" + _C + ".分母 " + \
            "FROM " + \
	            "CI_94" + _C + " " + \
            "WHERE " + \
	            "CI_94" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_94" + _C + ".年度 " 

def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_68 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_68" in wb.sheetnames:
        wb.remove(wb["CI_68"])
    wb.create_sheet("CI_68")
    sheet = wb["CI_68"]    

    # 書込み列の位置
    colCnt = 1

    #職員の予防接種
    temp職員の予防接種 = X_01.excuteSQL(getSQL_CI_68_職員の予防接種(HOSPITAL_ID, _C))

    if temp職員の予防接種 is None:
        print("エラー: 職員の予防接種の取得に失敗しました。")
        return -1

    # 1行目に病院名
    sheet.cell(1, 1 + (colCnt - 1) * 11).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, 1 + (colCnt - 1) * 11).value = "職員の予防接種"
    # 3行目に疾患名
    sheet.cell(3, 1 + (colCnt - 1) * 11).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, 1 + (colCnt - 1) * 11).value = ""

    # 5行名からヘッダを出力
    sheet.cell(5, 1 + (colCnt - 1) * 11).value = "年度"
    sheet.cell(5, 1 + (colCnt - 1) * 11 + 1).value = "割合"
    sheet.cell(5, 1 + (colCnt - 1) * 11 + 2).value = "分子"
    sheet.cell(5, 1 + (colCnt - 1) * 11 + 3).value = "分母"
    sheet.cell(5, 1 + (colCnt - 1) * 11 + 4).value = "割合_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * 11 + 5).value = "分子_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * 11 + 6).value = "分母_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * 11 + 7).value = "割合_比較用"
    sheet.cell(5, 1 + (colCnt - 1) * 11 + 8).value = "分子_比較用"
    sheet.cell(5, 1 + (colCnt - 1) * 11 + 9).value = "分母_比較用"

    # 6行目からデータ入力
    rowCnt = 6
    for tempRow in temp職員の予防接種:

        #年度
        sheet.cell(rowCnt, 1 + (colCnt - 1) * 11).value = tempRow[0]
        
        #割合
        if tempRow[1] == -1 or tempRow[1] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * 11 + 1).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * 11 + 1).value = tempRow[1] 
        #分子
        if tempRow[2] == -1 or tempRow[2] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * 11 + 2).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * 11 + 2).value = tempRow[2] 
        #分母
        if tempRow[3] == -1 or tempRow[3] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * 11 + 3).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * 11 + 3).value = tempRow[3] 

        #割合_ラベル
        if tempRow[1] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * 11 + 4).value = "N/A"
        elif tempRow[1] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * 11 + 4).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * 11 + 4).value = tempRow[1]
        #分子_ラベル
        if tempRow[2] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * 11 + 5).value = "N/A"
        elif tempRow[2] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * 11 + 5).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * 11 + 5).value = tempRow[2]
        #分母_ラベル
        if tempRow[3] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * 11 + 6).value = "N/A"
        elif tempRow[3] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * 11 + 6).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * 11 + 6).value = tempRow[3]

        #割合_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * 11 + 7).value = tempRow[1]
        #分子_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * 11 + 8).value = tempRow[2]
        #分母_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * 11 + 9).value = tempRow[3]

        rowCnt+=1
    colCnt+=1
    #wb.save(wbPath)
    print(" " + HOSPITAL_NAME + " CI_68 終了")


