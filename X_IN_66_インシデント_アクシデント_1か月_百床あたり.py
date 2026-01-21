import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

# 2024年度にCI番号変更
# CI_90　→　CI_66
# ただし、DBのテーブルは「CI_90」のまま
def getSQL_CI_66_インシデント_アクシデント_1か月_百床あたり_件数(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_90" + _C + ".病床数 ,  " + \
                "CI_90" + _C + ".年度 ,  " + \
	            "CI_90" + _C + ".月,  " + \
                "CI_90" + _C + ".一か月_百床あたり_発生件数 " + \
            "FROM " + \
	            "CI_90" + _C + " " + \
            "WHERE " + \
	            "CI_90" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_90" + _C + ".年度, CI_90" + _C + ".月 " 

def getSQL_CI_66_インシデント_アクシデント_月毎発生件数(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_90" + _C + ".病床数 ,  " + \
                "CI_90" + _C + ".年度 ,  " + \
	            "CI_90" + _C + ".月,  " + \
                "CI_90" + _C + ".月毎発生件数 " + \
            "FROM " + \
	            "CI_90" + _C + " " + \
            "WHERE " + \
	            "CI_90" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_90" + _C + ".年度, CI_90" + _C + ".月 " 

def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_66 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_66" in wb.sheetnames:
        wb.remove(wb["CI_66"])
    wb.create_sheet("CI_66")
    sheet = wb["CI_66"]    

    #インシデント_アクシデント_1か月_百床あたり_件数
    tempインシデント_アクシデント_1か月_百床あたり_件数 = X_01.excuteSQL(getSQL_CI_66_インシデント_アクシデント_1か月_百床あたり_件数(HOSPITAL_ID, _C))

    keyList = ["1ヶ月平均", "四月", "五月", "六月", "七月", "八月", "九月", "十月", 
                "十一月", "十二月", "一月", "二月", "三月" ]
    yearList = []
    dict_temp = {}

    if tempインシデント_アクシデント_1か月_百床あたり_件数 is None:
        print("エラー: インシデント_アクシデント_1か月_百床あたり_件数の取得に失敗しました。")
        return -1

    for a in tempインシデント_アクシデント_1か月_百床あたり_件数:
        # yearListに年が存在していなければ追加
        if a[1] not in yearList:
            yearList.append(a[1])
        #年と月をキーに辞書に追加
        dict_temp[str(a[1]) + str(a[2])] = a

    
    # 書込み列の繰り返し数
    roopColCnt = 1
    
    # 1行目に病院名
    sheet.cell(1, roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, roopColCnt).value = "インシデント_アクシデント_1か月_百床あたり_件数"
    # 3行目に疾患名
    sheet.cell(3, roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, roopColCnt).value = ""

    # 5行名からヘッダを出力
    tmpCnt = 0
    sheet.cell(5, roopColCnt).value = "病床数"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "年度"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "1ヶ月平均"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "四月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "五月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "六月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "七月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "八月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "九月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十一月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十二月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "一月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "二月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "三月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "1ヶ月平均_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "四月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "五月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "六月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "七月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "八月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "九月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十一月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十二月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "一月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "二月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "三月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "1ヶ月平均_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "四月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "五月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "六月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "七月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "八月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "九月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十一月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十二月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "一月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "二月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "三月_比較用"
    

    # 6行目からデータ入力
    rowCnt = 6
    for year in yearList:
        colCnt = 3
        for key in keyList:
            tempRow = dict_temp[str(year) + str(key)]
            #病床数は"1ヶ月平均"を出力
            if colCnt == 3:
                #病床数
                sheet.cell(rowCnt, roopColCnt).value = tempRow[0]
                #年度
                sheet.cell(rowCnt, roopColCnt + 1).value = tempRow[1]
            
            #グラフ作成用数値
            if tempRow[3] == -1 or tempRow[3] == -2:
                sheet.cell(rowCnt, colCnt).value = 0
            else:
                sheet.cell(rowCnt, colCnt).value = tempRow[3]

            #ラベル用数値
            if tempRow[3] == -1:
                sheet.cell(rowCnt, colCnt + 13).value = "N/A"
            elif tempRow[3] == -2:
                sheet.cell(rowCnt, colCnt + 13).value = "-"
            else:
                sheet.cell(rowCnt, colCnt + 13).value = tempRow[3]
            
            #比較用
            sheet.cell(rowCnt, colCnt + 26).value = tempRow[3]
            
            colCnt += 1
        rowCnt += 1

    
    #インシデント_アクシデント_月毎発生件数
    tempインシデント_アクシデント_月毎発生件数 = X_01.excuteSQL(getSQL_CI_66_インシデント_アクシデント_月毎発生件数(HOSPITAL_ID, _C))

    keyList = ["1ヶ月平均", "四月", "五月", "六月", "七月", "八月", "九月", "十月", 
                "十一月", "十二月", "一月", "二月", "三月" ]
    yearList = []
    dict_temp = {}

    if tempインシデント_アクシデント_月毎発生件数 is None:
        print("エラー: インシデント_アクシデント_月毎発生件数の取得に失敗しました。")
        return -1

    for a in tempインシデント_アクシデント_月毎発生件数:
        # yearListに年が存在していなければ追加
        if a[1] not in yearList:
            yearList.append(a[1])
        #年と月をキーに辞書に追加
        dict_temp[str(a[1]) + str(a[2])] = a

    
    
    # 書込み列の繰り返し数
    roopColCnt = tmpCnt + 3
    

    # 1行目に病院名
    sheet.cell(1, roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, roopColCnt).value = "インシデント_アクシデント_月毎発生件数"
    # 3行目に疾患名
    sheet.cell(3, roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, roopColCnt).value = ""

    # 5行名からヘッダを出力
    tmpCnt = 0
    sheet.cell(5, roopColCnt).value = "病床数"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "年度"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "1ヶ月平均"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "四月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "五月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "六月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "七月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "八月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "九月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十一月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十二月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "一月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "二月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "三月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "1ヶ月平均_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "四月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "五月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "六月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "七月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "八月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "九月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十一月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十二月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "一月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "二月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "三月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "1ヶ月平均_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "四月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "五月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "六月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "七月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "八月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "九月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十一月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十二月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "一月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "二月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "三月_比較用"


    # 6行目からデータ入力
    rowCnt = 6
    for year in yearList:
        colCnt = roopColCnt + 2
        for key in keyList:
            tempRow = dict_temp[str(year) + str(key)]
            #病床数は"1ヶ月平均"を出力
            if colCnt == roopColCnt + 2:
                #病床数
                sheet.cell(rowCnt, roopColCnt).value = tempRow[0]
                #年度
                sheet.cell(rowCnt, roopColCnt + 1).value = tempRow[1]
            
            #グラフ作成用数値
            if tempRow[3] == -1 or tempRow[3] == -2:
                sheet.cell(rowCnt, colCnt).value = 0
            else:
                sheet.cell(rowCnt, colCnt).value = tempRow[3]

            #ラベル用数値
            if tempRow[3] == -1:
                sheet.cell(rowCnt, colCnt + 13).value = "N/A"
            elif tempRow[3] == -2:
                sheet.cell(rowCnt, colCnt + 13).value = "-"
            else:
                sheet.cell(rowCnt, colCnt + 13).value = tempRow[3]
                
            #比較用
            sheet.cell(rowCnt, colCnt + 26).value = tempRow[3]
            
            colCnt += 1
        rowCnt += 1
    

    #wb.save(wbPath)
    print(" " + HOSPITAL_NAME + " CI_66 終了")

