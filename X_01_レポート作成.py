# 診療アウトカム評価事業　レポート作成　作成者　吉田　喬、作成日　2025/7/3
# 急性期指標レポート、アンケートレポート、慢性期指標レポートを作成する
# 各レポートは期間毎にまとめる。
# それぞれ病院名をそのまま表示させるレポート、病院名を番号に変換したレポートの２種類作成する
# 病院名を表示するレポートはNITが作成したPDFをそのまま使用、病院番号に変更する場合はエクセルファイルを使用する
# レポートの病院名を病院番号に変換、PDFに変換する
# 各指標の前に定義（PDF）を入れる
# ①急性期指標
#　　年度、グループ（急性期グループ、慢性期グループ）、期間（1Q、2Q、3Q、4Q、年間）
# ※注　令和6年度診療報酬改定により2024年度に指標の改修、追加を行った。
# 　　　2023年度まで、2024年度第1四半期、2024年度第2四半期以降で指標定義が異なるので注意。
# ②アンケート
# ③慢性期指標

import os
import sys
import shutil
import sqlite3
from datetime import datetime
from tkinter import N
import X_00_CONST as CONST
import X_02_急性期レポート作成 as ACUTE
import X_03_アンケート調査作成 as SURVEY
import X_04_慢性期指標レポート作成 as CHRONIC
import X_05_経年推移レポート as TRANSITION
import X_06_病院番号追記 as ADD_HOSPNO
import X_07_個別_急性期レポート作成 as INDIV_ACUTE
import X_08_個別_アンケート調査レポート作成 as INDIV_SURVEY
import X_09_個別_慢性期指標レポート作成 as INDIV_CHRONIC
import X_10_個別_死亡率_医療費レポート作成 as INDIV_DEATHRATE_COST
import X_11_個別レポート_グラフ作成 as INDIV_GRAPH
import X_12_個別レポート_PDF化 as INDIV_PDF
import openpyxl
from openpyxl.styles import PatternFill, Side, Border
import win32com.client

def exit_program():
    print("\n\n********診療アウトカム評価事業　レポート作成 終了********\n")
    sys.exit(0)

# DBにSQLを実行する関数（SELECT句を想定）
# 引数：dbPath：DBのパス、query：SQL文、params：SQL文のパラメータ（タプル）
def excuteSQL(query, params=None):
    try:
        # パスを正しく処理
        db_path = CONST.DB_PATH.replace('\\', '/')
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        #print(f"  実行するクエリ: {query}")
        if params:
            #print(f"  パラメータ: {params}")
            cursor.execute(query, params)
        else:
            cursor.execute(query)
        result = cursor.fetchall()
        #print(f"  実行結果 件数: {len(result)}")
        return result
    except sqlite3.Error as e:
        print(f"DBエラーが発生しました：{e}")
        print(f"  クエリ: {query}")
        print(f"  DBパス: {CONST.DB_PATH}")
    finally:
        if conn:
            conn.close()


# テーブル「HOSPITAL_PUBLIC_NO」に年度、PUBLIC_NO、HOSPITAL_NAMEを新規追加する
# 引数：hospital_data：病院データ（タプル）、target_year：年度
# 戻り値：0:成功、-1:失敗
def insertPublicNo(hospital_data, target_year):
    try:
        conn = sqlite3.connect(CONST.DB_PATH)
        cursor = conn.cursor()

        for hospital_id, hospital_name in hospital_data:
            public_no = f"{target_year}{hospital_id:04d}"  # PUBLIC_NOは年度 + 病院ID（4桁ゼロ埋め）
            query_insert = """
                INSERT INTO HOSPITAL_PUBLIC_NO (年度, PUBLIC_NO, HOSPITAL_NAME)
                VALUES (?, ?, ?)
            """
            cursor.execute(query_insert, (target_year, public_no, hospital_name))
            conn.commit()
            

        print("  データ追加が完了しました。")
        print(f"  追加件数：{len(hospital_data)}")
        return 0
    except sqlite3.Error as e:
        print(f"  データ追加中にエラーが発生しました：{e}")
        return -1
    finally:
        if conn:
            conn.close()
        return 0


#当該年度の平均在院日数（疾患別）から1Qのデータ提出病院のHospital_HOSPITAL_IDを取得し、
#HOSPITALテーブルから病院名を取得する
#引数：dbPath：DBのパス、target_year：対象年度
#戻り値：病院名とHOSPITAL_IDのリスト
#ERROR:-1
def getHospIDAndName(target_year):
    
    query_急性期 = "SELECT DISTINCT Hospital_HOSPITAL_ID FROM CI_01 WHERE 期 = '1Q' AND 年度 = ?"
    query_慢性期 = "SELECT DISTINCT Hospital_HOSPITAL_ID FROM CI_01_C WHERE 期 = '1Q' AND 年度 = ?"

    hospIDList_急性期 = excuteSQL(query_急性期, (target_year,))
    if hospIDList_急性期 is None:
        return -1
    #print("\n 急性期の病院IDリスト取得完了（参加病院全体（平均値、中央値）２件含む）")
    #print(" 取得件数：" + str(len(hospIDList_急性期)))
    #print(" 病院IDリスト：" + str(hospIDList_急性期))
    hospIDList_慢性期 = excuteSQL(query_慢性期, (target_year,))
    if hospIDList_慢性期 is None:
        return -1
    #print("\n 慢性期の病院IDリスト取得完了（参加病院全体（平均値、中央値）２件含む）")
    #print(" 取得件数：" + str(len(hospIDList_慢性期)))
    #print(" 病院IDリスト：" + str(hospIDList_慢性期))

    # hospIDList_急性期とhospIDList_慢性期から重複を除いた統合リストを作成
    hospIDList_統合 = list(set(hospIDList_急性期) | set(hospIDList_慢性期))
    #print("\n 統合病院IDリスト作成完了（参加病院全体（平均値、中央値）２件含む）")
    #print(" 統合リスト件数：" + str(len(hospIDList_統合)))
    #print(" 統合病院IDリスト：" + str(hospIDList_統合))

    # テーブル「HOSPITAL」から病院名とHOSPITAL_IDを取得する
    query_hospital = "SELECT HOSPITAL_ID, 病院名 FROM HOSPITAL WHERE HOSPITAL_ID IN ({seq})".format(
        seq=','.join(['?'] * len(hospIDList_統合))
    )

    hospIDList_統合_flat = [item[0] for item in hospIDList_統合]  # hospIDList_統合のID部分を抽出
    hospital_data = excuteSQL(query_hospital, hospIDList_統合_flat)
    if hospital_data is None or len(hospital_data) == 0 :
        return -1

    #print("\n 病院名とHOSPITAL_IDの取得完了（参加病院全体（平均値、中央値）２件含む）")
    #print(" 取得件数：" + str(len(hospital_data)))
    #print(" 病院データ：" + str(hospital_data))
    return hospital_data


def create_hospitalList(target_year, listPath, output_folder):
    print("\n参加病院名簿の作成を開始します。")
    #①名簿ファイルの読み取り    
    hospital_list = []
    try:
        #listPathのファイルを開く
        workbook = openpyxl.load_workbook(listPath)
        if workbook is None:
            print("\nエラー: 病院一覧ファイルの読み込みに失敗しました。年度={}".format(target_year))
            print("ファイルパス：{}".format(listPath))
            return -1
        
        #シート名一覧を取得
        sheet_names = workbook.sheetnames
        if sheet_names is None:
            print("\nエラー: 病院一覧ファイルのシート名が取得できません。年度={}".format(target_year))
            print("ファイルパス：{}".format(listPath))
            return -1

        sheet = None
        #シート名にtarget_yearが含まれているシートを開く
        for sheet_name in sheet_names:
            if str(target_year) in sheet_name:
                sheet = workbook[sheet_name]
                #print("\nsheet_name={}".format(sheet_name))
                break

        if sheet is None:
            print("\nエラー: 病院一覧ファイルのシートが見つかりません。年度={}".format(target_year))
            print("ファイルパス：{}".format(listPath))
            return -1

        #ヘッダ：1行目のデータを取得
        header_row = sheet[1]
        if header_row is None:
            print("\nエラー: 病院一覧ファイルのヘッダが見つかりません。年度={}".format(target_year))
            print("ファイルパス：{}".format(listPath))
            return -1

        #for i in range(len(header_row)):
        #    print("header_row[{}].value={}".format(i,header_row[i].value))


        header_index = {}
        cnt = 0
        for i in range(len(header_row)):
            #print("header_row[{}].value={}".format(i,header_row[i].value))
            if header_row[i].value == "No.":
                header_index["No."] = i
                cnt += 1
            if header_row[i].value == "都道府県":
                header_index["都道府県"] = i
                cnt += 1
            elif header_row[i].value == "施設名":
                header_index["施設名"] = i
                cnt += 1
            elif header_row[i].value == "総合計":
                header_index["総合計"] = i
                cnt += 1
            if cnt == 4:
                break

        #print("header_index={}".format(header_index))   

        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, values_only=True):
            if row[header_index["No."]] == "" or row[header_index["No."]] is None:
                break
            hosipital_data = []
            
            #print("No." + str(row[1]))
            #print("施設名=" + str(row[8]))
            #print("都道府県=" + str(row[15]))
            #print("総合計=" + str(row[19]))


            hosipital_data.append(row[header_index["都道府県"]])
            hosipital_data.append(row[header_index["施設名"]])
            hosipital_data.append(row[header_index["総合計"]])
            hospital_list.append(hosipital_data)
            #print("hosipital_data={}".format(hosipital_data))

        #hospital_listを都道府県順で並び替え
        hospital_list.sort(key=lambda x: CONST.pref_order[x[0]])    

        #for hospital in hospital_list:
        #    print(hospital)

        workbook.close()

        #print(hospital_list)
        print("\n名簿読み取り終了")
        print("参加病院数={}".format(len(hospital_list)))

    except:
        print("\nエラー: 病院一覧ファイルの作成に失敗しました。年度={}".format(target_year))
        print("ファイルパス：{}".format(listPath))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1

    #②00_表紙、病院一覧.xlsxに書き込み
    try:
        #00_表紙.xlsxを開く
        workbook = openpyxl.load_workbook(CONST.HOSP_LIST_FORMAT_PATH)
        if workbook is None:
            print("\nエラー: 表紙ファイルの読み込みに失敗しました。年度={}".format(target_year))
            print("ファイルパス：{}".format(os.path.join(output_folder, "00_表紙.xlsx")))
            return -1

        sheet = workbook["00-2_参加病院名簿"]
        if sheet is None:
            print("\nエラー: 参加病院名簿シートが見つかりません。年度={}".format(target_year))
            print("ファイルパス：{}".format(os.path.join(output_folder, "00_表紙.xlsx")))
            return -1
        
        #セルG1に指標タイプを入力
        sheet["G1"] = target_year
        idx = 5
        #参加病院名簿シートに書き込み
        for hospital in hospital_list:
            #5行目から書き込み
            
            sheet.cell(row=idx, column=1).value = idx - 4
            sheet.cell(row=idx, column=2).value = hospital[0]
            sheet.cell(row=idx, column=3).value = hospital[1]
            sheet.cell(row=idx, column=4).value = hospital[2]
            idx += 1
        
        # 線のスタイルを定義（細線）
        thin = Side(border_style="thin", color="000000")
        # 4辺に罫線を設定
        border = Border(top=thin, bottom=thin, left=thin, right=thin)

        for row in sheet["A4:D" + str(idx -1 )]:
            for cell in row:
                cell.border = border


        #上書き保存 DEBG
        #workbook.save(CONST.OVER_PAGE_PATH)
        workbook.save(os.path.join(output_folder, "00-2_参加病院名簿.xlsx"))
        if workbook is None:
            print("\nエラー: 参加病院名簿シートの保存に失敗しました。年度={}".format(target_year))


        #PDFに変換
        # Excel ファイルを開く
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Excel のウィンドウを非表示にする
        wb = excel.Workbooks.Open(os.path.join(output_folder, "00-2_参加病院名簿.xlsx"))
        newPath = str(os.path.join(output_folder, "00-2_参加病院名簿.pdf"))
        #print("newPath={}".format(newPath))

        wb.ExportAsFixedFormat(0, newPath)  # 0 は PDF 形式
        wb.Close()
        excel.Quit()

        #pdfファイルの作成が完了したら元のエクセルファイルを削除
        os.remove(os.path.join(output_folder, "00-2_参加病院名簿.xlsx"))

    except:
        print("\nエラー: 参加病院名簿シートの作成に失敗しました。年度={}".format(target_year))
        print("ファイルパス：{}".format(os.path.join(output_folder, "00_表紙.xlsx")))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1


def createDictPublicNo(target_year, hospital_data):
    # テーブル「HOSPITAL_PUBLIC_NO」に当該年度のデータがあるか確認する
    query_public_no = "SELECT PUBLIC_NO, HOSPITAL_NAME FROM HOSPITAL_PUBLIC_NO WHERE 年度 = ?"
    public_no_data = excuteSQL(query_public_no, (target_year,))


    if public_no_data:
        print("\n テーブル「HOSPITAL_PUBLIC_NO」に当該年度のデータが存在します。")
        print(" 取得件数：" + str(len(public_no_data)))

        # 病院名をキーにハッシュテーブルに公開用病院番号を格納
        dict_hosp_DB = {}
        for row in public_no_data:
            dict_hosp_DB[row[1]] = row[0]
            #print("dict_hosp_DB: key={}, value={}".format(row[1],row[0]))
        return dict_hosp_DB
    else:
        print("\n テーブル「HOSPITAL_PUBLIC_NO」に当該年度のデータが存在しません。")
        print(">>" + str(target_year) + "年度データに公開用病院番号を新規登録しますか？")
        print(" 1:新規登録する、2:中止する")
        userInput = input('>>')

        if userInput == "1":
            print("新規登録を開始します。")
            # テーブル「HOSPITAL_PUBLIC_NO」に年度、PUBLIC_NO、HOSPITAL_NAMEを新規追加する
            if(not insertPublicNo(hospital_data, target_year)):
                return -1
        
            public_no_data = excuteSQL(query_public_no, (target_year,))
            if public_no_data: 
                print(" 取得件数：" + str(len(public_no_data)))

                # 病院名をキーにハッシュテーブルに公開用病院番号を格納
                dict_hosp_DB = {}
                for row in public_no_data:
                    dict_hosp_DB[row[1]] = row[0]
                return dict_hosp_DB

            else:
                print(" 新規登録に失敗しました")
                return -1

        elif userInput == "2":
            print("中止します。")
            return -1
        else:
            print("不正な入力です。")
            return -1

    
    #print("\n 公開用病院番号データを出力します。")
    #for record in public_no_data:
    #    print(record)

#
# エクセルのファイル名をキーに病院名行、読み取り開始列を格納した辞書型
# key:エクセルのファイル名
# 0  :病院名行
# 1  :読み取り開始列
#
def getDict_HospRow_StartCol():
    tempDict = {
		"平均在院日数_疾患別" : (4,5),
		"平均在院日数_重症度別" : (4,6),
		"平均在院日数_年代別" : (4,6),
		"平均在院日数_性別" : (4,6),
        "死亡率_重症度別" :(4, 6),
        "死亡率_年代別" :(4, 6),
        "死亡率_性別" : (4, 6),
		"予定しない再入院率（30日以内再入院）_疾患別" : (4,4),
		"予定しない再入院率（30日以内再入院）_重症度別" : (4,5),
		"予定しない再入院率（30日以内再入院）_年代別" : (4,5),
		"予定しない再入院率（30日以内再入院）_性別" : (4,5),
        "医療費_重症度別":(4, 6),
        "医療費_年代別":(4, 6),
        "医療費_性別":(4, 6),
        "リスクレベルが「中」以上の手術を施行した患者の肺血栓塞栓症の予防対策の実施率" : (4,4),
		"肺血栓塞栓症の予防対策" : (4,4),
        "手術ありの患者の肺血栓塞栓症（予防対策の実施率）" : (4,4),
		"肺血栓塞栓症の発生率_管理料算定有無別" : (4,5),
		"褥瘡の持込率" : (4,5),
		"褥瘡の発生率" : (4,5),
		"d2（真皮までの損傷）以上の褥瘡発生率" : (4,5),
		"65歳以上の患者における認知症の保有率" : (4,5),
		"急性心筋梗塞における主要手術・治療実施率_重症度別" : (4,5),
		"急性心筋梗塞における主要手術・治療実施率_年代別" : (4,5),
		"急性心筋梗塞における主要手術・治療実施率_性別" : (4,5),
		"急性心筋梗塞における主要手術・治療実施率_Kコード別" : (4,5),
		"肺炎に対する入院当日の抗生物質使用率_重症度別" : (4,5),
		"出血性胃・十二指腸潰瘍内視鏡的治療の施行率" : (4,4),
		"急性脳梗塞患者リハビリテーション開始率" : (4,4),
		"胃がんの患者に対するESDの施行率" : (4,4),
		"胆嚢切除術患者に対する腹腔鏡下手術施行率" : (4,4),
		"虫垂炎の患者に対する手術施行率" : (4,4),
		"虫垂切除術患者に対する腹腔鏡下手術施行率" : (4,4),
		"帝王切開における全身麻酔施行率" : (4,4),
		"帝王切開における輸血施行率" : (4,4),
		"脳卒中地域連携パスの使用率" : (4,4),
		"脳卒中地域連携パスの使用率(転院先)" : (4,4),
		"大腿骨地域連携パスの使用率" : (4,4),
		"大腿骨頸部骨折連携パスの使用率(転院先)" : (4,4),
		"急性期病棟における退院調整の実施率" : (4,4),
		"介護施設・福祉施設からの入院率" : (4,5),
		"他の病院・診療所の病棟からの転院率" : (4,5),
		"介護施設・福祉施設等への退院率(年代別)" : (4,5),
		"介護施設・福祉施設等への退院率(退院先)" : (4,5),
		"他の病院・診療所の病棟への転院率" : (4,5),
		"自宅退院患者における在宅医療を受ける率" : (4,5),
		"中心静脈カテーテル挿入時の気胸発生率" : (4,4),
		"急性心筋梗塞患者におけるアスピリン_早期投与" : (4,4),
		"急性心筋梗塞患者におけるアスピリン_退院時投与" : (4,4),
		"Door_to_balloon" : (4,4),
		"誤嚥性肺炎患者に対する喉頭ファイバースコピーあるいは嚥下造影検査の実施率" : (4,4),
		"術後24時間以内の予防的抗菌薬の投与停止率" : (4,4),
		"術後48時間以内の予防的抗菌薬の投与停止率" : (4,4),
		"服薬指導（安全管理が必要な薬剤の服薬指導実施率）" : (4,4),
		"服薬指導（薬剤管理指導実施率）" : (4,4),
		"栄養指導" : (4,4),
		"血液培養の実施（血液培養実施率）" : (4,4),
		"広域スペクトル抗菌薬使用時の細菌培養実施率" : (4,4),
		"血液培養の実施（2セット実施率）" : (4,4),
        "血液培養の実施（2セット培養）" : (4,4),
        "血液培養2セット実施率" : (4,4),
		"65 歳以上の患者の入院早期の栄養アセスメント実施割合" : (4,4),
		"身体的拘束の実施率" : (4,4),
		"手術開始前1 時間以内の予防的抗菌薬投与率" : (4,4),
        "転倒・転落発生率" : (4,4),
        "転棟転落によるインシデント影響度分類レベル3b以上の発生率" : (4,4),
    }
    return tempDict



def getSQL_疾患マスタ():
    return "SELECT * FROM MASTA_DISEASE ORDER BY ORDER_NUM"

def getSQL_重症度マスタ():
    return "SELECT * FROM MASTA_SEVERITY ORDER BY ORDER_NUM"

def getSQL_年代マスタ():
    return "SELECT * FROM MASTA_AGE ORDER BY ORDER_NUM"

def getSQL_性別マスタ():
    return "SELECT * FROM MASTA_SEX ORDER BY ORDER_NUM"

def getSQL_管理料算定有無マスタ():
    return "SELECT * FROM MASTA_CALCULATION ORDER BY ORDER_NUM"

def getSQL_入院経路マスタ():
    return "SELECT * FROM MASTA_SOURCE ORDER BY ORDER_NUM"

def getSQL_Kコードマスタ():
    return "SELECT * FROM MASTA_K_CODE ORDER BY ORDER_NUM"

def getSQL_退院先マスタ():
    return "SELECT * FROM MASTA_DESTINATION ORDER BY ORDER_NUM"

def setMastaData():
    print("マスタデータをセットします。")
    # 疾患マスタ
    temp_DISEASE = excuteSQL(getSQL_疾患マスタ(), None)
    if temp_DISEASE is None:
        print("エラー: 疾患マスタの取得に失敗しました。")
        return -1
    temp_DISEASE_list = []
    for row in temp_DISEASE:
        temp_DISEASE_list.append(row)
    CONST.MASTA_DISEASE = temp_DISEASE_list
    
    # 重症度マスタ
    temp_SEVERITY = excuteSQL(getSQL_重症度マスタ(), None)
    if temp_SEVERITY is None:
        print("エラー: 重症度マスタの取得に失敗しました。")
        return -1
    temp_SEVERITY_list = []
    for row in temp_SEVERITY:
        temp_SEVERITY_list.append(row)
    CONST.MASTA_SEVERITY = temp_SEVERITY_list
    
    # 年代マスタ
    temp_AGE = excuteSQL(getSQL_年代マスタ(), None)
    if temp_AGE is None:
        print("エラー: 年代マスタの取得に失敗しました。")
        return -1
    temp_AGE_list = []
    for row in temp_AGE:
        temp_AGE_list.append(row)
    CONST.MASTA_AGE = temp_AGE_list
    
    # 性別マスタ
    temp_SEX = excuteSQL(getSQL_性別マスタ(), None)
    if temp_SEX is None:
        print("エラー: 性別マスタの取得に失敗しました。")
        return -1
    temp_SEX_list = []
    for row in temp_SEX:
        temp_SEX_list.append(row)
    CONST.MASTA_SEX = temp_SEX_list
    
    # 管理料算定有無マスタ
    temp_CALCULATION = excuteSQL(getSQL_管理料算定有無マスタ(), None)
    if temp_CALCULATION is None:
        print("エラー: 管理料算定有無マスタの取得に失敗しました。")
        return -1
    temp_CALCULATION_list = []
    for row in temp_CALCULATION:
        temp_CALCULATION_list.append(row)
    CONST.MASTA_CALCULATION = temp_CALCULATION_list
    
    # 入院経路マスタ
    temp_SOURCE = excuteSQL(getSQL_入院経路マスタ(), None)
    if temp_SOURCE is None:
        print("エラー: 入院経路マスタの取得に失敗しました。")
        return -1
    temp_SOURCE_list = []
    for row in temp_SOURCE:
        temp_SOURCE_list.append(row)
    CONST.MASTA_SOURCE = temp_SOURCE_list
    
    # Kコードマスタ
    temp_K_CODE = excuteSQL(getSQL_Kコードマスタ(), None)
    if temp_K_CODE is None:
        print("エラー: Kコードマスタの取得に失敗しました。")
        return -1
    temp_K_CODE_list = []
    for row in temp_K_CODE:
        temp_K_CODE_list.append(row)
    CONST.MASTA_K_CODE = temp_K_CODE_list
    
    # 退院先マスタ  
    temp_DESTINATION = excuteSQL(getSQL_退院先マスタ(), None)
    if temp_DESTINATION is None:
        print("エラー: 退院先マスタの取得に失敗しました。")
        return -1
    temp_DESTINATION_list = []
    for row in temp_DESTINATION:    
        temp_DESTINATION_list.append(row)
    CONST.MASTA_DESTINATION = temp_DESTINATION_list
    
    print("マスタデータのセットに成功しました。")
    return 0    















if __name__ == "__main__":
    print("\n\n********診療アウトカム評価事業　レポート作成 開始********")
    #DB接続 
    print("\nDB接続を開始します。")
    conn = sqlite3.connect(CONST.DB_PATH)
    cursor = conn.cursor()
    if cursor is None:
        print("DB接続に失敗しました。\n")
        sys.exit(0)
    else:
        print("DB接続に成功しました。\n")

    print("作成レポートを選んでください。")
    print("四半期、期間レポート" + "\n"      +
          "  １：急性期指標レポート、２：アンケート調査レポート、３：慢性期指標レポート、４：経年推移レポート、５：参加病院一覧に病院名変換後の病院番号を追記。" +"\n" +
          "個別レポート  作成" + "\n" +
          "  ６：急性期指標レポート、７：アンケート調査レポート、８：慢性期指標レポート、９：死亡率・医療費レポート、" + "\n" +
          "個別レポート  後処理" + "\n" +
          "  １０：個別レポートのグラフ作成、１１：個別レポートのPDF化")
    

    user_input = input(":>>")
    if user_input not in ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11" ]:
        print("エラー: 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11（半角） のいずれかを入力してください。")
        sys.exit(0)

    # 急性期指標レポート
    if user_input == "1":
        if ACUTE.create_report() == -1:
            exit_program()

    # アンケート調査レポート
    if user_input == "2":
        if SURVEY.create_report() == -1:
            exit_program()

    # 慢性期指標レポート
    if user_input == "3":
        if CHRONIC.create_report() == -1:
            exit_program()

    # 経年推移レポート
    if user_input == "4":
        if TRANSITION.create_report() == -1:
            exit_program()
    
    # 参加病院一覧に病院名変換後の病院番号を追記
    if user_input == "5":
        if ADD_HOSPNO.add_hospital_number() == -1:
            exit_program()
    
    # 個別レポート（急性期指標）
    if user_input == "6":
        if INDIV_ACUTE.create_report() == -1:
            exit_program()

    # 個別レポート（アンケート調査）
    if user_input == "7":
        if INDIV_SURVEY.create_report() == -1:
            exit_program()

    # 個別レポート（慢性期指標）
    if user_input == "8":
        if INDIV_CHRONIC.create_report() == -1:
            exit_program()

    # 個別レポート（死亡率・医療費）    
    if user_input == "9":
        if INDIV_DEATHRATE_COST.create_report() == -1:
            exit_program()

    # 個別レポート（グラフ作成）
    if user_input == "10":
        if INDIV_GRAPH.create_report() == -1:
            exit_program()

    # 個別レポート（PDF化）
    if user_input == "11":
        if INDIV_PDF.create_report_pypdf() == -1:
            exit_program()


    exit_program()





