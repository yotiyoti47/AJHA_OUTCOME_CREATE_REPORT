import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01
# 2024年度にCI番号変更
# CI_92　→　CI_59_2
# ただし、DBのテーブルは「CI_92」のまま
def getSQL_CI_59_2_転倒_転落_月別_転倒_転落件数(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_92" + _C + ".年度 ,  " + \
                "CI_92" + _C + ".月,  " + \
                "CI_92" + _C + ".月別_転倒_転落件数 " + \
            "FROM " + \
	            "CI_92" + _C + " " + \
            "WHERE " + \
	            "CI_92" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_92" + _C + ".年度, CI_92" + _C + ".月 " 

def getSQL_CI_59_2_転倒_転落_入院中の患者に発生した転倒_転落件数(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_92" + _C + ".年度 ,  " + \
                "CI_92" + _C + ".月,  " + \
                "CI_92" + _C + ".入院中の患者に発生した転倒_転落件数 " + \
            "FROM " + \
	            "CI_92" + _C + " " + \
            "WHERE " + \
	            "CI_92" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_92" + _C + ".年度, CI_92" + _C + ".月 " 

def getSQL_CI_59_2_転倒_転落_入院患者延べ数(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_92" + _C + ".年度 ,  " + \
                "CI_92" + _C + ".月,  " + \
                "CI_92" + _C + ".入院患者延べ数 " + \
            "FROM " + \
	            "CI_92" + _C + " " + \
            "WHERE " + \
	            "CI_92" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_92" + _C + ".年度, CI_92" + _C + ".月 " 


def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_59_2 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_59_2" in wb.sheetnames:
        wb.remove(wb["CI_59_2"])
    wb.create_sheet("CI_59_2")
    sheet = wb["CI_59_2"]    

    #転倒_転落_月別_転倒_転落件数
    temp転倒_転落_月別_転倒_転落件数 = X_01.excuteSQL(getSQL_CI_59_2_転倒_転落_月別_転倒_転落件数(HOSPITAL_ID, _C))

    keyList = ["1ヶ月平均", "四月", "五月", "六月", "七月", "八月", "九月", "十月", 
                "十一月", "十二月", "一月", "二月", "三月" ]
    yearList = []
    dict_temp = {}

    if temp転倒_転落_月別_転倒_転落件数 is None:
        print("エラー: 転倒_転落_月別_転倒_転落件数の取得に失敗しました。")
        return -1

    for a in temp転倒_転落_月別_転倒_転落件数:
        # yearListに年が存在していなければ追加
        if a[0] not in yearList:
            yearList.append(a[0])
        #年と月をキーに辞書に追加
        dict_temp[str(a[0]) + str(a[1])] = a

    
    # 書込み列の繰り返し数
    roopColCnt = 1
    
    # 1行目に病院名
    sheet.cell(1, roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, roopColCnt).value = "転倒_転落件数"
    # 3行目に疾患名
    sheet.cell(3, roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, roopColCnt).value = ""

    # 5行名からヘッダを出力
    sheet.cell(5, roopColCnt).value = "年度"
    tmpCnt = 1

    #ヘッダ　グラフ作成用
    for a in keyList:
        sheet.cell(5, roopColCnt + tmpCnt).value = a
        tmpCnt+=1
    
    #ヘッダ　ラベル用
    for a in keyList:
        sheet.cell(5, roopColCnt + tmpCnt).value = a + "_ラベル"
        tmpCnt+=1
    
    #ヘッダ　比較用
    for a in keyList:
        sheet.cell(5, roopColCnt + tmpCnt).value = a + "_比較用"
        tmpCnt+=1

    # 6行目からデータ入力
    rowCnt = 6
    for year in yearList:
        colCnt = 2
        for key in keyList:
            tempRow = dict_temp[str(year) + str(key)]
            #年度は"1ヶ月平均"を出力
            if colCnt == 2:
                sheet.cell(rowCnt, roopColCnt).value = tempRow[0]

            #グラフ作成用数値
            if tempRow[2] == -1 or tempRow[2] == -2:
                sheet.cell(rowCnt, colCnt).value = 0
            else:
                sheet.cell(rowCnt, colCnt).value = tempRow[2]

            #ラベル用数値
            if tempRow[2] == -1:
                sheet.cell(rowCnt, colCnt + 13).value = "N/A"
            elif tempRow[2] == -2:
                sheet.cell(rowCnt, colCnt + 13).value = "-"
            else:
                sheet.cell(rowCnt, colCnt + 13).value = tempRow[2]
            
            #比較用
            sheet.cell(rowCnt, colCnt + 26).value = tempRow[2]
            
            colCnt += 1
        rowCnt += 1

    #転倒_転落_入院中の患者に発生した転倒_転落件数
    temp転倒_転落_入院中の患者に発生した転倒_転落件数 = X_01.excuteSQL(getSQL_CI_59_2_転倒_転落_入院中の患者に発生した転倒_転落件数(HOSPITAL_ID, _C))

    keyList = ["1ヶ月平均", "四月", "五月", "六月", "七月", "八月", "九月", "十月", 
                "十一月", "十二月", "一月", "二月", "三月" ]
    yearList = []
    dict_temp = {}

    if temp転倒_転落_入院中の患者に発生した転倒_転落件数 is None:
        print("エラー: 転倒_転落_入院中の患者に発生した転倒_転落件数の取得に失敗しました。")
        return -1

    for a in temp転倒_転落_入院中の患者に発生した転倒_転落件数:
        # yearListに年が存在していなければ追加
        if a[0] not in yearList:
            yearList.append(a[0])
        #年と月をキーに辞書に追加
        dict_temp[str(a[0]) + str(a[1])] = a

            
    # 書込み列の繰り返し数
    roopColCnt = tmpCnt + 2
    
    # 1行目に病院名
    sheet.cell(1, roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, roopColCnt).value = "入院中の患者に発生した転倒_転落件数"
    # 3行目に疾患名
    sheet.cell(3, roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, roopColCnt).value = ""

    # 5行名からヘッダを出力
    sheet.cell(5, roopColCnt).value = "年度"
    tmpCnt = 1

    #ヘッダ　グラフ作成用
    for a in keyList:
        sheet.cell(5, roopColCnt + tmpCnt).value = a
        tmpCnt+=1
    
    #ヘッダ　ラベル用
    for a in keyList:
        sheet.cell(5, roopColCnt + tmpCnt).value = a + "_ラベル"
        tmpCnt+=1
    
    #ヘッダ　比較用
    for a in keyList:
        sheet.cell(5, roopColCnt + tmpCnt).value = a + "_比較用"
        tmpCnt+=1


    # 6行目からデータ入力
    rowCnt = 6
    for year in yearList:
        colCnt = roopColCnt + 1
        for key in keyList:
            tempRow = dict_temp[str(year) + str(key)]
            #病床数は"1ヶ月平均"を出力
            if colCnt == roopColCnt + 1:
                sheet.cell(rowCnt, roopColCnt).value = tempRow[0]

             #グラフ作成用数値
            if tempRow[2] == -1 or tempRow[2] == -2:
                sheet.cell(rowCnt, colCnt).value = 0
            else:
                sheet.cell(rowCnt, colCnt).value = tempRow[2]

            #ラベル用数値
            if tempRow[2] == -1:
                sheet.cell(rowCnt, colCnt + 13).value = "N/A"
            elif tempRow[2] == -2:
                sheet.cell(rowCnt, colCnt + 13).value = "-"
            else:
                sheet.cell(rowCnt, colCnt + 13).value = tempRow[2]
            
            #比較用
            sheet.cell(rowCnt, colCnt + 26).value = tempRow[2]
            colCnt += 1
        rowCnt += 1
    

 #入院患者延べ数
    temp入院患者延べ数 = X_01.excuteSQL(getSQL_CI_59_2_転倒_転落_入院患者延べ数(HOSPITAL_ID, _C))

    keyList = ["1ヶ月平均", "四月", "五月", "六月", "七月", "八月", "九月", "十月", 
                "十一月", "十二月", "一月", "二月", "三月" ]
    yearList = []
    dict_temp = {}

    if temp入院患者延べ数 is None:
        print("エラー: 入院患者延べ数の取得に失敗しました。")
        return -1

    for a in temp入院患者延べ数:
        # yearListに年が存在していなければ追加
        if a[0] not in yearList:
            yearList.append(a[0])
        #年と月をキーに辞書に追加
        dict_temp[str(a[0]) + str(a[1])] = a

                    
    # 書込み列の繰り返し数
    roopColCnt = roopColCnt + tmpCnt + 1

    # 1行目に病院名
    sheet.cell(1, roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, roopColCnt).value = "入院患者延べ数"
    # 3行目に疾患名
    sheet.cell(3, roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, roopColCnt).value = ""

    # 5行名からヘッダを出力
    sheet.cell(5, roopColCnt).value = "年度"
    tmpCnt = 1

    #ヘッダ　グラフ作成用
    for a in keyList:
        sheet.cell(5, roopColCnt + tmpCnt).value = a
        tmpCnt+=1
    
    #ヘッダ　ラベル用
    for a in keyList:
        sheet.cell(5, roopColCnt + tmpCnt).value = a + "_ラベル"
        tmpCnt+=1
    
    #ヘッダ　比較用
    for a in keyList:
        sheet.cell(5, roopColCnt + tmpCnt).value = a + "_比較用"
        tmpCnt+=1


    # 6行目からデータ入力
    rowCnt = 6
    for year in yearList:
        colCnt = roopColCnt + 1
        for key in keyList:
            tempRow = dict_temp[str(year) + str(key)]
            #病床数は"1ヶ月平均"を出力
            if colCnt == roopColCnt + 1:
                sheet.cell(rowCnt, roopColCnt).value = tempRow[0]

             #グラフ作成用数値
            if tempRow[2] == -1 or tempRow[2] == -2:
                sheet.cell(rowCnt, colCnt).value = 0
            else:
                sheet.cell(rowCnt, colCnt).value = tempRow[2]

            #ラベル用数値
            if tempRow[2] == -1:
                sheet.cell(rowCnt, colCnt + 13).value = "N/A"
            elif tempRow[2] == -2:
                sheet.cell(rowCnt, colCnt + 13).value = "-"
            else:
                sheet.cell(rowCnt, colCnt + 13).value = tempRow[2]
            
            #比較用
            sheet.cell(rowCnt, colCnt + 26).value = tempRow[2]
            colCnt += 1
        rowCnt += 1

    #wb.save(wbPath)
    print(" " + HOSPITAL_NAME + " CI_59_2 終了")

