import sqlite3
#from typing_extensions import TypeVarTuple
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

def getSQL_疾患別患者数(HOSPITAL_ID, _C):
    return "SELECT * FROM NUMBER_OF_PATIENTS_BY_DISEASE" + _C + " WHERE Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " "  \
            "ORDER BY 年度, 期, Disease_DISEASE_ID"

def getSQL_CI_02_平均在院日数(HOSPITAL_ID, disease_ID, Severity_ID, _C):
    return  "SELECT  " + \
	            "CI_02" + _C + ".年度 ,  " + \
	            "CI_02" + _C + ".期,  " + \
                "CI_02" + _C + ".平均在院日数 " + \
            "FROM " + \
	            "CI_02" + _C + " " + \
            "WHERE " + \
	            "NOT CI_02" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_02" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " AND " + \
	            "CI_02" + _C + ".Disease_DISEASE_ID = " + str(disease_ID) + " AND " + \
                "CI_02" + _C + ".Severity_SEVERITY_ID = " + str(Severity_ID) +  \
            " ORDER BY CI_02" + _C + ".年度, CI_02" + _C + ".期 " 

def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_02 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_02" in wb.sheetnames:
        ws = wb["CI_02"]
        for row in ws.rows:
            for cell in row:
                cell.value = None
        
        #wb.remove(wb["CI_02"])

    wb.create_sheet("CI_02")
    sheet = wb["CI_02"]    

    #当該病院の疾患別患者数を取得
    list_疾患別患者数 = X_01.excuteSQL(getSQL_疾患別患者数(HOSPITAL_ID, _C))
    if list_疾患別患者数 is None:
        print("エラー: 疾患別患者数の取得に失敗しました。")
        return -1

    dict_疾患別患者数 = {}

    #疾患別患者数を年度、期、DISEASE_IDをキーに辞書に追加
    for temp疾患別患者数 in list_疾患別患者数:
        # 年度、期、DISEASE_IDがキー
        key = str(temp疾患別患者数[1]) + temp疾患別患者数[2] + str(temp疾患別患者数[4])
        #print("key:" + key + " value:" + str(temp疾患別患者数[3])) 
        dict_疾患別患者数[key] = temp疾患別患者数[3]

    # 書込み列の位置
    colCnt = 1

    # 書込み列の繰り返し数
    roopColCnt = 9

    #疾患の数だけループ
    for tempDisease in CONST.MASTA_DISEASE:
        
        #対象疾患の重症度を取得
        sevList = []
        for tempSev in CONST.MASTA_SEVERITY:
            if tempSev[3] == tempDisease[0]:
                sevList.append(tempSev)

        #対象疾患の重症度の数だけループ
        for tempSev in sevList:

            #当該疾患の平均在院日数を取得
            #print(HOSPITAL_NAME + " " + tempDisease[1] + " " + tempSev[1])
            #print(HOSPITAL_NAME + " " + tempDisease[1] + " " + tempSev[1])
            #print(getSQL_CI_02_平均在院日数(HOSPITAL_ID, tempDisease[0], tempSev[0]))
            temp平均在院日数 = X_01.excuteSQL(getSQL_CI_02_平均在院日数(HOSPITAL_ID, tempDisease[0], tempSev[0], _C))
            if temp平均在院日数 is None:
                print("エラー: 平均在院日数の取得に失敗しました。")
                return -1

            #print("疾患名 " + tempDisease[1])

            # 1行目に病院名
            sheet.cell(1, 1 + (colCnt - 1) * roopColCnt).value = HOSPITAL_NAME
            # 2行目に指標名
            sheet.cell(2, 1 + (colCnt - 1) * roopColCnt).value = "平均在院日数_疾患別_重症度別"
            # 3行目に疾患名
            sheet.cell(3, 1 + (colCnt - 1) * roopColCnt).value = tempDisease[1]
            # 4行名に重症度、年代別、性別
            sheet.cell(4, 1 + (colCnt - 1) * roopColCnt).value = tempSev[1]

            # 5行名からヘッダを出力
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt).value = "年度"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 1).value = "期"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 2).value = "平均在院日数"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 3).value = "疾患別患者数"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 4).value = "平均在院日数_ラベル"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 5).value = "疾患別患者数_ラベル"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 6).value = "平均在院日数_比較用"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 7).value = "疾患別患者数_比較用"

            # 6行目からデータ入力
            rowCnt = 6
            for tempRow in temp平均在院日数:
                #辞書から疾患別患者数を取り出すためのキー
                key = str(tempRow[0]) + str(tempRow[1]) + str(tempDisease[0])

                #年度
                sheet.cell(rowCnt, 1 + (colCnt - 1) * roopColCnt).value = tempRow[0]
                #期
                sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 1).value = tempRow[1]
                #平均在院日数
                if tempRow[2] == -1 or tempRow[2] == -2:
                    sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = 0
                else:
                    sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = tempRow[2]
                #疾患別患者数
                if key in dict_疾患別患者数:
                    if tempDisease[0] == 25 or  dict_疾患別患者数[key] == -1 or dict_疾患別患者数[key] == -2:
                        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 3).value = 0
                    else:
                        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 3).value = dict_疾患別患者数[key]
                #平均在院日数_ラベル
                if tempRow[2] == -1:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 4).value = "N/A"
                elif tempRow[2] == -2:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 4).value = "-"
                else:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 4).value = tempRow[2]
                #疾患別患者数_ラベル
                if key in dict_疾患別患者数:
                    if tempDisease[0] == 25 or dict_疾患別患者数[key] == -2:
                        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "-"
                    elif dict_疾患別患者数[key]== -1:
                        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "N/A"            
                    else:
                        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = dict_疾患別患者数[key]
                #平均在院日数_比較用
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = tempRow[2]
                #疾患別患者数_比較用
                if key in dict_疾患別患者数:
                    if tempDisease[0] == 25:
                        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = 0
                    else:
                        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = dict_疾患別患者数[key]
                rowCnt+=1
            colCnt+=1
    #wb.save(wbPath)
    print(" " + HOSPITAL_NAME + " CI_02 終了")

def getRepAgeData_重症度別(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_02 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_02" in wb.sheetnames:
        wb.remove(wb["CI_02"])
    wb.create_sheet("CI_02")
    sheet = wb["CI_02"]    

    dic_平均在院日数 = {}

    #疾患の数だけループ
    for tempDisease in CONST.MASTA_DISEASE:
        
        #対象疾患の重症度を取得
        sevList = []
        for tempSev in CONST.MASTA_SEVERITY:
            if tempSev[3] == tempDisease[0]:
                sevList.append(tempSev)

        #対象疾患の重症度の数だけループ
        for tempSev in sevList:

            #当該疾患の平均在院日数を取得
            #print(HOSPITAL_NAME + " " + tempDisease[1] + " " + tempSev[1])
            #print(HOSPITAL_NAME + " " + tempDisease[1] + " " + tempSev[1])
            #print(getSQL_CI_02_平均在院日数(HOSPITAL_ID, tempDisease[0], tempSev[0]))

            #疾患毎、重症度毎の平均在院日数リストを取得し、疾患名、重症度をキーにリストを保存
            temp平均在院日数 = X_01.excuteSQL(getSQL_CI_02_平均在院日数(HOSPITAL_ID, tempDisease[0], tempSev[0], _C))
            dic_平均在院日数[str(tempDisease[0]) + str(tempSev[0])] = temp平均在院日数
            #print("疾患名 " + tempDisease[1])
    
    # 当該疾患データの書き込み開始位置
    targetColCnt = 1

    #疾患の数だけループ       
    for tempDisease in CONST.MASTA_DISEASE:

        #対象疾患の重症度を取得
        sevList = []
        for tempSev in CONST.MASTA_SEVERITY:
            if tempSev[3] == tempDisease[0]:
                sevList.append(tempSev)
        
        wroopCnt = 0
        # 3回繰り返す　0:グラフ作成用値、1:表示用値（-,N/A等）、2:比較用値
        for i in range(3):
            if i == 0:  
                # 1行目に病院名
                sheet.cell(1, targetColCnt).value = HOSPITAL_NAME
                # 2行目に指標名
                sheet.cell(2, targetColCnt).value = "平均在院日数_疾患別_重症度別"
                # 3行目に疾患名
                sheet.cell(3, targetColCnt) .value = tempDisease[1]

                # 5行名からヘッダを出力
                sheet.cell(5, targetColCnt).value = "年度"
                sheet.cell(5, targetColCnt + 1).value = "期"

            #対象疾患の重症度の数だけループ
            for tempSev in sevList:
                #辞書から平均在院日数リストを取得
                tempList = dic_平均在院日数[str(tempDisease[0]) + str(tempSev[0])]
                
                # 6行目からデータ入力
                rowCnt = 6
                for tempRow in tempList:
                    #print(tempRow)
                    # 当該疾患、最初の重症度の時のみ年度、期のデータを出力
                    if wroopCnt == 0:
                        #年度
                        sheet.cell(rowCnt, targetColCnt).value = tempRow[0]
                        #期
                        sheet.cell(rowCnt, targetColCnt + 1).value = tempRow[1]
                    
                    # 0:グラフ作成用値
                    if i == 0:

                        #ヘッダ　重症度
                        sheet.cell(5, targetColCnt + 2 + wroopCnt).value = tempSev[1]     
                        
                        #平均在院日数
                        if tempRow[2] == -1 or tempRow[2] == -2:
                            sheet.cell(rowCnt, targetColCnt + 2 + wroopCnt).value = 0
                        else:
                            sheet.cell(rowCnt, targetColCnt + 2 + wroopCnt).value = tempRow[2]   
                        
                    # 1:表示用値（-,N/A等）
                    elif i == 1:
                        #ヘッダ　重症度
                        sheet.cell(5, targetColCnt + 2 + wroopCnt).value = tempSev[1] + "_ラベル"
                        
                        #平均在院日数_ラベル
                        if tempRow[2] == -1:
                            sheet.cell(rowCnt,  targetColCnt + 2 + wroopCnt).value = "N/A"
                        elif tempRow[2] == -2:
                            sheet.cell(rowCnt,  targetColCnt + 2 + wroopCnt).value = "-"
                        else:
                            sheet.cell(rowCnt,  targetColCnt + 2 + wroopCnt).value = tempRow[2]
                        
                    # 2:比較用値
                    elif i == 2:
                        #ヘッダ　重症度
                        sheet.cell(5, targetColCnt + 2 + wroopCnt).value = tempSev[1] + "_比較用"

                        #平均在院日数_比較用
                        sheet.cell(rowCnt,  targetColCnt + 2 + wroopCnt).value = tempRow[2]
                    rowCnt+=1
                wroopCnt+=1
        targetColCnt = targetColCnt + wroopCnt + 2 + 1

    print(" " + HOSPITAL_NAME + " CI_02 終了")

