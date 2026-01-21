import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

def getSQL_CI_56_65_歳以上の患者の入院早期の栄養アセスメント実施割合(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_56" + _C + ".年度 ,  " + \
	            "CI_56" + _C + ".期,  " + \
                "CI_56" + _C + "._65歳以上の患者の入院早期の栄養アセスメント実施割合, " + \
                "CI_56" + _C + ".分母のうち入院後48時間以内に栄養アセスメントが実施された患者数, " + \
                "CI_56" + _C + "._65歳以上の退院患者数 " + \
            "FROM " + \
	            "CI_56" + _C + " " + \
            "WHERE " + \
	            "NOT CI_56" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_56" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_56" + _C + ".年度, CI_56" + _C + ".期 " 

def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_56 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_56" in wb.sheetnames:
        wb.remove(wb["CI_56"])
    wb.create_sheet("CI_56")
    sheet = wb["CI_56"]    

    # 書込み列の位置
    colCnt = 1

    # 書込み列の繰り返し数
    roopColCnt = 12
    
    #血液培養の実施_2セット実施率
    temp65歳以上の患者の入院早期の栄養アセスメント実施割合 = X_01.excuteSQL(getSQL_CI_56_65_歳以上の患者の入院早期の栄養アセスメント実施割合(HOSPITAL_ID, _C))

    # 1行目に病院名
    sheet.cell(1, 1 + (colCnt - 1) * roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, 1 + (colCnt - 1) * roopColCnt).value = "65 歳以上の患者の入院早期の栄養アセスメント実施割合"
    # 3行目に疾患名
    sheet.cell(3, 1 + (colCnt - 1) * roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, 1 + (colCnt - 1) * roopColCnt).value = ""

    # 5行名からヘッダを出力
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt).value = "年度"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 1).value = "期"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 2).value = "65歳以上の患者の入院早期の栄養アセスメント実施割合"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 3).value = "分母のうち入院後48時間以内に栄養アセスメントが実施された患者数"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 4).value = "65歳以上の退院患者数"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 5).value = "65歳以上の患者の入院早期の栄養アセスメント実施割合_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 6).value = "血分母のうち入院後48時間以内に栄養アセスメントが実施された患者数_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 7).value = "65歳以上の退院患者数_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 8).value = "65歳以上の患者の入院早期の栄養アセスメント実施割合_比較用"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 9).value = "分母のうち入院後48時間以内に栄養アセスメントが実施された患者数_比較用"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 10).value = "65歳以上の退院患者数_比較用"

    if temp65歳以上の患者の入院早期の栄養アセスメント実施割合 is None:
        print("エラー: 血液培養の実施_2セット実施率の取得に失敗しました。")
        return -1

    # 6行目からデータ入力
    rowCnt = 6
    for tempRow in temp65歳以上の患者の入院早期の栄養アセスメント実施割合:

        #年度
        sheet.cell(rowCnt, 1 + (colCnt - 1) * roopColCnt).value = tempRow[0]
        #期
        sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 1).value = tempRow[1]
        
        #65歳以上の患者の入院早期の栄養アセスメント実施割合
        if tempRow[2] == -1 or tempRow[2] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = tempRow[2]  / 100
        #分母のうち入院後48時間以内に栄養アセスメントが実施された患者数
        if tempRow[3] == -1 or tempRow[3] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = tempRow[3] 
        #65歳以上の退院患者数
        if tempRow[4] == -1 or tempRow[4] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 4).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 4).value = tempRow[4] 

        #65歳以上の患者の入院早期の栄養アセスメント実施割合_ラベル
        if tempRow[2] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "N/A"
        elif tempRow[2] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = tempRow[2] / 100
        #分母のうち入院後48時間以内に栄養アセスメントが実施された患者数ラベル
        if tempRow[3] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = "N/A"
        elif tempRow[3] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = tempRow[3]
        #65歳以上の退院患者数_ラベル
        if tempRow[4] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = "N/A"
        elif tempRow[4] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = tempRow[4]

        #血65歳以上の患者の入院早期の栄養アセスメント実施割合_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 8).value = tempRow[2] / 100
        #分母のうち入院後48時間以内に栄養アセスメントが実施された患者数_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 9).value = tempRow[3]
        #65歳以上の退院患者数_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 10).value = tempRow[4]

        rowCnt+=1
    colCnt+=1
    #wb.save(wbPath)
    print(" " + HOSPITAL_NAME + " CI_56 終了")

