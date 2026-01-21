import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

def getSQL_疾患別患者数(HOSPITAL_ID, _C):
    return "SELECT * FROM NUMBER_OF_PATIENTS_BY_DISEASE" + _C + " WHERE Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " "  \
            "ORDER BY 年度, 期, Disease_DISEASE_ID"

def getSQL_CI_10_再入院率_疾患別_年代別(HOSPITAL_ID, disease_ID, Age_ID, _C):
    return  "SELECT  " + \
	            "CI_10" + _C + ".年度 ,  " + \
	            "CI_10" + _C + ".期,  " + \
                "CI_10" + _C + ".再入院率 " + \
            "FROM " + \
	            "CI_10" + _C + " " + \
            "WHERE " + \
	            "NOT CI_10" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_10" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " AND " + \
	            "CI_10" + _C + ".Disease_DISEASE_ID = " + str(disease_ID) + " AND " + \
                "CI_10" + _C + ".Age_AGE_ID = " + str(Age_ID) +  \
            " ORDER BY CI_10" + _C + ".年度, CI_10" + _C + ".期 " 


def getRepAgeData_年代別(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_10 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_10" in wb.sheetnames:
        ws = wb["CI_10"]
        for row in ws.rows:
            for cell in row:
                cell.value = None
        
        #wb.remove(wb["CI_10"])

    wb.create_sheet("CI_10")
    sheet = wb["CI_10"]    

    #当該病院の疾患別患者数を取得
    list_疾患別患者数 = X_01.excuteSQL(getSQL_疾患別患者数(HOSPITAL_ID, _C))
    if list_疾患別患者数 is None:
        print("エラー: 疾患別患者数の取得に失敗しました。")
        return -1

    dict_疾患別患者数 = {}

    #疾患別患者数を年度、期、DISEASE_IDをキーに辞書に追加
    for temp疾患別患者数 in list_疾患別患者数:
        # 年度、期、DISEASE_IDがキー
        key = str(temp疾患別患者数[1]) + temp疾患別患者数[2] + str(temp疾患別患者数[4])
        #print("key:" + key + " value:" + str(temp疾患別患者数[3])) 
        dict_疾患別患者数[key] = temp疾患別患者数[3]

    dic_再入院率 = {}

    #疾患の数だけループ
    for tempDisease in CONST.MASTA_DISEASE:
        
        #年代の数だけループ
        for tempAge in CONST.MASTA_AGE:

            #当該疾患の平均在院日数を取得
            #print(HOSPITAL_NAME + " " + tempDisease[1] + " " + tempSev[1])
            #print(HOSPITAL_NAME + " " + tempDisease[1] + " " + tempSev[1])
            #print(getSQL_CI_03_平均在院日数(HOSPITAL_ID, tempDisease[0], tempSev[0]))

            #疾患毎、重症度毎の平均在院日数リストを取得し、疾患名、重症度をキーにリストを保存
            temp再入院率 = X_01.excuteSQL(getSQL_CI_10_再入院率_疾患別_年代別(HOSPITAL_ID, tempDisease[0], tempAge[0], _C))
            
            key1 = ""   
            if tempDisease[0] < 10:
                key1 = "0" + str(tempDisease[0])
            else:   
                key1 = str(tempDisease[0])
            key2 = ""
            if tempAge[0] < 10:
                key2 = "0" + str(tempAge[0])
            else:   
                key2 = str(tempAge[0])
            dic_再入院率[key1 + key2] = temp再入院率
            #print("疾患名 " + tempDisease[1])
    
    # 当該疾患データの書き込み開始位置
    targetColCnt = 1

    #疾患の数だけループ       
    for tempDisease in CONST.MASTA_DISEASE:

        wroopCnt = 0
        # 3回繰り返す　0:グラフ作成用値、1:表示用値（-,N/A等）、2:比較用値
        for i in range(3):
            if i == 0:  
                # 1行目に病院名
                sheet.cell(1, targetColCnt).value = HOSPITAL_NAME
                # 2行目に指標名
                sheet.cell(2, targetColCnt).value = "予定しない再入院率_疾患別_年代別"
                # 3行目に疾患名
                sheet.cell(3, targetColCnt) .value = tempDisease[1]

                # 5行名からヘッダを出力
                sheet.cell(5, targetColCnt).value = "年度"
                sheet.cell(5, targetColCnt + 1).value = "期"

            #年代の数だけループ
            for tempAge in CONST.MASTA_AGE:
                #辞書から平均在院日数リストを取得
                key1 = ""   
                if tempDisease[0] < 10:
                    key1 = "0" + str(tempDisease[0])
                else:   
                    key1 = str(tempDisease[0])
                key2 = ""
                if tempAge[0] < 10:
                    key2 = "0" + str(tempAge[0])
                else:   
                    key2 = str(tempAge[0])
                tempList = dic_再入院率[key1 + key2]
                
                # 6行目からデータ入力
                rowCnt = 6
                for tempRow in tempList:
                    #print(tempRow)
                    # 当該疾患、最初の重症度の時のみ年度、期のデータを出力
                    if wroopCnt == 0:
                        #年度
                        sheet.cell(rowCnt, targetColCnt).value = tempRow[0]
                        #期
                        sheet.cell(rowCnt, targetColCnt + 1).value = tempRow[1]
                    
                    # 0:グラフ作成用値
                    if i == 0:

                        #ヘッダ　年代
                        sheet.cell(5, targetColCnt + 2 + wroopCnt).value = tempAge[1]     
                        
                        #再入院率
                        if tempRow[2] == -1 or tempRow[2] == -2:
                            sheet.cell(rowCnt, targetColCnt + 2 + wroopCnt).value = 0
                        else:
                            sheet.cell(rowCnt, targetColCnt + 2 + wroopCnt).value = tempRow[2]   
                        
                    # 1:表示用値（-,N/A等）
                    elif i == 1:
                        #ヘッダ　年代
                        sheet.cell(5, targetColCnt + 2 + wroopCnt).value = tempAge[1] + "_ラベル"
                        
                        #再入院率_ラベル
                        if tempRow[2] == -1:
                            sheet.cell(rowCnt,  targetColCnt + 2 + wroopCnt).value = "N/A"
                        elif tempRow[2] == -2:
                            sheet.cell(rowCnt,  targetColCnt + 2 + wroopCnt).value = "-"
                        else:
                            sheet.cell(rowCnt,  targetColCnt + 2 + wroopCnt).value = tempRow[2]
                        
                    # 2:比較用値
                    elif i == 2:
                        #ヘッダ　年代
                        sheet.cell(5, targetColCnt + 2 + wroopCnt).value = tempAge[1] + "_比較用"

                        #再入院率_比較用
                        sheet.cell(rowCnt,  targetColCnt + 2 + wroopCnt).value = tempRow[2]
                    rowCnt+=1
                wroopCnt+=1
        targetColCnt = targetColCnt + wroopCnt + 2 + 1

    print(" " + HOSPITAL_NAME + " CI_10 終了")


