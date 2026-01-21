import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

# 2024年度にCI番号変更
# CI_91　→　CI_67
# ただし、DBのテーブルは「CI_91」のまま
def getSQL_CI_67_インシデント_アクシデント_医師の占める割合(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_91" + _C + ".年度 ,  " + \
                "CI_91" + _C + ".月,  " + \
                "CI_91" + _C + ".インシデント_アクシデント_全報告中医師の占める割合 " + \
            "FROM " + \
	            "CI_91" + _C + " " + \
            "WHERE " + \
	            "CI_91" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_91" + _C + ".年度, CI_91" + _C + ".月 " 

def getSQL_CI_67_医師が提出したインシデント_アクシデント報告総件数(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_91" + _C + ".年度 ,  " + \
                "CI_91" + _C + ".月,  " + \
                "CI_91" + _C + ".医師が提出したインシデント_アクシデント報告総件数 " + \
            "FROM " + \
	            "CI_91" + _C + " " + \
            "WHERE " + \
	            "CI_91" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_91" + _C + ".年度, CI_91" + _C + ".月 " 

def getSQL_CI_67_調査期間中の月毎のインシデント_アクシデント報告総件数(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_91" + _C + ".年度 ,  " + \
                "CI_91" + _C + ".月,  " + \
                "CI_91" + _C + ".調査期間中の月毎のインシデント_アクシデント報告総件数 " + \
            "FROM " + \
	            "CI_91" + _C + " " + \
            "WHERE " + \
	            "CI_91" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_91" + _C + ".年度, CI_91" + _C + ".月 " 

def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_67 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_67" in wb.sheetnames:
        wb.remove(wb["CI_67"])
    wb.create_sheet("CI_67")
    sheet = wb["CI_67"]    

    #インシデント_アクシデント_医師の占める割合
    tempインシデント_アクシデント_医師の占める割合 = X_01.excuteSQL(getSQL_CI_67_インシデント_アクシデント_医師の占める割合(HOSPITAL_ID, _C))

    keyList = ["1ヶ月平均", "四月", "五月", "六月", "七月", "八月", "九月", "十月", 
                "十一月", "十二月", "一月", "二月", "三月" ]
    yearList = []
    dict_temp = {}

    if tempインシデント_アクシデント_医師の占める割合 is None:
        print("エラー: インシデント_アクシデント_医師の占める割合の取得に失敗しました。")
        return -1

    for a in tempインシデント_アクシデント_医師の占める割合:
        # yearListに年が存在していなければ追加
        if a[0] not in yearList:
            yearList.append(a[0])
        #年と月をキーに辞書に追加
        dict_temp[str(a[0]) + str(a[1])] = a

    # 書込み列の繰り返し数
    roopColCnt = 1
    
    # 1行目に病院名
    sheet.cell(1, roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, roopColCnt).value = "インシデント_アクシデント_医師の占める割合"
    # 3行目に疾患名
    sheet.cell(3, roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, roopColCnt).value = ""

    # 5行名からヘッダを出力
    tmpCnt = 0
    sheet.cell(5, roopColCnt).value = "年度"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "1ヶ月平均"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "四月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "五月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "六月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "七月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "八月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "九月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十一月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十二月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "一月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "二月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "三月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "1ヶ月平均_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "四月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "五月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "六月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "七月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "八月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "九月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十一月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十二月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "一月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "二月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "三月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "1ヶ月平均_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "四月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "五月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "六月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "七月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "八月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "九月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十一月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十二月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "一月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "二月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "三月_比較用"

    # 6行目からデータ入力
    rowCnt = 6
    for year in yearList:
        colCnt = 2
        for key in keyList:
            tempRow = dict_temp[str(year) + str(key)]
            #年度は"1ヶ月平均"を出力
            if colCnt == 2:
                sheet.cell(rowCnt, roopColCnt).value = tempRow[0]

            #グラフ作成用数値
            if tempRow[2] == -1 or tempRow[2] == -2:
                sheet.cell(rowCnt, colCnt).value = 0
            else:
                sheet.cell(rowCnt, colCnt).value = tempRow[2]

            #ラベル用数値
            if tempRow[2] == -1:
                sheet.cell(rowCnt, colCnt + 13).value = "N/A"
            elif tempRow[2] == -2:
                sheet.cell(rowCnt, colCnt + 13).value = "-"
            else:
                sheet.cell(rowCnt, colCnt + 13).value = tempRow[2]
            
            #比較用
            sheet.cell(rowCnt, colCnt + 26).value = tempRow[2]
            
            colCnt += 1
        rowCnt += 1

    
    #医師が提出したインシデント_アクシデント報告総件数
    temp医師が提出したインシデント_アクシデント報告総件数 = X_01.excuteSQL(getSQL_CI_67_医師が提出したインシデント_アクシデント報告総件数(HOSPITAL_ID, _C))

    keyList = ["1ヶ月平均", "四月", "五月", "六月", "七月", "八月", "九月", "十月", 
                "十一月", "十二月", "一月", "二月", "三月" ]
    yearList = []
    dict_temp = {}

    if temp医師が提出したインシデント_アクシデント報告総件数 is None:
        print("エラー: 医師が提出したインシデント_アクシデント報告総件数の取得に失敗しました。")
        return -1

    for a in temp医師が提出したインシデント_アクシデント報告総件数:
        # yearListに年が存在していなければ追加
        if a[0] not in yearList:
            yearList.append(a[0])
        #年と月をキーに辞書に追加
        dict_temp[str(a[0]) + str(a[1])] = a

    
    # 書込み列の繰り返し数
    roopColCnt = tmpCnt + 3
    
    # 1行目に病院名
    sheet.cell(1, roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, roopColCnt).value = "医師が提出したインシデント_アクシデント報告総件数"
    # 3行目に疾患名
    sheet.cell(3, roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, roopColCnt).value = ""

    # 5行名からヘッダを出力
    tmpCnt = 0
    sheet.cell(5, roopColCnt).value = "年度"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "1ヶ月平均"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "四月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "五月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "六月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "七月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "八月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "九月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十一月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十二月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "一月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "二月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "三月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "1ヶ月平均_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "四月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "五月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "六月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "七月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "八月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "九月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十一月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十二月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "一月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "二月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "三月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "1ヶ月平均_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "四月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "五月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "六月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "七月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "八月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "九月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十一月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十二月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "一月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "二月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "三月_比較用"


    # 6行目からデータ入力
    rowCnt = 6
    for year in yearList:
        colCnt = roopColCnt + 1
        for key in keyList:
            tempRow = dict_temp[str(year) + str(key)]
            #病床数は"1ヶ月平均"を出力
            if colCnt == roopColCnt + 1:
                sheet.cell(rowCnt, roopColCnt).value = tempRow[0]

             #グラフ作成用数値
            if tempRow[2] == -1 or tempRow[2] == -2:
                sheet.cell(rowCnt, colCnt).value = 0
            else:
                sheet.cell(rowCnt, colCnt).value = tempRow[2]

            #ラベル用数値
            if tempRow[2] == -1:
                sheet.cell(rowCnt, colCnt + 13).value = "N/A"
            elif tempRow[2] == -2:
                sheet.cell(rowCnt, colCnt + 13).value = "-"
            else:
                sheet.cell(rowCnt, colCnt + 13).value = tempRow[2]
            
            #比較用
            sheet.cell(rowCnt, colCnt + 26).value = tempRow[2]
            colCnt += 1
        rowCnt += 1
    

 #調査期間中の月毎のインシデント_アクシデント報告総件数
    temp調査期間中の月毎のインシデント_アクシデント報告総件数 = X_01.excuteSQL(getSQL_CI_67_調査期間中の月毎のインシデント_アクシデント報告総件数(HOSPITAL_ID, _C))

    keyList = ["1ヶ月平均", "四月", "五月", "六月", "七月", "八月", "九月", "十月", 
                "十一月", "十二月", "一月", "二月", "三月" ]
    yearList = []
    dict_temp = {}

    if temp調査期間中の月毎のインシデント_アクシデント報告総件数 is None:
        print("エラー: 調査期間中の月毎のインシデント_アクシデント報告総件数の取得に失敗しました。")
        return -1

    for a in temp調査期間中の月毎のインシデント_アクシデント報告総件数:
        # yearListに年が存在していなければ追加
        if a[0] not in yearList:
            yearList.append(a[0])
        #年と月をキーに辞書に追加
        dict_temp[str(a[0]) + str(a[1])] = a

        
    # 書込み列の繰り返し数
    roopColCnt = roopColCnt + tmpCnt + 2
    
    # 1行目に病院名
    sheet.cell(1, roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, roopColCnt).value = "調査期間中の月毎のインシデント_アクシデント報告総件数"
    # 3行目に疾患名
    sheet.cell(3, roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, roopColCnt).value = ""

    # 5行名からヘッダを出力
    tmpCnt = 0
    sheet.cell(5, roopColCnt).value = "年度"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "1ヶ月平均"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "四月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "五月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "六月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "七月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "八月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "九月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十一月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十二月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "一月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "二月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "三月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "1ヶ月平均_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "四月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "五月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "六月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "七月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "八月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "九月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十一月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十二月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "一月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "二月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "三月_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "1ヶ月平均_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "四月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "五月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "六月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "七月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "八月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "九月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十一月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "十二月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "一月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "二月_比較用"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "三月_比較用"


    # 6行目からデータ入力
    rowCnt = 6
    for year in yearList:
        colCnt = roopColCnt + 1
        for key in keyList:
            tempRow = dict_temp[str(year) + str(key)]
            #病床数は"1ヶ月平均"を出力
            if colCnt == roopColCnt + 1:
                sheet.cell(rowCnt, roopColCnt).value = tempRow[0]

             #グラフ作成用数値
            if tempRow[2] == -1 or tempRow[2] == -2:
                sheet.cell(rowCnt, colCnt).value = 0
            else:
                sheet.cell(rowCnt, colCnt).value = tempRow[2]

            #ラベル用数値
            if tempRow[2] == -1:
                sheet.cell(rowCnt, colCnt + 13).value = "N/A"
            elif tempRow[2] == -2:
                sheet.cell(rowCnt, colCnt + 13).value = "-"
            else:
                sheet.cell(rowCnt, colCnt + 13).value = tempRow[2]
            
            #比較用
            sheet.cell(rowCnt, colCnt + 26).value = tempRow[2]
            colCnt += 1
        rowCnt += 1



    #wb.save(wbPath)
    print(" " + HOSPITAL_NAME + " CI_91 終了")


