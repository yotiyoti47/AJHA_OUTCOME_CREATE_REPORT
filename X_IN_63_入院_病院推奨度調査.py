import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

# 2024年度にCI番号変更
# CI_82　→　CI_63
# ただし、DBのテーブルは「CI_82」のまま
def getSQL_CI_63_入院_病院推奨度調査_回答率(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_82" + _C + ".年度 ,  " + \
	            "CI_82" + _C + ".期,  " + \
                "CI_82" + _C + ".回答率, " + \
                "CI_82" + _C + ".回答数, " + \
                "CI_82" + _C + ".客体数 " + \
            "FROM " + \
	            "CI_82" + _C + " " + \
            "WHERE " + \
	            "NOT CI_82" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_82" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_82" + _C + ".年度, CI_82" + _C + ".期 " 

def getSQL_CI_63_入院_病院推奨度調査_満足度_割合(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_82" + _C + ".年度 ,  " + \
	            "CI_82" + _C + ".期,  " + \
                "CI_82" + _C + ".満足_割合, " + \
                "CI_82" + _C + ".やや満足_割合, " + \
                "CI_82" + _C + ".やや不満_割合, " + \
                "CI_82" + _C + ".不満_割合, " + \
                "CI_82" + _C + ".無効回答_割合 " + \
            "FROM " + \
	            "CI_82" + _C + " " + \
            "WHERE " + \
	            "NOT CI_82" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_82" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_82" + _C + ".年度, CI_82" + _C + ".期 " 

def getSQL_CI_63_入院_病院推奨度調査_満足度_件数(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_82" + _C + ".年度 ,  " + \
	            "CI_82" + _C + ".期,  " + \
                "CI_82" + _C + ".満足, " + \
                "CI_82" + _C + ".やや満足, " + \
                "CI_82" + _C + ".やや不満, " + \
                "CI_82" + _C + ".不満, " + \
                "CI_82" + _C + ".無効回答 " + \
            "FROM " + \
	            "CI_82" + _C + " " + \
            "WHERE " + \
	            "NOT CI_82" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_82" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_82" + _C + ".年度, CI_82" + _C + ".期 " 

def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_63 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_63" in wb.sheetnames:
        wb.remove(wb["CI_63"])
    wb.create_sheet("CI_63")
    sheet = wb["CI_63"]    

    # 書込み列の位置
    colCnt = 1

    # 書込み列の繰り返し数
    roopColCnt = 1    

    #入院_病院推奨度調査_回答率
    temp入院_病院推奨度調査_回答率 = X_01.excuteSQL(getSQL_CI_63_入院_病院推奨度調査_回答率(HOSPITAL_ID, _C))

    # 1行目に病院名
    sheet.cell(1, 1).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, 1).value = "入院_病院推奨度調査_回答率"
    # 3行目に疾患名
    sheet.cell(3, 1).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, 1).value = ""

    # 5行名からヘッダを出力
    sheet.cell(5, roopColCnt).value = "年度"
    sheet.cell(5, roopColCnt + 1).value = "期"
    sheet.cell(5, roopColCnt + 2).value = "回答率"
    sheet.cell(5, roopColCnt + 3).value = "回答数"
    sheet.cell(5, roopColCnt + 4).value = "客体数"
    sheet.cell(5, roopColCnt + 5).value = "回答率_ラベル"
    sheet.cell(5, roopColCnt + 6).value = "回答数_ラベル"
    sheet.cell(5, roopColCnt + 7).value = "客体数_ラベル"
    sheet.cell(5, roopColCnt + 8).value = "回答率_比較用"
    sheet.cell(5, roopColCnt + 9).value = "回答数_比較用"
    sheet.cell(5, roopColCnt + 10).value = "客体数_比較用"
    
    if temp入院_病院推奨度調査_回答率 is None:
        print("エラー: 入院_病院推奨度調査_回答率の取得に失敗しました。")
        return -1
    
    # 6行目からデータ入力
    rowCnt = 6
    for tempRow in temp入院_病院推奨度調査_回答率:

        #年度
        sheet.cell(rowCnt, roopColCnt).value = tempRow[0]
        #期
        sheet.cell(rowCnt, roopColCnt + 1).value = tempRow[1]
        
        #回答率
        if tempRow[2] == -1 or tempRow[2] == -2:
            sheet.cell(rowCnt, roopColCnt + 2).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 2).value = tempRow[2] 
        #回答数
        if tempRow[3] == -1 or tempRow[3] == -2:
            sheet.cell(rowCnt, roopColCnt + 3).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 3).value = tempRow[3] 
        #客体数
        if tempRow[4] == -1 or tempRow[4] == -2:
            sheet.cell(rowCnt, roopColCnt + 4).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 4).value = tempRow[4] 

        #回答率_ラベル
        if tempRow[2] == -1:
            sheet.cell(rowCnt,  roopColCnt + 5).value = "N/A"
        elif tempRow[2] == -2:
            sheet.cell(rowCnt,  roopColCnt + 5).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 5).value = tempRow[2]
        #回答数_ラベル
        if tempRow[3] == -1:
            sheet.cell(rowCnt,  roopColCnt + 6).value = "N/A"
        elif tempRow[3] == -2:
            sheet.cell(rowCnt,  roopColCnt + 6).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 6).value = tempRow[3]
        #客体数_ラベル
        if tempRow[4] == -1:
            sheet.cell(rowCnt,  roopColCnt + 7).value = "N/A"
        elif tempRow[4] == -2:
            sheet.cell(rowCnt,  roopColCnt + 7).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 7).value = tempRow[4]

        #回答率_比較用
        sheet.cell(rowCnt,  roopColCnt + 8).value = tempRow[2]
        #回答数_比較用
        sheet.cell(rowCnt,  roopColCnt + 9).value = tempRow[3]
        #客体数_比較用
        sheet.cell(rowCnt,  roopColCnt + 10).value = tempRow[4]

        rowCnt+=1
    colCnt+=1


    # 書込み列の繰り返し数
    roopColCnt = 13    

    #getSQL_CI_82_入院_病院推奨度調査_満足度_割合
    temp入院_病院推奨度調査_満足度_割合 = X_01.excuteSQL(getSQL_CI_63_入院_病院推奨度調査_満足度_割合(HOSPITAL_ID, _C))

    # 1行目に病院名
    sheet.cell(1, roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, roopColCnt).value = "入院_病院推奨度調査_満足度_割合"
    # 3行目に疾患名
    sheet.cell(3, roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, roopColCnt).value = ""

    # 5行名からヘッダを出力
    sheet.cell(5, roopColCnt).value = "年度"
    sheet.cell(5, roopColCnt + 1).value = "期"
    sheet.cell(5, roopColCnt + 2).value = "満足_割合"
    sheet.cell(5, roopColCnt + 3).value = "やや満足_割合"
    sheet.cell(5, roopColCnt + 4).value = "やや不満_割合"
    sheet.cell(5, roopColCnt + 5).value = "不満_割合"
    sheet.cell(5, roopColCnt + 6).value = "無効回答_割合"
    sheet.cell(5, roopColCnt + 7).value = "満足_割合_ラベル"
    sheet.cell(5, roopColCnt + 8).value = "やや満足_割合_ラベル"
    sheet.cell(5, roopColCnt + 9).value = "やや不満_割合_ラベル"
    sheet.cell(5, roopColCnt + 10).value = "不満_割合_ラベル"
    sheet.cell(5, roopColCnt + 11).value = "無効回答_割合_ラベル"
    sheet.cell(5, roopColCnt + 12).value = "満足_割合_比較用"
    sheet.cell(5, roopColCnt + 13).value = "やや満足_割合_比較用"
    sheet.cell(5, roopColCnt + 14).value = "やや不満_割合_比較用"
    sheet.cell(5, roopColCnt + 15).value = "不満_割合_比較用"
    sheet.cell(5, roopColCnt + 16).value = "無効回答_割合_比較用"

    if temp入院_病院推奨度調査_満足度_割合 is None:
        print("エラー: 入院_病院推奨度調査_満足度_割合の取得に失敗しました。")
        return -1

    # 6行目からデータ入力
    rowCnt = 6
    for tempRow in temp入院_病院推奨度調査_満足度_割合:

        #年度
        sheet.cell(rowCnt, roopColCnt).value = tempRow[0]
        #期
        sheet.cell(rowCnt, roopColCnt + 1).value = tempRow[1]
        
        #満足_割合
        if tempRow[2] == -1 or tempRow[2] == -2:
            sheet.cell(rowCnt, roopColCnt + 2).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 2).value = tempRow[2] 
        #やや満足_割合
        if tempRow[3] == -1 or tempRow[3] == -2:
            sheet.cell(rowCnt, roopColCnt + 3).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 3).value = tempRow[3] 
        #やや不満_割合
        if tempRow[4] == -1 or tempRow[4] == -2:
            sheet.cell(rowCnt, roopColCnt + 4).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 4).value = tempRow[4] 
        #不満_割合
        if tempRow[5] == -1 or tempRow[5] == -2:
            sheet.cell(rowCnt, roopColCnt + 5).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 5).value = tempRow[5] 
        #無効回答_割合
        if tempRow[6] == -1 or tempRow[6] == -2:
            sheet.cell(rowCnt, roopColCnt + 6).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 6).value = tempRow[6] 

        #満足_割合_ラベル
        if tempRow[2] == -1:
            sheet.cell(rowCnt,  roopColCnt + 7).value = "N/A"
        elif tempRow[2] == -2:
            sheet.cell(rowCnt,  roopColCnt + 7).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 7).value = tempRow[2]
        #やや満足_割合_ラベル
        if tempRow[3] == -1:
            sheet.cell(rowCnt,  roopColCnt + 8).value = "N/A"
        elif tempRow[3] == -2:
            sheet.cell(rowCnt,  roopColCnt + 8).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 8).value = tempRow[3]
        #やや不満_割合_ラベル
        if tempRow[4] == -1:
            sheet.cell(rowCnt,  roopColCnt + 9).value = "N/A"
        elif tempRow[4] == -2:
            sheet.cell(rowCnt,  roopColCnt + 9).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 9).value = tempRow[4]
        #不満_割合_ラベル
        if tempRow[5] == -1:
            sheet.cell(rowCnt,  roopColCnt + 10).value = "N/A"
        elif tempRow[5] == -2:
            sheet.cell(rowCnt,  roopColCnt + 10).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 10).value = tempRow[5]
        #無効回答_割合_ラベル
        if tempRow[6] == -1:
            sheet.cell(rowCnt,  roopColCnt + 11).value = "N/A"
        elif tempRow[6] == -2:
            sheet.cell(rowCnt,  roopColCnt + 11).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 11).value = tempRow[6]

        #満足_割合_比較用
        sheet.cell(rowCnt,  roopColCnt + 12).value = tempRow[2]
        #やや満足_割合_比較用
        sheet.cell(rowCnt,  roopColCnt + 13).value = tempRow[3]
        #やや不満_割合_比較用
        sheet.cell(rowCnt,  roopColCnt + 14).value = tempRow[4]
        #不満_割合_比較用
        sheet.cell(rowCnt,  roopColCnt + 15).value = tempRow[5]
        #無効回答_割合_比較用
        sheet.cell(rowCnt,  roopColCnt + 16).value = tempRow[6]

        rowCnt+=1
    colCnt+=1



    # 書込み列の繰り返し数
    roopColCnt = 31  


    #入院_病院推奨度調査_満足度_件数
    temp入院_病院推奨度調査_満足度_件数 = X_01.excuteSQL(getSQL_CI_63_入院_病院推奨度調査_満足度_件数(HOSPITAL_ID, _C))

    # 1行目に病院名
    sheet.cell(1, roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, roopColCnt).value = "入院_病院推奨度調査_満足度_件数"
    # 3行目に疾患名
    sheet.cell(3, roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, roopColCnt).value = ""

    # 5行名からヘッダを出力
    sheet.cell(5, roopColCnt).value = "年度"
    sheet.cell(5, roopColCnt + 1).value = "期"
    sheet.cell(5, roopColCnt + 2).value = "満足"
    sheet.cell(5, roopColCnt + 3).value = "やや満足"
    sheet.cell(5, roopColCnt + 4).value = "やや不満"
    sheet.cell(5, roopColCnt + 5).value = "不満"
    sheet.cell(5, roopColCnt + 6).value = "無効回答"
    sheet.cell(5, roopColCnt + 7).value = "満足_ラベル"
    sheet.cell(5, roopColCnt + 8).value = "やや満足_ラベル"
    sheet.cell(5, roopColCnt + 9).value = "やや不満_ラベル"
    sheet.cell(5, roopColCnt + 10).value = "不満_ラベル"
    sheet.cell(5, roopColCnt + 11).value = "無効回答_ラベル"
    sheet.cell(5, roopColCnt + 12).value = "満足_比較用"
    sheet.cell(5, roopColCnt + 13).value = "やや満足_比較用"
    sheet.cell(5, roopColCnt + 14).value = "やや不満_比較用"
    sheet.cell(5, roopColCnt + 15).value = "不満_比較用"
    sheet.cell(5, roopColCnt + 16).value = "無効回答_比較用"

    if temp入院_病院推奨度調査_満足度_件数 is None:
        print("エラー: 入院_病院推奨度調査_満足度_件数の取得に失敗しました。")
        return -1

    # 6行目からデータ入力
    rowCnt = 6
    for tempRow in temp入院_病院推奨度調査_満足度_件数:

        #年度
        sheet.cell(rowCnt, roopColCnt).value = tempRow[0]
        #期
        sheet.cell(rowCnt, roopColCnt + 1).value = tempRow[1]
        
        #満足
        if tempRow[2] == -1 or tempRow[2] == -2:
            sheet.cell(rowCnt, roopColCnt + 2).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 2).value = tempRow[2] 
        #やや満足
        if tempRow[3] == -1 or tempRow[3] == -2:
            sheet.cell(rowCnt, roopColCnt + 3).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 3).value = tempRow[3] 
        #やや不満
        if tempRow[4] == -1 or tempRow[4] == -2:
            sheet.cell(rowCnt, roopColCnt + 4).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 4).value = tempRow[4] 
        #不満
        if tempRow[5] == -1 or tempRow[5] == -2:
            sheet.cell(rowCnt, roopColCnt + 5).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 5).value = tempRow[5] 
        #無効回答
        if tempRow[6] == -1 or tempRow[6] == -2:
            sheet.cell(rowCnt, roopColCnt + 6).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 6).value = tempRow[6] 

        #満足_ラベル
        if tempRow[2] == -1:
            sheet.cell(rowCnt,  roopColCnt + 7).value = "N/A"
        elif tempRow[2] == -2:
            sheet.cell(rowCnt,  roopColCnt + 7).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 7).value = tempRow[2]
        #やや満足_ラベル
        if tempRow[3] == -1:
            sheet.cell(rowCnt,  roopColCnt + 8).value = "N/A"
        elif tempRow[3] == -2:
            sheet.cell(rowCnt,  roopColCnt + 8).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 8).value = tempRow[3]
        #やや不満_ラベル
        if tempRow[4] == -1:
            sheet.cell(rowCnt,  roopColCnt + 9).value = "N/A"
        elif tempRow[4] == -2:
            sheet.cell(rowCnt,  roopColCnt + 9).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 9).value = tempRow[4]
        #不満_ラベル
        if tempRow[5] == -1:
            sheet.cell(rowCnt,  roopColCnt + 10).value = "N/A"
        elif tempRow[5] == -2:
            sheet.cell(rowCnt,  roopColCnt + 10).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 10).value = tempRow[5]
        #無効回答_ラベル
        if tempRow[6] == -1:
            sheet.cell(rowCnt,  roopColCnt + 11).value = "N/A"
        elif tempRow[6] == -2:
            sheet.cell(rowCnt,  roopColCnt + 11).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 11).value = tempRow[6]

        #満足_比較用
        sheet.cell(rowCnt,  roopColCnt + 12).value = tempRow[2]
        #やや満足_比較用
        sheet.cell(rowCnt,  roopColCnt + 13).value = tempRow[3]
        #やや不満_比較用
        sheet.cell(rowCnt,  roopColCnt + 14).value = tempRow[4]
        #不満_比較用
        sheet.cell(rowCnt,  roopColCnt + 15).value = tempRow[5]
        #無効回答_比較用
        sheet.cell(rowCnt,  roopColCnt + 16).value = tempRow[6]

        rowCnt+=1
    colCnt+=1

    #wb.save(wbPath)
    print(" " + HOSPITAL_NAME + " CI_63 終了")

