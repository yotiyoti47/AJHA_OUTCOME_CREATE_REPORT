import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

def getSQL_CI_22_急性心筋梗塞における主要手術_治療実施率_性別(HOSPITAL_ID, Sex_ID, _C):
    return  "SELECT  " + \
	            "CI_22" + _C + ".年度 ,  " + \
	            "CI_22" + _C + ".期,  " + \
                "CI_22" + _C + ".性別実施率, " + \
                "CI_22" + _C + ".性別症例数 " + \
            "FROM " + \
	            "CI_22" + _C + " " + \
            "WHERE " + \
	            "NOT CI_22" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_22" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " AND " + \
	            "CI_22" + _C + ".Sex_SEX_ID = " + str(Sex_ID) +  \
            " ORDER BY CI_22" + _C + ".年度, CI_22" + _C + ".期 " 

def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_22 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_22" in wb.sheetnames:
        wb.remove(wb["CI_22"])
    wb.create_sheet("CI_22")
    sheet = wb["CI_22"]    

    # 書込み列の位置
    colCnt = 1

    # 書込み列の繰り返し数
    roopColCnt = 9

    #性別マスタの数だけループ
    for tempSex in CONST.MASTA_SEX:

        #当該疾患の平均在院日数を取得
        #print(HOSPITAL_NAME + " " + tempDisease[1] + " " + tempSev[1])
        #print(HOSPITAL_NAME + " " + tempDisease[1] + " " + tempSev[1])
        #print(getSQL_CI_02_平均在院日数(HOSPITAL_ID, tempDisease[0], tempSev[0]))
        temp性別実施率 = X_01.excuteSQL(getSQL_CI_22_急性心筋梗塞における主要手術_治療実施率_性別(HOSPITAL_ID, tempSex[0], _C))
        #print("疾患名 " + tempDisease[1])

        # 1行目に病院名
        sheet.cell(1, 1 + (colCnt - 1) * roopColCnt).value = HOSPITAL_NAME
        # 2行目に指標名
        sheet.cell(2, 1 + (colCnt - 1) * roopColCnt).value = "急性心筋梗塞における主要手術_治療実施率_性別"
        # 3行目に疾患名
        sheet.cell(3, 1 + (colCnt - 1) * roopColCnt).value = tempSex[1]
        # 4行名に重症度、年代別、性別
        sheet.cell(4, 1 + (colCnt - 1) * roopColCnt).value = ""

        # 5行名からヘッダを出力
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt).value = "年度"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 1).value = "期"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 2).value = "性別実施率"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 3).value = "性別症例数"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 4).value = "性別実施率_ラベル"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 5).value = "性別症例数_ラベル"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 6).value = "性別実施率_比較用"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 7).value = "性別症例数_比較用"

        if temp性別実施率 is None:
            print("エラー: 性別実施率の取得に失敗しました。")
            return -1

        # 6行目からデータ入力
        rowCnt = 6
        for tempRow in temp性別実施率:

            #年度
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt).value = tempRow[0]
            #期
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 1).value = tempRow[1]

            #性別実施率
            if tempRow[2] == -1 or tempRow[2] == -2:
                sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = 0
            else:
                sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = tempRow[2] / 100
            #性別症例数
            if tempRow[3] == -1 or tempRow[3] == -2:
                sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = 0
            else:
                sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = tempRow[3]  

            #性別実施率_ラベル
            if tempRow[2] == -1:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 4).value = "N/A"
            elif tempRow[2] == -2:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 4).value = "-"
            else:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 4).value = tempRow[2] / 100
            #性別症例数_ラベル
            if tempRow[3] == -1:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "N/A"
            elif tempRow[3] == -2:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "-"
            else:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = tempRow[3]

            #性別実施率_比較用
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = tempRow[2] / 100
            #性別症例数_比較用
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = tempRow[3]
                
            rowCnt+=1
        colCnt+=1
    #wb.save(wbPath)
    print(" " + HOSPITAL_NAME + " CI_22 終了")


