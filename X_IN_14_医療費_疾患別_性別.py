import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

def getSQL_CI_14_医療費_疾患別_性別(disease_ID, sex_ID, _C):
    return  "SELECT  " + \
	            "CI_14" + _C + ".年度 ,  " + \
	            "CI_14" + _C + ".期,  " + \
                "CI_14" + _C + ".入院費用, " + \
                "CI_14" + _C + ".入院単価日毎, " + \
                "CI_14" + _C + ".症例数 " + \
            "FROM " + \
	            "CI_14" + _C + " " + \
            "WHERE " + \
	            "NOT CI_14" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_14" + _C + ".Disease_DISEASE_ID = " + str(disease_ID) + " AND " + \
                "CI_14" + _C + ".Sex_SEX_ID = " + str(sex_ID) +  \
            " ORDER BY CI_14" + _C + ".年度, CI_14" + _C + ".期, CI_14" + _C + ".Disease_DISEASE_ID,  CI_14" + _C + ".Sex_SEX_ID" 

def getRepAgeData(wb, is_acute):
    print(" 全体集計" + " CI_14 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_14" in wb.sheetnames:
        wb.remove(wb["CI_14"])
    wb.create_sheet("CI_14")
    sheet = wb["CI_14"]    

    # 書込み列の位置
    colCnt = 1

    #疾患の数だけループ
    for tempDisease in CONST.MASTA_DISEASE:
        
        #性別の数だけループ
        for tempSex in CONST.MASTA_SEX:

            #print("疾患名 " + tempDisease[1] + " 疾患名 " + tempSev[1])
            #print(getSQL_CI_05_死亡率_疾患別_重症度別(tempDisease[0], tempSev[0]))
            temp医療費 = X_01.excuteSQL(getSQL_CI_14_医療費_疾患別_性別(tempDisease[0], tempSex[0], _C))
            #print("疾患名 " + tempDisease[1])

            # 1行目に病院名
            sheet.cell(1, 1 + (colCnt - 1) * 12).value = "全体"
            # 2行目に指標名
            sheet.cell(2, 1 + (colCnt - 1) * 12).value = "医療費_疾患別_性別"
            # 3行目に疾患名
            sheet.cell(3, 1 + (colCnt - 1) * 12).value = tempDisease[1]
            # 4行名に重症度、年代別、性別
            sheet.cell(4, 1 + (colCnt - 1) * 12).value = tempSex[1]

            # 5行名からヘッダを出力
            sheet.cell(5, 1 + (colCnt - 1) * 12).value = "年度"
            sheet.cell(5, 1 + (colCnt - 1) * 12 + 1).value = "期"
            sheet.cell(5, 1 + (colCnt - 1) * 12 + 2).value = "入院費用"
            sheet.cell(5, 1 + (colCnt - 1) * 12 + 3).value = "入院単価日毎"
            sheet.cell(5, 1 + (colCnt - 1) * 12 + 4).value = "症例数"
            sheet.cell(5, 1 + (colCnt - 1) * 12 + 5).value = "入院費用_ラベル"
            sheet.cell(5, 1 + (colCnt - 1) * 12 + 6).value = "入院単価日毎_ラベル"
            sheet.cell(5, 1 + (colCnt - 1) * 12 + 7).value = "症例数_ラベル"
            sheet.cell(5, 1 + (colCnt - 1) * 12 + 8).value = "入院費用_比較用"
            sheet.cell(5, 1 + (colCnt - 1) * 12 + 9).value = "入院単価日毎_比較用"
            sheet.cell(5, 1 + (colCnt - 1) * 12 + 10).value = "症例数_比較用"
            
            if temp医療費 is None:
                print("エラー: 医療費の取得に失敗しました。")
                return -1

            # 6行目からデータ入力
            rowCnt = 6
            for tempRow in temp医療費:

                #年度
                sheet.cell(rowCnt, 1 + (colCnt - 1) * 12).value = tempRow[0]
                #期
                sheet.cell(rowCnt, 1 +  (colCnt - 1) * 12 + 1).value = tempRow[1]

                #入院費用
                if tempRow[2] == -1 or tempRow[2] == -2:
                    sheet.cell(rowCnt, 1 +  (colCnt - 1) * 12 + 2).value = 0
                else:
                    sheet.cell(rowCnt, 1 +  (colCnt - 1) * 12 + 2).value = tempRow[2]          
                #入院単価日毎
                if tempRow[3] == -1 or tempRow[3] == -2:
                    sheet.cell(rowCnt, 1 +  (colCnt - 1) * 12 + 3).value = 0
                else:
                    sheet.cell(rowCnt, 1 +  (colCnt - 1) * 12 + 3).value = tempRow[3]   
                #症例数
                if tempRow[4] == -1 or tempRow[4] == -2:
                    sheet.cell(rowCnt, 1 +  (colCnt - 1) * 12 + 4).value = 0
                else:
                    sheet.cell(rowCnt, 1 +  (colCnt - 1) * 12 + 4).value = tempRow[4]   
                                
                #入院費用_ラベル
                if tempRow[2] == -1:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * 12 + 5).value = "N/A"
                elif tempRow[2] == -2:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * 12 + 5).value = "-"
                else:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * 12 + 5).value = tempRow[2]                
                #入院単価日毎_ラベル
                if tempRow[3] == -1:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * 12 + 6).value = "N/A"
                elif tempRow[3] == -2:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * 12 + 6).value = "-"
                else:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * 12 + 6).value = tempRow[3]
                #症例数_ラベル
                if tempRow[4] == -1:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * 12 + 7).value = "N/A"
                elif tempRow[4] == -2:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * 12 + 7).value = "-"
                else:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * 12 + 7).value = tempRow[4]

                #入院費用_比較用
                sheet.cell(rowCnt,  1 + (colCnt - 1) * 12 + 8).value = tempRow[2]
                #入院単価日毎_比較用
                sheet.cell(rowCnt,  1 + (colCnt - 1) * 12 + 9).value = tempRow[3]
                #症例数_比較用
                sheet.cell(rowCnt,  1 + (colCnt - 1) * 12 + 10).value = tempRow[4]

                rowCnt+=1
            colCnt+=1
    #wb.save(wbPath)
    print(" 全体集計" + " CI_14 終了")

