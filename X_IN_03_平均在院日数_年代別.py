import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

def getSQL_CI_03_平均在院日数(HOSPITAL_ID, disease_ID, Age_ID, _C):
    return  "SELECT  " + \
	            "CI_03" + _C + ".年度 ,  " + \
	            "CI_03" + _C + ".期,  " + \
                "CI_03" + _C + ".平均在院日数 " + \
            "FROM " + \
	            "CI_03" + _C + " " + \
            "WHERE " + \
	            "NOT CI_03" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_03" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " AND " + \
	            "CI_03" + _C + ".Disease_DISEASE_ID = " + str(disease_ID) + " AND " + \
                "CI_03" + _C + ".Age_AGE_ID = " + str(Age_ID) +  \
            " ORDER BY CI_03" + _C + ".年度, CI_03" + _C + ".期 " 

def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME):
    print(HOSPITAL_NAME + " CI_03 開始")



    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_03" in wb.sheetnames:
        wb.remove(wb["CI_03"])
    wb.create_sheet("CI_03")
    sheet = wb["CI_03"]    

    # 書込み列の位置
    colCnt = 1

    # 書込み列の繰り返し数
    roopColCnt = 6

    #疾患の数だけループ
    for tempDisease in CONST.MASTA_DISEASE:
        # 主要疾患（ID=25）は対象外
        if tempDisease[0] == 25:
            break

        #年代の数だけループ
        for tempAge in CONST.MASTA_AGE:

            #当該疾患の平均在院日数を取得
            #print(HOSPITAL_NAME + " " + tempDisease[1] + " " + tempSev[1])
            temp平均在院日数 = X_01.excuteSQL(getSQL_CI_03_平均在院日数(HOSPITAL_ID, tempDisease[0], tempAge[0]))
            #print("疾患名 " + tempDisease[1])

            # 1行目に病院名
            sheet.cell(1, 1 + (colCnt - 1) * roopColCnt).value = HOSPITAL_NAME
            # 2行目に指標名
            sheet.cell(2, 1 + (colCnt - 1) * roopColCnt).value = "平均在院日数_疾患別_年代別"
            # 3行目に疾患名
            sheet.cell(3, 1 + (colCnt - 1) * roopColCnt).value = tempDisease[1]
            # 4行名に重症度、年代別、性別
            sheet.cell(4, 1 + (colCnt - 1) * roopColCnt).value = tempAge[1]

            # 5行名からヘッダを出力
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt).value = "年度"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 1).value = "期"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 2).value = "平均在院日数"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 3).value = "平均在院日数_ラベル"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 4).value = "平均在院日数_比較用"

            # 6行目からデータ入力
            rowCnt = 6
            for tempRow in temp平均在院日数:
                #辞書から疾患別患者数を取り出すためのキー
                key = str(tempRow[0]) + str(tempRow[1]) + str(tempDisease[0])

                #年度
                sheet.cell(rowCnt, 1 + (colCnt - 1) * roopColCnt).value = tempRow[0]
                #期
                sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 1).value = tempRow[1]
                #平均在院日数
                if tempRow[2] == -1 or tempRow[2] == -2:
                    sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = 0
                else:
                    sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = tempRow[2]          
                #平均在院日数_ラベル
                if tempRow[2] == -1:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 3).value = "N/A"
                elif tempRow[2] == -2:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 3).value = "-"
                else:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 3).value = tempRow[2]
                #平均在院日数_比較用
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 4).value = tempRow[2]
                rowCnt+=1
            colCnt+=1
    #wb.save(wbPath)
    print(HOSPITAL_NAME + " CI_03 終了")

def getRepAgeData_年代別(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(HOSPITAL_NAME + " CI_03 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"
        
    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_03" in wb.sheetnames:
        wb.remove(wb["CI_03"])
    wb.create_sheet("CI_03")
    sheet = wb["CI_03"]    

    dic_平均在院日数 = {}

    #疾患の数だけループ
    for tempDisease in CONST.MASTA_DISEASE:
        
        #年代の数だけループ
        for tempAge in CONST.MASTA_AGE:

            #当該疾患の平均在院日数を取得
            #print(HOSPITAL_NAME + " " + tempDisease[1] + " " + tempSev[1])
            #print(HOSPITAL_NAME + " " + tempDisease[1] + " " + tempSev[1])
            #print(getSQL_CI_03_平均在院日数(HOSPITAL_ID, tempDisease[0], tempSev[0]))

            #疾患毎、重症度毎の平均在院日数リストを取得し、疾患名、重症度をキーにリストを保存
            #print(getSQL_CI_03_平均在院日数(HOSPITAL_ID, tempDisease[0], tempAge[0], _C))
            temp平均在院日数 = X_01.excuteSQL(getSQL_CI_03_平均在院日数(HOSPITAL_ID, tempDisease[0], tempAge[0], _C))
            dic_平均在院日数[str(tempDisease[0]) + str(tempAge[0])] = temp平均在院日数
            #print("疾患名 " + tempDisease[1])
    
    # 当該疾患データの書き込み開始位置
    targetColCnt = 1

    #疾患の数だけループ       
    for tempDisease in CONST.MASTA_DISEASE:

        wroopCnt = 0
        # 3回繰り返す　0:グラフ作成用値、1:表示用値（-,N/A等）、2:比較用値
        for i in range(3):
            if i == 0:  
                # 1行目に病院名
                sheet.cell(1, targetColCnt).value = HOSPITAL_NAME
                # 2行目に指標名
                sheet.cell(2, targetColCnt).value = "平均在院日数_疾患別_年代別"
                # 3行目に疾患名
                sheet.cell(3, targetColCnt) .value = tempDisease[1]

                # 5行名からヘッダを出力
                sheet.cell(5, targetColCnt).value = "年度"
                sheet.cell(5, targetColCnt + 1).value = "期"

            #年代の数だけループ
            for tempAge in CONST.MASTA_AGE:
                #辞書から平均在院日数リストを取得
                tempList = dic_平均在院日数[str(tempDisease[0]) + str(tempAge[0])]
                
                # 6行目からデータ入力
                rowCnt = 6
                for tempRow in tempList:
                    #print(tempRow)
                    # 当該疾患、最初の重症度の時のみ年度、期のデータを出力
                    if wroopCnt == 0:
                        #年度
                        sheet.cell(rowCnt, targetColCnt).value = tempRow[0]
                        #期
                        sheet.cell(rowCnt, targetColCnt + 1).value = tempRow[1]
                    
                    # 0:グラフ作成用値
                    if i == 0:

                        #ヘッダ　年代
                        sheet.cell(5, targetColCnt + 2 + wroopCnt).value = tempAge[1]     
                        
                        #平均在院日数
                        if tempRow[2] == -1 or tempRow[2] == -2:
                            sheet.cell(rowCnt, targetColCnt + 2 + wroopCnt).value = 0
                        else:
                            sheet.cell(rowCnt, targetColCnt + 2 + wroopCnt).value = tempRow[2]   
                        
                    # 1:表示用値（-,N/A等）
                    elif i == 1:
                        #ヘッダ　年代
                        sheet.cell(5, targetColCnt + 2 + wroopCnt).value = tempAge[1] + "_ラベル"
                        
                        #平均在院日数_ラベル
                        if tempRow[2] == -1:
                            sheet.cell(rowCnt,  targetColCnt + 2 + wroopCnt).value = "N/A"
                        elif tempRow[2] == -2:
                            sheet.cell(rowCnt,  targetColCnt + 2 + wroopCnt).value = "-"
                        else:
                            sheet.cell(rowCnt,  targetColCnt + 2 + wroopCnt).value = tempRow[2]
                        
                    # 2:比較用値
                    elif i == 2:
                        #ヘッダ　年代
                        sheet.cell(5, targetColCnt + 2 + wroopCnt).value = tempAge[1] + "_比較用"

                        #平均在院日数_比較用
                        sheet.cell(rowCnt,  targetColCnt + 2 + wroopCnt).value = tempRow[2]
                    rowCnt+=1
                wroopCnt+=1
        targetColCnt = targetColCnt + wroopCnt + 2 + 1

    print(HOSPITAL_NAME + " CI_03 終了")
