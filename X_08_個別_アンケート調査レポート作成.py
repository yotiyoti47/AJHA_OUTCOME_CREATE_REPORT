import copy
import re
from datetime import datetime
import os
import sys
import X_00_CONST as CONST
import X_01_レポート作成 as X_01
import openpyxl
import shutil
import X_IN_61_入院_患者満足度調査 as IN_61
import X_IN_62_入院_医師満足度調査 as IN_62
import X_IN_63_入院_病院推奨度調査 as IN_63
import X_IN_64_外来_患者満足度調査 as IN_64
import X_IN_65_外来_医師満足度調査 as IN_65
import X_IN_66_インシデント_アクシデント_1か月_百床あたり as IN_66
import X_IN_67_インシデント_アクシデント_医師の占める割合 as IN_67
import X_IN_68_職員の予防接種 as IN_68


#出力フォルダを設定する ※各レポートで個別実装
#引数：target_year：対象年度、target_group：対象グループ、target_period：対象期間
#戻り値：出力先フォルダのパス
#ERROR:-1
def set_output_folder(target_year):
    
    try:
        #出力先に当該年度のフォルダがあるか確認する
        output_folder = CONST.OUTPUT_FOLDER
        output_folder = os.path.join(output_folder, str(target_year))
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        #output_folder内に「05_個別レポート」フォルダがあるか確認する
        individual_folder = os.path.join(output_folder, "05_個別レポート")
        if not os.path.exists(individual_folder):
            os.makedirs(individual_folder)
    
        #survey_folder内に当該グループのフォルダがあるか確認する
        survey_folder = os.path.join(individual_folder, "02_アンケート調査")
        if not os.path.exists(survey_folder):
            os.makedirs(survey_folder)
    except:
        print("エラー: 出力先フォルダの設定に失敗しました。年度={}".format(target_year))
        return -1
      
    #report_type_folder内のファイル、フォルダを全て削除
    delete_file_count = 0
    for file in os.listdir(survey_folder):
        try:
            if os.path.isfile(os.path.join(survey_folder, file)):
                os.remove(os.path.join(survey_folder, file))
                #削除したファイル数をカウント
                delete_file_count += 1
            elif os.path.isdir(os.path.join(survey_folder, file)):
                shutil.rmtree(os.path.join(survey_folder, file))
                #削除したフォルダ数をカウント
                delete_file_count += 1
        except:
            print("フォルダ：{}".format(survey_folder))
            print(f"エラー: {file} の削除に失敗しました。")
            print("エラー内容：{}".format(sys.exc_info()[1]))
            return -1

    print("削除したファイル数：{}".format(delete_file_count))
    return survey_folder


# 過去に一度でも以下のアンケート調査にデータを提出した病院を取得
# 入院_患者満足度調査
# 入院_医師満足度調査
# 入院_病院推奨度調査
# 外来_患者満足度調査
# 外来_医師満足度調査
# インシデント_アクシデント_1か月_百床あたり
# インシデント_アクシデント_医師の占める割合
# 職員の予防接種 
def get_survey_hospitals():
    print("\n 過去にアンケート提出した病院を取得します。 ")

    hospID_list = []

    for temp_surveyName in CONST.SURVEY_GROUPS:

        tableName = ""
        # 0 "80_入院_患者満足度調査"
        # 1 "81_入院_医師満足度調査"
        # 2 "82_入院_病院推奨度調査"
        # 3 "83_外来_患者満足度調査"
        # 4 "84_外来_医師満足度調査"
        # 5 "90_インシデント_アクシデント_1か月_百床あたり"
        # 6 "91_インシデント_アクシデント_医師の占める割合"
        # 7 "92_転倒_転落"
        # 8 "93_転倒_転落_3b以上"
        # 9 "94_職員の予防接種"
        if temp_surveyName == CONST.SURVEY_GROUPS[0]:
            tableName = "CI_80"
        elif temp_surveyName == CONST.SURVEY_GROUPS[1]:
            tableName = "CI_81"
        elif temp_surveyName == CONST.SURVEY_GROUPS[2]:
            tableName = "CI_82"
        elif temp_surveyName == CONST.SURVEY_GROUPS[3]:
            tableName = "CI_83"
        elif temp_surveyName == CONST.SURVEY_GROUPS[4]:
            tableName = "CI_84"
        elif temp_surveyName == CONST.SURVEY_GROUPS[5]:
            tableName = "CI_90"
        elif temp_surveyName == CONST.SURVEY_GROUPS[6]:
            tableName = "CI_91"
        elif temp_surveyName == CONST.SURVEY_GROUPS[7]:
            tableName = "CI_92"
        elif temp_surveyName == CONST.SURVEY_GROUPS[8]:
            tableName = "CI_93"
        elif temp_surveyName == CONST.SURVEY_GROUPS[9]:
            tableName = "CI_94"
        else:
            print("エラー: アンケート調査名が不正です。")
            return -1

        sql1 = "SELECT DISTINCT Hospital_HOSPITAL_ID FROM " + tableName 
        sql2 = "SELECT DISTINCT Hospital_HOSPITAL_ID FROM " + tableName + "_C"
        hospID1 =X_01.excuteSQL(sql1, None)
        hospID2 =X_01.excuteSQL(sql1, None)
    
        if hospID1 is None and hospID2 is None:
            print("エラー: 対象病院の取得に失敗しました。")
            return -1

        if hospID1 is not None:
            for a in range(len(hospID1)):
                hospID_list.append(hospID1[a][0])

        if hospID2 is not None:
            for a in range(len(hospID2)):
                hospID_list.append(hospID2[a][0])

    #hospID_listから重複を除く
    hospID_list = list(set(hospID_list))

    return hospID_list




# レポート作成対象となる病院を取得
#  2024年度にデータを提出した病院かつ
# 過去に１度でも以下のアンケート調査にデータを提出した病院が対象
def get_target_hospitals(target_year):
    print("\nレポート作成対象となる病院を取得します。")

    hospID_list = []

    ##
    ##急性期グループ
    ##
    sql_getMaxYear_急性期 = "SELECT MAX(年度) FROM TABLE_STATUS" + " WHERE QI_NAME = '01_平均在院日数_疾患別' AND STATUS = '登録'"
    maxYear_急性期 =X_01.excuteSQL(sql_getMaxYear_急性期, None)
    if maxYear_急性期 is None:
        print("エラー: 最新年度の取得に失敗しました。")
        return -1
        #print("maxYear_急性期")
    #   print(maxYear_急性期[0][0])

    sql_getMaxQuater_急性期 = "SELECT MAX(期) FROM TABLE_STATUS" + " WHERE QI_NAME = '01_平均在院日数_疾患別' AND STATUS = '登録'" \
        + " AND 年度 = " + str(maxYear_急性期[0][0] ) + " "
    maxQuater_急性期 =X_01.excuteSQL(sql_getMaxQuater_急性期, None)
    if maxQuater_急性期 is None:
        print("エラー: 最新期の取得に失敗しました。")
        return -1

    CONST.INDIV_TARGET_QUARTER_急性期 = maxQuater_急性期[0][0]  

    sql_getTargetHospitals_急性期  = "SELECT DISTINCT Hospital_HOSPITAL_ID FROM CI_01" + " WHERE 年度 = " \
        + str(maxYear_急性期[0][0]) + " AND 期 = '" + maxQuater_急性期[0][0]  + "' AND Disease_DISEASE_ID = 25 " \
        + " AND 平均在院日数 <> -2  "
    hospID_急性期 =X_01.excuteSQL(sql_getTargetHospitals_急性期, None)
    if hospID_急性期 is None:
        print("エラー: 対象病院の取得に失敗しました。")
        return -1
        
    for a in range(len(hospID_急性期)):
        hospID_list.append(hospID_急性期[a][0])

    ##
    ##慢性期グループ
    ##
    sql_getMaxYear_慢性期 = "SELECT MAX(年度) FROM TABLE_STATUS_C" + " WHERE QI_NAME = '01_平均在院日数_疾患別' AND STATUS = '登録'"
    maxYear_慢性期 =X_01.excuteSQL(sql_getMaxYear_慢性期, None)
    if maxYear_慢性期 is None:
        print("エラー: 最新年度の取得に失敗しました。")
        return -1

    sql_getMaxQuater_慢性期 = "SELECT MAX(期) FROM TABLE_STATUS_C" + " WHERE QI_NAME = '01_平均在院日数_疾患別' AND STATUS = '登録'" \
        + " AND 年度 = " + str(maxYear_慢性期[0][0] ) + " "
    maxQuater_慢性期 =X_01.excuteSQL(sql_getMaxQuater_慢性期, None)
    if maxQuater_慢性期 is None:
        print("エラー: 最新期の取得に失敗しました。")
        return -1

    CONST.INDIV_TARGET_QUARTER_慢性期 = maxQuater_慢性期[0][0]

    sql_getTargetHospitals_慢性期  = "SELECT DISTINCT Hospital_HOSPITAL_ID FROM CI_01_C" + " WHERE 年度 = " \
        + str(maxYear_慢性期[0][0]) + " AND 期 = '" + maxQuater_慢性期[0][0]  + "' AND Disease_DISEASE_ID = 25 " \
        + " AND 平均在院日数 <> -2  "
    hospID_慢性期 =X_01.excuteSQL(sql_getTargetHospitals_慢性期, None)
    if hospID_慢性期 is None:
        print("エラー: 対象病院の取得に失敗しました。")
        return -1

    for a in range(len(hospID_慢性期)):
        hospID_list.append(hospID_慢性期[a][0])

    #hospID_listから重複を除く
    hospID_list = list(set(hospID_list))

    #過去にアンケート提出した病院を取得
    survey_hospID_list = get_survey_hospitals()
    if survey_hospID_list != -1:
        #hospID_listとsurvey_hospID_listで重複している値を取得
        hospID_list = list(set(hospID_list) & set(survey_hospID_list))





    hosIDs = ",".join(str(hospID) for hospID in hospID_list)

    sql_target_hospName = "SELECT 病院名, HOSPITAL_ID FROM HOSPITAL WHERE HOSPITAL_ID IN (" + hosIDs+ " ) ORDER BY 都道府県コード, 病院名_よみがな"
    temp_hospName_list = X_01.excuteSQL(sql_target_hospName, None)
    if temp_hospName_list is None:
        print("エラー: 対象病院の取得に失敗しました。")
        return -1

    hospName_list = []
    for a in range(len(temp_hospName_list)):
        hospName_list.append((temp_hospName_list[a][0],temp_hospName_list[a][1]))

    print("hospName_list")
    print(hospName_list)

    #print("急性期 年度{},期{}の対象病院数:{}".format(maxYear_急性期[0][0], maxQuater_急性期[0][0], len(hospID_list)))
    #print("慢性期 年度{},期{}の対象病院数:{}".format(maxYear_慢性期[0][0], maxQuater_慢性期[0][0], len(hospID_list)))
    print("{} ：対象病院数:{}".format(target_year, len(hospName_list)))
    return hospName_list





















def copyFormatFile(target_year, hospName_list, output_folder):
    print("\nフォーマットファイルをコピーします。")

    #病院の数だけ繰り返し
    repCnt = 1

    temp_repCnt = 0
    #output_folder内のxlsmファイルを取得
    xlsm_file_list = [f for f in os.listdir(output_folder) if f.endswith('.xlsm')]

    for xlsm_file in xlsm_file_list:
        xlsm_file_name = os.path.splitext(xlsm_file )[0]
        #xlsm_file_name内にある数値２桁を取得
        xlsm_file_name_num = re.search(r'\d{2}', xlsm_file_name)
       
        if xlsm_file_name_num is not None:
            xlsm_file_name_num = xlsm_file_name_num.group()
        else:
            print("エラー: 数値２桁が取得できませんでした。")
            print("ファイル名:{}".format(xlsm_file_name))
            return -1

        xlsm_file_name_num = int(xlsm_file_name_num)

        if xlsm_file_name_num > temp_repCnt:
            temp_repCnt = xlsm_file_name_num

    repCnt = temp_repCnt + 1
          
    try:
        for hospName in hospName_list:
        #print(str(repCnt) + ":" + hospName + " 開始")
        
            format_file_path = CONST.INDIV_REP_FRMT_SURVEY_PATH
            if not os.path.exists(format_file_path):
                print("エラー: フォーマットファイルが存在しません。")
                print("フォーマットファイルのパス:{}".format(format_file_path))
                return -1

            temp_repCnt = ""
            if repCnt <= 9:
                temp_repCnt = "0" + str(repCnt)
            else:
                temp_repCnt = str(repCnt)

            newName = "個別レポート_アンケート調査_" + str(temp_repCnt) + "_" + hospName[0] + ".xlsm"

            #format_file_pathをnewNameでコピー
            shutil.copy(format_file_path, os.path.join(output_folder, newName))            
            repCnt += 1

    except:
        print("エラー: フォーマットファイルのコピーに失敗しました。")
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1

    print("フォーマットファイルのコピーに成功しました。")
    return 0


def create_indiv_report(target_create_date, hosp_list, output_folder):
    print("\n個別レポートを作成します。")

    #output_folder内のxlsmファイルを取得
    xlsm_file_list = [f for f in os.listdir(output_folder) if f.endswith('.xlsm')]
    if len(xlsm_file_list) == 0:
        #存在しない場合もあるので正常終了で返す
        print("エラー: xlsmファイルが存在しません。")
        return 0

    # hospName_list
    # 0 病院名
    # 1 病院ID
    for hosp in hosp_list:
        for xlsm_file in xlsm_file_list:
            if hosp[0] in xlsm_file:
                print("{} を作成します。".format(xlsm_file))
                print("病院ID:{}".format(hosp[1]))
                print("病院名:{}".format(hosp[0]))

                wb = openpyxl.load_workbook(os.path.join(output_folder, xlsm_file), keep_vba=True)

                if hosp[0] == "竹口病院":
                    is_acute = "慢性期"

                ws = wb["レポート_表紙"]
                ws["CF4"] = hosp[0]
                ws["CF5"] = target_create_date.strftime("%Y/%m/%d")

                # 入院_患者満足度調査
                IN_61.getRepAgeData(wb, hosp[1], hosp[0], "急性期")
                # 入院_医師満足度調査
                IN_62.getRepAgeData(wb, hosp[1], hosp[0], "急性期")
                # 入院_病院推奨度調査
                IN_63.getRepAgeData(wb, hosp[1], hosp[0], "急性期")
                # 外来_患者満足度調査
                IN_64.getRepAgeData(wb, hosp[1], hosp[0], "急性期")
                # 外来_医師満足度調査
                IN_65.getRepAgeData(wb, hosp[1], hosp[0], "急性期")
                # インシデント_アクシデント_1か月_百床あたり
                IN_66.getRepAgeData(wb, hosp[1], hosp[0], "急性期")
                # インシデント_アクシデント_医師の占める割合
                IN_67.getRepAgeData(wb, hosp[1], hosp[0], "急性期")
                
                # 転倒_転落、転倒_転落_3b以上は2024年度より急性期レポートへ
                # R6診療報酬改定により、転倒_転落、転倒_転落_3b以上は様式１、様式３に
                # 組み込まれたためアンケート調査の対象外
    
                # 職員の予防接種
                IN_68.getRepAgeData(wb, hosp[1], hosp[0], "急性期")
     
                wb.save(os.path.join(output_folder, xlsm_file))
                wb.close()

                






def create_report():
    print("\n経年比較  アンケート調査レポートを作成します。")

    print("\n年度（西暦4桁、半角）を入力してください。")
    user_input = input(":>>")
    if not user_input.isdigit() or len(user_input) != 4:
        print("\nエラー: 西暦4桁の半角数字を入力してください。")
        return -1

    target_year = int(user_input)

    print("\nレポートの作成日をyyyy/m/d形式で入力してください。")
    user_input = input(":>>")
    if not re.match(r'^\d{4}/\d{1,2}/\d{1,2}$', user_input):
        print("\nエラー: 日付の形式が不正です。yyyy/m/d形式で入力してください。")
        return -1

    target_create_date = datetime.strptime(user_input, "%Y/%m/%d")
    print(f"\n対象作成日: {target_create_date.strftime('%Y/%m/%d')}\n")

    #出力先フォルダの設定
    output_folder = set_output_folder(target_year)
    if output_folder is None or output_folder == -1:
        return -1

    #作成対象となる病院を取得
    #まれに慢性期だけにデータを提出している病院がいるので注意！

    # 0 病院名
    # 1 病院ID
    hospID_list = get_target_hospitals(target_year)
    if hospID_list == -1:
        return -1

    if copyFormatFile(target_year, hospID_list, output_folder) == -1:   
        return -1

    # マスタデータをセット
    X_01.setMastaData()

    create_indiv_report(target_create_date, hospID_list, output_folder)
    



    return 0


