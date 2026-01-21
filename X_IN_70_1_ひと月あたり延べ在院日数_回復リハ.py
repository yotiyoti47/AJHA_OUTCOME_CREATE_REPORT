import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

# 2024年度にCI番号変更
# CI_103　→　CI_70_1
# ただし、DBのテーブルは「CI_103」のまま
def getSQL_CI_70_1_ひと月あたり延べ在院日数_回復リハ(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_103" + _C + ".年度 ,  " + \
	            "CI_103" + _C + ".月, " + \
                "CI_103" + _C + ".病床数, " + \
                "CI_103" + _C + ".延べ在院日数 " + \
            "FROM " + \
	            "CI_103" + _C + " " + \
            "WHERE " + \
	            "CI_103" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_103" + _C + ".年度, CI_103" + _C + ".月 " 


def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_70_1 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_70_1" in wb.sheetnames:
        wb.remove(wb["CI_70_1"])
    wb.create_sheet("CI_70_1")
    sheet = wb["CI_70_1"]    

    #ひと月あたり延べ在院日数_回復リハ
    tempひと月あたり延べ在院日数_回復リハ = X_01.excuteSQL(getSQL_CI_70_1_ひと月あたり延べ在院日数_回復リハ(HOSPITAL_ID, _C))

    if tempひと月あたり延べ在院日数_回復リハ is None:
        print("エラー: ひと月あたり延べ在院日数_回復リハの取得に失敗しました。")
        return -1

    keyList = ["1ヶ月平均", "四月", "五月", "六月", 
                "七月", "八月", "九月", "十月", "十一月", "十二月", "一月", 
                "二月", "三月" ]
    yearList = []
    dict_temp = {}

    for a in tempひと月あたり延べ在院日数_回復リハ:
        # yearListに年が存在していなければ追加
        if a[0] not in yearList:
            yearList.append(a[0])
        #年と月をキーに辞書に追加
        dict_temp[str(a[0]) + str(a[1])] = a

    
    # 書込み列の繰り返し数
    roopColCnt = 1
    
    # 1行目に病院名
    sheet.cell(1, roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, roopColCnt).value = "ひと月あたり延べ在院日数_回復リハ"
    # 3行目に疾患名
    sheet.cell(3, roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, roopColCnt).value = ""

    # 5行名からヘッダを出力
    sheet.cell(5, roopColCnt).value = "病床数"
    tmpCnt = 1
    sheet.cell(5, roopColCnt + tmpCnt).value = "年度"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "月"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "ひと月あたり延べ在院日数"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "ひと月あたり延べ在院日数_ラベル"
    tmpCnt+=1
    sheet.cell(5, roopColCnt + tmpCnt).value = "ひと月あたり延べ在院日数_比較用"

    # 6行目からデータ入力
    rowCnt = 6
    for year in yearList:
        for key in keyList:
            tempRow = dict_temp[str(year) + str(key)]
            #病床数
            sheet.cell(rowCnt, roopColCnt).value = tempRow[2]
            tmpCnt = 1
            #年度
            sheet.cell(rowCnt, roopColCnt + tmpCnt).value = tempRow[0]
            tmpCnt+=1
            #月
            sheet.cell(rowCnt, roopColCnt + tmpCnt).value = tempRow[1]
            tmpCnt+=1

            #グラフ作成用数値　
            if tempRow[3] == -1 or tempRow[3] == -2:
                sheet.cell(rowCnt, roopColCnt + tmpCnt).value = 0
            else:
                sheet.cell(rowCnt, roopColCnt + tmpCnt).value = tempRow[3] 
            tmpCnt+=1

            #ラベル用数値
            if tempRow[3] == -1:
                sheet.cell(rowCnt, roopColCnt + tmpCnt).value = "N/A"
            elif tempRow[3] == -2:
                sheet.cell(rowCnt, roopColCnt + tmpCnt).value = "-"
            else:
                sheet.cell(rowCnt, roopColCnt + tmpCnt).value = tempRow[3]
            tmpCnt+=1
            #比較用
            sheet.cell(rowCnt, roopColCnt + tmpCnt).value = tempRow[3]
            rowCnt += 1

    print(" " + HOSPITAL_NAME + " CI_70_1 終了")

