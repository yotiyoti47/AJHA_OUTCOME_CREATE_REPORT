import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

def getSQL_疾患別患者数(HOSPITAL_ID, _C):
    return "SELECT * FROM NUMBER_OF_PATIENTS_BY_DISEASE" + _C + " WHERE Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " "  \
            "ORDER BY 年度, 期, Disease_DISEASE_ID"

def getSQL_CI_19_65歳以上の患者における認知症の保有率_入院経路別(HOSPITAL_ID, Source_ID, _C):
    return  "SELECT  " + \
	            "CI_19" + _C + ".年度 ,  " + \
	            "CI_19" + _C + ".期,  " + \
                "CI_19" + _C + ".入院経路別保有率, " + \
                "CI_19" + _C + ".入院経路別症例数 " + \
            "FROM " + \
	            "CI_19" + _C + " " + \
            "WHERE " + \
	            "NOT CI_19" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_19" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " AND " + \
	            "CI_19" + _C + ".Source_SOURCE_ID = " + str(Source_ID) +  \
            " ORDER BY CI_19" + _C + ".年度, CI_19" + _C + ".期 " 

def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(HOSPITAL_NAME + " CI_19 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_19" in wb.sheetnames:
        wb.remove(wb["CI_19"])
    wb.create_sheet("CI_19")
    sheet = wb["CI_19"]    

    # 書込み列の位置
    colCnt = 1
    
    # 書込み列の繰り返し数
    roopColCnt = 9

    #入院経路の数だけループ
    for tempSource in CONST.MASTA_SOURCE:

        #65歳以上の患者における認知症の保有率を取得
        #print(HOSPITAL_NAME + " " + tempSource[1] )
        #print(getSQL_CI_19_65歳以上の患者における認知症の保有率_入院経路別(HOSPITAL_ID, tempSource[0]))
        temp認知症の保有率 = X_01.excuteSQL(getSQL_CI_19_65歳以上の患者における認知症の保有率_入院経路別(HOSPITAL_ID, tempSource[0], _C))
        #print("疾患名 " + tempDisease[1])

        # 1行目に病院名
        sheet.cell(1, 1 + (colCnt - 1) * roopColCnt).value = HOSPITAL_NAME
        # 2行目に指標名
        sheet.cell(2, 1 + (colCnt - 1) * roopColCnt).value = "65歳以上の患者における認知症の保有率_入院経路別"
        # 3行目に疾患名
        sheet.cell(3, 1 + (colCnt - 1) * roopColCnt).value = tempSource[1]
        # 4行名に重症度、年代別、性別
        sheet.cell(4, 1 + (colCnt - 1) * roopColCnt).value = ""

        # 5行名からヘッダを出力
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt).value = "年度"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 1).value = "期"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 2).value = "入院経路別保有率"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 3).value = "入院経路別症例数"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 4).value = "入院経路別保有率_ラベル"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 5).value = "入院経路別症例数_ラベル"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 6).value = "入院経路別保有率_比較用"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 7).value = "入院経路別症例数_比較用"

        # 6行目からデータ入力
        rowCnt = 6
        for tempRow in temp認知症の保有率:
            #print(tempRow)
            #年度
            sheet.cell(rowCnt, 1 + (colCnt - 1) * roopColCnt).value = tempRow[0]
            #期
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 1).value = tempRow[1]
            
            #入院経路別保有率
            if tempRow[2] == -1 or tempRow[2] == -2:
                sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = 0
            else:
                sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = tempRow[2] / 100   
            #入院経路別症例数
            if tempRow[3] == -1 or tempRow[3] == -2:
                sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = 0
            else:
                sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = tempRow[3]       
                   
            #入院経路別保有率_ラベル
            if tempRow[2] == -1:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 4).value = "N/A"
            elif tempRow[2] == -2:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 4).value = "-"
            else:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 4).value = tempRow[2] / 100
            #入院経路別症例数_ラベル
            if tempRow[3] == -1:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "N/A"
            elif tempRow[3] == -2:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "-"
            else:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = tempRow[3]
            
            #入院経路別保有率_比較用
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = tempRow[2] / 100
            #入院経路別症例数_比較用
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = tempRow[3]
                        
            rowCnt+=1
        colCnt+=1
    
    #wb.save(wbPath)
    print(HOSPITAL_NAME + " CI_19 終了")


def getRepAgeData_old(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_19 開始")
    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"
    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_19" in wb.sheetnames:
        ws = wb["CI_19"]
        for row in ws.rows:
            for cell in row:
                cell.value = None
        
        #wb.remove(wb["CI_19"])

    wb.create_sheet("CI_19")
    sheet = wb["CI_19"]    

    #当該病院の疾患別患者数を取得
    list_疾患別患者数 = X_01.excuteSQL(getSQL_疾患別患者数(HOSPITAL_ID, _C))
    if list_疾患別患者数 is None:
        print("エラー: 疾患別患者数の取得に失敗しました。")
        return -1

    dict_疾患別患者数 = {}

    #疾患別患者数を年度、期、DISEASE_IDをキーに辞書に追加
    for temp疾患別患者数 in list_疾患別患者数:
        # 年度、期、DISEASE_IDがキー
        key = str(temp疾患別患者数[1]) + temp疾患別患者数[2] + str(temp疾患別患者数[4])
        #print("key:" + key + " value:" + str(temp疾患別患者数[3])) 
        dict_疾患別患者数[key] = temp疾患別患者数[3]

    # 書込み列の位置
    colCnt = 1
    
    # 書込み列の繰り返し数
    roopColCnt = 12

    #入院経路の数だけループ
    for tempSource in CONST.MASTA_SOURCE:

        #65歳以上の患者における認知症の保有率を取得
        #print(HOSPITAL_NAME + " " + tempSource[1] )
        #print(getSQL_CI_19_65歳以上の患者における認知症の保有率_入院経路別(HOSPITAL_ID, tempSource[0]))
        temp認知症の保有率 = X_01.excuteSQL(getSQL_CI_19_65歳以上の患者における認知症の保有率_入院経路別(HOSPITAL_ID, tempSource[0], _C))
        #print("疾患名 " + tempDisease[1])

        # 1行目に病院名
        sheet.cell(1, 1 + (colCnt - 1) * roopColCnt).value = HOSPITAL_NAME
        # 2行目に指標名
        sheet.cell(2, 1 + (colCnt - 1) * roopColCnt).value = "65歳以上の患者における認知症の保有率_入院経路別"
        # 3行目に疾患名
        sheet.cell(3, 1 + (colCnt - 1) * roopColCnt).value = tempSource[1]
        # 4行名に重症度、年代別、性別
        sheet.cell(4, 1 + (colCnt - 1) * roopColCnt).value = ""

        # 5行名からヘッダを出力
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt).value = "年度"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 1).value = "期"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 2).value = "入院経路別保有率"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 3).value = "入院経路別症例数"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 4).value = "疾患別患者数"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 5).value = "入院経路別保有率_ラベル"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 6).value = "入院経路別症例数_ラベル"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 7).value = "疾患別患者数_ラベル"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 8).value = "入院経路別保有率_比較用"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 9).value = "入院経路別症例数_比較用"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 10).value = "疾患別患者数_比較用"

        if temp認知症の保有率 is None:
            print("エラー: 認知症の保有率の取得に失敗しました。")
            return -1

        # 6行目からデータ入力
        rowCnt = 6
        for tempRow in temp認知症の保有率:
            #print(tempRow)
            #辞書から疾患別患者数を取り出すためのキー（疾患IDは25を使用）
            key = str(tempRow[0]) + str(tempRow[1]) + "25"
            
            #年度
            sheet.cell(rowCnt, 1 + (colCnt - 1) * roopColCnt).value = tempRow[0]
            #期
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 1).value = tempRow[1]
            
            #入院経路別保有率
            if tempRow[2] == -1 or tempRow[2] == -2:
                sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = 0
            else:
                sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = tempRow[2] / 100   
            #入院経路別症例数
            if tempRow[3] == -1 or tempRow[3] == -2:
                sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = 0
            else:
                sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = tempRow[3]
            #疾患別患者数
            if key in dict_疾患別患者数:
                if dict_疾患別患者数[key] == -1 or dict_疾患別患者数[key] == -2:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 4).value = 0
                else:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 4).value = dict_疾患別患者数[key]
                   
            #入院経路別保有率_ラベル
            if tempRow[2] == -1:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "N/A"
            elif tempRow[2] == -2:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "-"
            else:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = tempRow[2] / 100
            #入院経路別症例数_ラベル
            if tempRow[3] == -1:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = "N/A"
            elif tempRow[3] == -2:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = "-"
            else:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = tempRow[3]
            #疾患別患者数_ラベル
            if key in dict_疾患別患者数:
                if dict_疾患別患者数[key] == -2:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = "-"
                elif dict_疾患別患者数[key] == -1:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = "N/A"
                else:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = dict_疾患別患者数[key]
            
            #入院経路別保有率_比較用
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 8).value = tempRow[2] / 100
            #入院経路別症例数_比較用
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 9).value = tempRow[3]
            #疾患別患者数_比較用
            if key in dict_疾患別患者数:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 10).value = dict_疾患別患者数[key]
                        
            rowCnt+=1
        colCnt+=1
    
    #wb.save(wbPath)
    print(" " + HOSPITAL_NAME + " CI_19 終了")

