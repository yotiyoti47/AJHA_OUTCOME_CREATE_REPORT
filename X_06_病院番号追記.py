




import tkinter as tk
from tkinter import filedialog
import X_00_CONST as CONST
import openpyxl
import os
import X_01_レポート作成 as COMMON









def add_hospital_number():
    print("\n参加病院一覧に病院番号を追記します。")

    print("\n年度（西暦4桁、半角）を入力してください。")
    user_input = input(":>>")
    if not user_input.isdigit() or len(user_input) != 4:
        print("\nエラー: 西暦4桁の半角数字を入力してください。")
        return -1

    target_year = int(user_input)
    print(f"\n対象年度: {str(target_year)}\n")


    print("\n参加病院一覧ファイルを指定してください。")
    #ファイルダイアログを開く
    root = tk.Tk()
    #root.withdraw()
    list_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not list_path:
        print("\nエラー: 参加病院一覧ファイルが指定されていません。")
        return -1
    

    #当該年度の平均在院日数（疾患別）から1Qのデータ提出病院のHospital_HOSPITAL_IDを取得し、
    #HOSPITALテーブルから病院ID、病院名を取得する
    print("target_year={}年度第1四半期「平均在院日数（疾患別）」テーブルからデータ提出病院のID、病院名を取得".format(target_year))
    hospital_data = COMMON.getHospIDAndName(target_year)
    if hospital_data is None or hospital_data == -1:
        print("\nエラー: 病院データの取得に失敗しました。年度={}、".format(target_year))
        return -1   
    #cnt = 1
    #for data in hospital_data:
    #    print("No={} data: name={}: no={}".format(cnt,data[1],data[0]))
    #    cnt += 1

    # hospital_dataから「参加病院全体」が含まれている要素を削除する
    hospital_data = [data for data in hospital_data if "参加病院全体" not in data[1]]
    print("\n フィルタリング後の病院データ")
    print("病院データ 取得件数={}、年度={}\n".format(str(len(hospital_data)),target_year, ))

    # 辞書型
    # key:病院名、value:公開用病院番号
    dict_public_no = COMMON.createDictPublicNo(target_year, hospital_data)
    if dict_public_no is None or dict_public_no == -1:
        print("\nエラー: 公開用病院番号の作成に失敗しました。年度={}".format(target_year))
        return -1
    
    #エクセルファイルを開
    try:
        wb = openpyxl.load_workbook(list_path)
        excel_filename = os.path.basename(list_path)
        print(f"選択されたファイル名: {excel_filename}")

        ws = wb.active
        print(f"シート名: {ws.title}")

        # ヘッダ（1行目データ）を取得
        first_row = [cell.value for cell in ws[1]]

        hospNAme_idx =0
        hospNo_idx = 0
        for col in first_row:
            if col == "施設名":
                hospNAme_idx = first_row.index(col) + 1
                continue
            elif col == "公開用病院番号":
                hospNo_idx = first_row.index(col) + 1  

            if hospNAme_idx != 0 and hospNo_idx != 0: 
                break

        keys = dict_public_no.keys()
        for row in ws.iter_rows(min_row=2, max_col=max(hospNAme_idx, hospNo_idx), values_only=False):
            # 病院名を取得
            if row[hospNAme_idx-1].value is None:
                # 病院名が空白の場合は終了
                break

            tempHospName = row[hospNAme_idx-1].value

            ismatch = False
            for key in keys:
                if key in tempHospName:       
                    # 病院名が辞書に存在する場合、公開用病院番号を設定
                    public_no = dict_public_no[key]
                    print(f"病院名: {tempHospName} に対応する公開用病院番号: {public_no} を発見しました。")
                    # 病院名と公開用病院番号を設定
                    ws.cell(row=row[0].row, column=hospNo_idx, value=public_no)
                    ismatch = True
                    break
  
            if not ismatch:
                # 病院名が辞書に存在しない場合、空白を設定
                print(f"\nエラー: 病院名が一致しません。")
                print(f"病院名: {tempHospName} は辞書に存在しません。公開用病院番号は設定されません。")
                return -1

        # エクセルファイルを保存
        wb.save(list_path)
        wb.close()
        print(f"\n病院番号の追記が完了しました。ファイル名: {excel_filename}")

    except Exception as e:
        print(f"\nエラー: エクセルファイルの読み込みに失敗しました。詳細: {e}")
        return -1

    
