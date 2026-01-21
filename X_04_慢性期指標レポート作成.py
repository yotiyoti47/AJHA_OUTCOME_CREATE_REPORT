import re
from datetime import datetime
import tkinter as tk
from tkinter import filedialog
import X_01_レポート作成 as COMMON
import X_00_CONST as CONST
import os
import sys
import openpyxl
from openpyxl.styles import PatternFill
import win32com.client
import shutil
#from PyPDF2 import PdfReader, PdfWriter
from pypdf import PdfReader, PdfWriter
from openpyxl import load_workbook



LIST_KEY_FILENAMES = [
    [["ひと月あたり症例数", "回復リハ"],"69_1_ひと月あたり症例数（回復リハ）レポート"],
    [["ひと月あたり症例数", "地域包括ケア"],"69_2_ひと月あたり症例数（地域包括ケア）レポート"],
    [["ひと月あたり症例数", "療養"],"69_3_ひと月あたり症例数（療養）レポート"],
    [["ひと月あたり延べ在院日数", "回復リハ"],"70_1_ひと月あたり延べ在院日数（回復リハ）レポート"],
    [["ひと月あたり延べ在院日数", "地域包括ケア"],"70_2_ひと月あたり延べ在院日数（地域包括ケア）レポート"],
    [["ひと月あたり延べ在院日数", "療養"],"70_3_ひと月あたり延べ在院日数（療養）レポート"],
    [["平均在院日数", "回復リハ"],"71_1_平均在院日数（回復リハ）レポート"],
    [["平均在院日数", "地域包括ケア"],"71_2_平均在院日数（地域包括ケア）レポート"],
    [["平均在院日数", "療養"],"71_3_平均在院日数（療養）レポート"],
    [["ADLスコアの改善率", "回復リハ"],"72_1_ADLスコアの改善率（回復リハ）レポート"],
    [["ADLスコアの改善率", "地域包括ケア"],"72_2_ADLスコアの改善率（地域包括ケア）レポート"],
    [["ADLスコアの改善率", "療養"],"72_3_ADLスコアの改善率（療養）レポート"],
    [["FIM得点の改善率", "回復リハ"],"73_1_FIM得点の改善率（回復リハ）レポート"],
    [["医療区分の改善率", "療養"],"74_1_医療区分の改善率（療養）レポート"],
    [["疾患別リハ単位数_運動器", "回復リハ"],"75_1_疾患別リハ単位数_運動器（回復リハ）レポート"],
    [["疾患別リハ単位数_呼吸器", "回復リハ"],"75_2_疾患別リハ単位数_呼吸器（回復リハ）レポート"],
    [["疾患別リハ単位数_心大血管疾患", "回復リハ"],"75_3_疾患別リハ単位数_心大血管疾患（回復リハ）レポート"],
    [["疾患別リハ単位数_脳血管疾患", "回復リハ"],"75_4_疾患別リハ単位数_脳血管疾患（回復リハ）レポート"],
    [["疾患別リハ単位数_廃用症候群", "回復リハ"],"75_5_疾患別リハ単位数_廃用症候群（回復リハ）レポート"],
    [["紹介率", "回復リハ"],"76_1_紹介率（回復リハ）レポート"],
    [["紹介率", "地域包括ケア"],"76_2_紹介率（地域包括ケア）レポート"],
    [["紹介率", "療養"],"76_3_紹介率（療養）レポート"],
    [["在宅復帰率", "回復リハ"],"77_1_在宅復帰率（回復リハ）レポート"],
    [["在宅復帰率", "地域包括ケア"],"77_2_在宅復帰率（地域包括ケア）レポート"],
    [["在宅復帰率", "療養"],"77_3_在宅復帰率（療養）レポート"],
    [["医療区分別の症例構成割合", "療養"],"78_1_医療区分別の症例構成割合（療養）レポート"],
    [["薬剤管理指導料の算定率", "療養"],"79_1_薬剤管理指導料の算定率（療養）レポート"],
    [["退院時薬剤情報管理指導料の算定率", "療養"],"80_1_退院時薬剤情報管理指導料の算定率（療養）レポート"],
    [["目標設定等支援・管理料の算定率", "回復リハ"],"81_1_目標設定等支援・管理料の算定率（回復リハ）レポート"],
    [["目標設定等支援・管理料の算定率", "療養"],"81_2_目標設定等支援・管理料の算定率（療養）レポート"],
    [["退院時リハビリテーション指導料の算定率", "療養"],"82_1_退院時リハビリテーション指導料の算定率（療養）レポート"],
    [["退院前訪問指導料の算定率", "療養"],"83_1_退院前訪問指導料の算定率（療養）レポート"],
    [["要介護度", "回復リハ"],"84_1_要介護度（回復リハ）レポート"],
    [["要介護度", "地域包括ケア"],"84_2_要介護度（地域包括ケア）レポート"],
    [["要介護度", "療養"],"84_3_要介護度（療養）レポート"],
    [["要介護情報", "胃瘻・腸瘻", "回復リハ"],"85_1_1_要介護情報_胃瘻・腸瘻（回復リハ）レポート"],
    [["要介護情報", "胃瘻・腸瘻", "地域包括ケア"],"85_1_2_要介護情報_胃瘻・腸瘻（地域包括ケア）レポート"],
    [["要介護情報", "胃瘻・腸瘻", "療養"],"85_1_3_要介護情報_胃瘻・腸瘻（療養）レポート"],
    [["要介護情報", "経鼻胃管", "回復リハ"],"85_2_1_要介護情報_経鼻胃管（回復リハ）レポート"],
    [["要介護情報", "経鼻胃管", "地域包括ケア"],"85_2_2_要介護情報_経鼻胃管（地域包括ケア）レポート"],
    [["要介護情報", "経鼻胃管", "療養"],"85_2_3_要介護情報_経鼻胃管（療養）レポート"],
    [["要介護情報", "摂食・嚥下機能障害", "回復リハ"],"85_3_1_要介護情報_摂食・嚥下機能障害（回復リハ）レポート"],
    [["要介護情報", "摂食・嚥下機能障害", "地域包括ケア"],"85_3_2_要介護情報_摂食・嚥下機能障害（地域包括ケア）レポート"],
    [["要介護情報", "摂食・嚥下機能障害", "療養"],"85_3_3_要介護情報_摂食・嚥下機能障害（療養）レポート"],
    [["要介護情報", "中心静脈栄養", "回復リハ"],"85_4_1_要介護情報_中心静脈栄養（回復リハ）レポート"],
    [["要介護情報", "中心静脈栄養", "地域包括ケア"],"85_4_2_要介護情報_中心静脈栄養（地域包括ケア）レポート"],
    [["要介護情報", "中心静脈栄養", "療養"],"85_4_3_要介護情報_中心静脈栄養（療養）レポート"],
    [["要介護情報", "低栄養", "回復リハ"],"85_5_1_要介護情報_低栄養（回復リハ）レポート"],
    [["要介護情報", "低栄養", "地域包括ケア"],"85_5_2_要介護情報_低栄養（地域包括ケア）レポート"],
    [["要介護情報", "低栄養", "療養"],"85_5_3_要介護情報_低栄養（療養）レポート"],
    [["要介護情報", "末梢静脈栄養", "回復リハ"],"85_6_1_要介護情報_末梢静脈栄養（回復リハ）レポート"],
    [["要介護情報", "末梢静脈栄養", "地域包括ケア"],"85_6_2_要介護情報_末梢静脈栄養（地域包括ケア）レポート"],
    [["要介護情報", "末梢静脈栄養", "療養"],"85_6_3_要介護情報_末梢静脈栄養（療養）レポート"],
]

#出力フォルダを設定する ※各レポートで個別実装
#引数：target_year：対象年度、target_group：対象グループ、target_period：対象期間
#戻り値：出力先フォルダのパス
#ERROR:-1
def set_output_folder(target_year,target_group, target_report_type):
    
    try:
        #出力先に当該年度のフォルダがあるか確認する
        output_folder = CONST.OUTPUT_FOLDER
        output_folder = os.path.join(output_folder, str(target_year))
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        #output_folder内に「03_慢性期指標」フォルダがあるか確認する
        chronic_folder = os.path.join(output_folder, "03_慢性期指標")
        if not os.path.exists(chronic_folder):
            os.makedirs(chronic_folder)
    
        #chronic_folder内に当該グループのフォルダがあるか確認する
        group_folder = os.path.join(chronic_folder, target_group)
        if not os.path.exists(group_folder):
            os.makedirs(group_folder)
    

        #period_folder内に当該レポート種類のフォルダがあるか確認する
        report_type_folder = os.path.join(group_folder, target_report_type)
        if not os.path.exists(report_type_folder):
            os.makedirs(report_type_folder)
    except:
        print("エラー: 出力先フォルダの設定に失敗しました。年度={}、グループ={}、レポート種類={}".format(target_year,target_group,target_report_type))
        return -1
      
    #report_type_folder内のファイル、フォルダを全て削除
    delete_file_count = 0
    for file in os.listdir(report_type_folder):
        try:
            os.remove(os.path.join(report_type_folder, file))
            #削除したファイル数をカウント
            delete_file_count += 1
        except:
            print("フォルダ：{}".format(report_type_folder))
            print(f"エラー: {file} の削除に失敗しました。")
            print("エラー内容：{}".format(sys.exc_info()[1]))
            return -1

    print("削除したファイル数：{}".format(delete_file_count))
    return report_type_folder


def create_cover_page(target_year,target_group, target_report_type, target_create_date, output_folder):
    print("\n表紙ファイルの作成を開始します。")
    try:
        #表紙ファイルを開く
        over_page_path = CONST.COVER_PAGE_FORMAT_PATH_CHRONIC
        #openpyxl で開く
        workbook = openpyxl.load_workbook(over_page_path)
        if workbook is None:
            print("\nエラー: 表紙ファイルの読み込みに失敗しました。年度={}、グループ={}、レポート種類={}".format(target_year,target_group,target_report_type))
            print("ファイルパス：{}".format(over_page_path))
            return -1

        #シート名「00-1_表紙」を開く
        sheet = workbook["00-1_表紙"]
        #セルCD1に指標タイプを入力
        sheet["CD1"] = "慢性期指標"
        #セルCD2に年度を入力
        sheet["CD2"] = target_year
        #セルCD3にグループを入力
        if(target_group == "01_急性期グループ"):
            sheet["CD3"] = "急性期グループ"
        elif(target_group == "02_慢性期グループ"):
            sheet["CD3"] = "慢性期グループ"
        else:
            print("\nエラー: グループが不正です。年度={}、グループ={}、レポート種類={}".format(target_year,target_group,target_report_type))
            return -1

        #セルCD5に作成日を入力
        sheet["CD5"] = target_create_date
        #セルA4の色を変更
        sheet["A4"].fill = PatternFill(start_color=CONST.REPORT_INDCTR_COLOR[2], end_color=CONST.REPORT_INDCTR_COLOR[2], fill_type="solid")
        #セルA22の色を変更
        color_index = 0
        if(target_group == "01_急性期グループ"):
            color_index = 0
        elif(target_group == "02_慢性期グループ"):
            color_index = 1
        sheet["A22"].fill = PatternFill(start_color=CONST.REPORT_GROUP_COLOR[color_index], end_color=CONST.REPORT_GROUP_COLOR[color_index], fill_type="solid")
    
        output_folder = os.path.join(output_folder, "00-1_表紙.xlsx")
        workbook.save(output_folder)
        if workbook is None:
            print("\nエラー: 表紙ファイルの保存に失敗しました。年度={}、グループ={}、レポート種類={}".format(target_year,target_group,target_report_type))
            print("ファイルパス：{}".format(output_folder))
            return -1

        #PDFに変換
        # Excel ファイルを開く
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Excel のウィンドウを非表示にする
        wb = excel.Workbooks.Open(output_folder)
        newPath = str(output_folder).replace(".xlsx",".pdf")
        wb.ExportAsFixedFormat(0, newPath)  # 0 は PDF 形式
        wb.Close()
        excel.Quit()

        #pdfファイルの作成が完了したら元のエクセルファイルを削除
        os.remove(output_folder)

        print(f"\n表紙ファイルを作成しました。ファイルパス: {output_folder}\n")
        return 0
    except:
        print("\nエラー: 表紙ファイルの作成に失敗しました。年度={}、グループ={}、レポート種類={}".format(target_year,target_group,target_report_type))
        print("ファイルパス：{}".format(output_folder))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1

#NIT納品レポートを作業用フォルダにコピー
#引数：target_year：対象年度、target_group：対象グループ、target_period：対象期間、target_report_type：レポート種類
#ERROR:-1
def copy_source_report(target_year,target_group, target_report_type, output_folder):
    print("\nNIT納品レポートのコピーを開始します。")

    TEMP =""
    if(target_report_type == "01_病院名"):
        TEMP = "01_PDF"
    elif(target_report_type == "02_病院番号"):
        TEMP = "02_EXCEL"
    else:
        print("\nエラー: レポート種類が不正です。年度={}、グループ={}、レポート種類={}".format(target_year,target_group,target_report_type))
        return -1

    #NIT納品レポートのフォルダを取得
    nit_report_folder = CONST.NIT_REPORT_FOLDER
    nit_report_folder = os.path.join(nit_report_folder, str(target_year))
    nit_report_folder = os.path.join(nit_report_folder, "03_慢性期指標")
    nit_report_folder = os.path.join(nit_report_folder, target_group)
    nit_report_folder = os.path.join(nit_report_folder, TEMP)

    extension = ".pdf"
    if(target_report_type == "02_病院番号"):
        extension = ".xlsx"

    #nit_report_folderが存在しない場合はエラー  
    if not os.path.exists(nit_report_folder):
        print("\nエラー: NIT納品レポートのフォルダが存在しません。年度={}、グループ={}、レポート種類={}".format(target_year,target_group,target_report_type))
        print("フォルダパス：{}".format(nit_report_folder))
        return -1
    #nit_report_folder内にファイルが存在しない場合はエラー
    if not os.listdir(nit_report_folder):
        print("\nエラー: NIT納品レポートのフォルダにファイルが存在しません。年度={}、グループ={}、レポート種類={}".format(target_year,target_group,target_report_type))
        print("フォルダパス：{}".format(nit_report_folder))
        return -1

    try:
        cnt = 1
        #TRANS_FILE_NAMES内のファイルをコピー
        #ファイル名に含まれている特定の単語でファイル名を変換
        for file in os.listdir(nit_report_folder):
            #拡張子を除いたファイル名を取得
            file_name = os.path.splitext(file)[0]            

            for item in LIST_KEY_FILENAMES:
                is_match = False
                keys = ""
                for key in item[0]:    
                    keys = keys + ", " + key
                    if key in file_name:
                        is_match = True
                    else:
                        is_match = False
                        break
                
                if is_match:
                    #print(" " + file_name + " keys  " + keys )
                    new_file_name = item[1] + extension
                    new_file_name = os.path.join(output_folder, new_file_name)
                    shutil.copy(os.path.join(nit_report_folder, file), new_file_name)
                    cnt += 1
                    break
                    
                    
            if is_match == False:
                print("\nエラー: レポートファイルが名称が一致しません。年度={}、グループ={}、レポート種類={}".format(target_year,target_group,target_report_type))
                print("ファイル名：{}".format(file))
                return -1


        print("\nNIT納品レポートのコピーに成功しました。年度={}、グループ={}、レポート種類={}".format(target_year,target_group,target_report_type))
        print("フォルダパス：{}".format(nit_report_folder))
        print("出力フォルダパス：{}".format(output_folder))
        #コピーしたファイルの数を出力
        print("コピーしたファイルの数：{}".format(str(cnt)))
        return 0

    except:
        print("\nエラー: NIT納品レポートのコピーに失敗しました。年度={}、グループ={}、レポート種類={}".format(target_year,target_group,target_report_type))
        print("フォルダパス：{}".format(nit_report_folder))
        print("出力フォルダパス：{}".format(output_folder))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1


def copy_TEIGI_FILES(target_year, output_folder):
    print("\n指標定義ファイルのコピーを開始します。")
    try:
        TEIGI_FILES_PATH = CONST.TEIGI_FILES_CHRONIC

        #TEIGI_FILES_PATH内のpdfファイルをoutput_folderにコピー
        for file in os.listdir(TEIGI_FILES_PATH):
            if file.endswith(".pdf"):
                shutil.copy(os.path.join(TEIGI_FILES_PATH, file), os.path.join(output_folder, file))

        return 0
    except:
        print("\nエラー: 指標定義ファイルのコピーに失敗しました。年度={}".format(target_year))
        print("ファイルパス：{}".format(TEIGI_FILES_PATH))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1


#def combine_report(target_year, target_group, target_report_type, output_folder):
    print("\nレポートのPDF結合を開始します。")
    try:
        #output_folder内のpdfファイルを結合
        pdf_files = [f for f in os.listdir(output_folder) if f.endswith(".pdf")]
        pdf_files.sort()
        if not pdf_files:
            print("\nエラー: レポートファイルが見つかりません。年度={}、グループ={}、レポート種類={}、フォルダパス={}".format(target_year,target_group,target_report_type,output_folder))
            return -1

        report_file_name = str(target_year) + "年度_慢性期指標レポート_" + target_group + "_" + target_report_type + ".pdf"
        #出力先はoutput_folderの2階層上のフォルダに作成
        reportoutput_folder = os.path.join(os.path.dirname(output_folder), "..")


        #PDFファイルを結合、ファイル名をしおりに追加
        writer = PdfWriter()
        current_page = 0

        for file in pdf_files:
            reader = PdfReader(os.path.join(output_folder, file))
            num_pages = len(reader.pages)

            # 各ページをwriterに追加
            for page in reader.pages:
                writer.add_page(page)

            # ファイル名をしおりとして追加（拡張子なしで追加もOK）
            title = os.path.splitext(os.path.basename(file))[0]
            title = title.replace("_0", "")
            if "レポート" in title:
                current_page += num_pages
                continue
            writer.add_outline_item(title=title, page_number=current_page)
            # 次のしおりのページ位置を把握
            current_page += num_pages

            
            
        # 保存
        with open(os.path.join(reportoutput_folder, report_file_name), "wb") as f:
            writer.write(f)

        print("\nレポートのPDF結合を完了しました。:{}".format(report_file_name))

        #pdf_filesを結合
        #pdf_merger = PdfMerger()
        #for file in pdf_files:
        #    pdf_merger.append(os.path.join(output_folder, file))
        #pdf_merger.write(os.path.join(reportoutput_folder, report_file_name))

    except:
        print("\nエラー: レポートの結合に失敗しました。年度={}、レポート種類={}、フォルダパス={}".format(target_year,target_report_type,output_folder))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1

def combine_report_pypdf(target_year, target_group, target_report_type, output_folder):
    print("\nレポートのPDF結合を開始します。")
    try:
        # output_folder内のPDFファイルを取得・ソート
        pdf_files = sorted(f for f in os.listdir(output_folder) if f.lower().endswith(".pdf"))
        if not pdf_files:
            print(f"\nエラー: レポートファイルが見つかりません。年度={target_year}、グループ={target_group}、レポート種類={target_report_type}、フォルダパス={output_folder}")
            return -1

        report_file_name = f"{target_year}年度_慢性期指標レポート_{target_group}_{target_report_type}.pdf"
        reportoutput_folder = os.path.abspath(os.path.join(output_folder, "..", ".."))
        os.makedirs(reportoutput_folder, exist_ok=True)  # 念のためフォルダがなければ作成

        writer = PdfWriter()
        current_page = 0

        for file in pdf_files:
            full_path = os.path.join(output_folder, file)
            reader = PdfReader(full_path)
            num_pages = len(reader.pages)

            for page in reader.pages:
                writer.add_page(page)

            title = os.path.splitext(file)[0].replace("_0", "")
            if "レポート" in title:
                current_page += num_pages
                continue

            writer.add_outline_item(title=title, page_number=current_page)
            current_page += num_pages

        output_path = os.path.join(reportoutput_folder, report_file_name)
        with open(output_path, "wb") as f:
            writer.write(f)

        print(f"\nレポートのPDF結合を完了しました。: {report_file_name}")
        return 0

    except Exception as e:
        print(f"\nエラー: レポートの結合に失敗しました。年度={target_year}、レポート種類={target_report_type}、フォルダパス={output_folder}")
        print("エラー内容：", e)
        return -1


def tranaHospNameToPublicNO(target_year, output_folder, dict_hosp_DB):
    print("\n病院名を病院番号に変換を開始します。")
    #output_folder内のエクセルファイルを取得
    excel_files = [f for f in os.listdir(output_folder) if f.endswith(".xlsx")]
    if not excel_files:
        print("\nエラー: エクセルファイルが見つかりません。年度={}、フォルダパス={}".format(target_year,output_folder))
        return -1
    try:
        for file in excel_files:
            print("file:{}".format(file))
            startRow = 5
            wb = load_workbook(os.path.join(output_folder, file))
            ws = wb["帳票"]
            col_data = [cell.value for cell in ws['B']]
            for i in range(startRow, len(col_data) + 1, 1):
                if(ws.cell(row=i, column=2).value == "合計" or ws.cell(row=i, column=2).value == None):
                    break
                pNo = dict_hosp_DB[ws.cell(row=i, column=2).value]         
                ws.cell(row=i, column=2, value=str(pNo)) 
        
            wb.save(os.path.join(output_folder, file))  # 保存
            wb.close()

            print("病院名を番号変換を完了しました。:{}".format(file))
            print("PDFに変換を開始します。")
            #PDFに変換
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False  # Excel のウィンドウを非表示にする
            wb = excel.Workbooks.Open(os.path.join(output_folder, file))
            # 指定シートを取得
            ws = wb.Sheets("帳票")
            newPath = str(os.path.join(output_folder, file.replace(".xlsx",".pdf")))

            ws.ExportAsFixedFormat(0, newPath)  # 0 は PDF 形式
            wb.Close()
            excel.Quit()

            print("PDFに変換を完了しました。:{}".format(file))
        return 0
    except Exception as e:
        print(f"\n  病院名を番号変換時にエラーが発生しました。: {e}")
        print("  エラー内容：{}".format(sys.exc_info()[1]))
        print("  dict_hosp_DB:{}".format(dict_hosp_DB))
        return -1

















def create_report():
    print("\nアンケート調査レポートを作成します。")

    print("\n年度（西暦4桁、半角）を入力してください。")
    user_input = input(":>>")
    if not user_input.isdigit() or len(user_input) != 4:
        print("\nエラー: 西暦4桁の半角数字を入力してください。")
        return -1

    target_year = int(user_input)
    print(f"\n対象年度: {str(target_year)}\n")

    print("\nグループ（急性期グループ、慢性期グループ）を選んでください。")
    print("1:急性期グループ、2:慢性期グループ")
    user_input = input(":>>")
    if user_input not in ["1", "2"]:
        print("\nエラー: 急性期グループ、慢性期グループのいずれかを入力してください。")
        return -1

    target_group = CONST.GROUPS[int(user_input)-1]
    print(f"\n対象グループ: {target_group}\n")

    print("\nレポートの種類を選んでください。")
    print("1:病院名、2:病院番号")
    user_input = input(":>>")
    if user_input not in ["1", "2"]:
        print("\nエラー: 病院名、病院番号のいずれかを入力してください。")
        return -1
    
    target_report_type = CONST.REPORT_TYPES[int(user_input) -1]
    print(f"\n対象レポート種類: {target_report_type}\n")

    print("\nレポートの作成日をyyyy/m/d形式で入力してください。")
    user_input = input(":>>")
    if not re.match(r'^\d{4}/\d{1,2}/\d{1,2}$', user_input):
        print("\nエラー: 日付の形式が不正です。yyyy/m/d形式で入力してください。")
        return -1

    target_create_date = datetime.strptime(user_input, "%Y/%m/%d")
    print(f"\n対象作成日: {target_create_date.strftime('%Y/%m/%d')}\n")

    print("\n参加病院一覧ファイルを指定してください。")
    #ファイルダイアログを開く
    root = tk.Tk()
    #root.withdraw()
    list_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not list_path:
        print("\nエラー: 参加病院一覧ファイルが指定されていません。")
        return -1
    

    #当該年度の平均在院日数（疾患別）から1Qのデータ提出病院のHospital_HOSPITAL_IDを取得し、
    #HOSPITALテーブルから病院ID、病院名を取得する
    print("target_year={}年度第1四半期「平均在院日数（疾患別）」テーブルからデータ提出病院のID、病院名を取得".format(target_year))
    hospital_data = COMMON.getHospIDAndName(target_year)
    if hospital_data is None or hospital_data == -1:
        print("\nエラー: 病院データの取得に失敗しました。年度={}、".format(target_year))
        return -1   
    #cnt = 1
    #for data in hospital_data:
    #    print("No={} data: name={}: no={}".format(cnt,data[1],data[0]))
    #    cnt += 1

    # hospital_dataから「参加病院全体」が含まれている要素を削除する
    hospital_data = [data for data in hospital_data if "参加病院全体" not in data[1]]
    print("\n フィルタリング後の病院データ")
    print("病院データ 取得件数={}、年度={}\n".format(str(len(hospital_data)),target_year, ))

    # 辞書型
    # key:病院名、value:公開用病院番号
    dict_public_no = COMMON.createDictPublicNo(target_year, hospital_data)
    if dict_public_no is None or dict_public_no == -1:
        print("\nエラー: 公開用病院番号の作成に失敗しました。年度={}".format(target_year))
        return -1

    #出力先フォルダの設定
    output_folder = set_output_folder(target_year, target_group, target_report_type)
    if output_folder is None or output_folder == -1:
        return -1

    #表紙を作成
    if(create_cover_page(target_year,target_group,target_report_type, target_create_date, output_folder) == -1):
        return -1

    #病院一覧を作成
    if(COMMON.create_hospitalList(target_year, list_path, output_folder) == -1):
        return -1

    #NIT納品レポートを作業用フォルダにコピー
    if(copy_source_report(target_year,target_group,target_report_type, output_folder) == -1):
        return -1

    #指標定義ファイルを作業用フォルダにコピー
    if(copy_TEIGI_FILES(target_year, output_folder) == -1):
        return -1


    if(target_report_type == CONST.REPORT_TYPES[1]):
        if(tranaHospNameToPublicNO(target_year, output_folder, dict_public_no) == -1):
            return -1

    #レポートを結合
    if(combine_report_pypdf(target_year, target_group, target_report_type, output_folder) == -1):
        return -1