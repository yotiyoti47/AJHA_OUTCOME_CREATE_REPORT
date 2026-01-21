import re
from datetime import datetime
import os
import sys
import X_00_CONST as CONST
import openpyxl
import win32com.client
from openpyxl.styles import PatternFill
import shutil
#from PyPDF2 import PdfReader, PdfWriter
from pypdf import PdfReader, PdfWriter


DICT_TRANSITION_REPORT_NAMES = {
"肺血栓塞栓症の予防対策実施率": "1-1_015_肺血栓塞栓症の予防対策実施率レポート.pdf",
"肺血栓塞栓症の発生率": "1-2_016_肺血栓塞栓症の発生率レポート.pdf",
"褥瘡の持込率": "1-3_017_褥瘡の持込率レポート.pdf",
"褥瘡の発生率": "1-4_018_褥瘡の発生率レポート.pdf",
"65歳以上の患者における認知症の保有率": "1-5_019_65歳以上の患者における認知症の保有率レポート.pdf",
"急性心筋梗塞における主要手術・治療実施率": "2-1_020-023_急性心筋梗塞における主要手術・治療実施率レポート.pdf",
"肺炎に対する入院当日の抗生物質使用率": "2-2_024_肺炎に対する入院当日の抗生物質使用率レポート.pdf",
"出血性胃・十二指腸潰瘍内視鏡的治療の施行率": "2-3_025_出血性胃・十二指腸潰瘍内視鏡的治療の施行率レポート.pdf",
"急性脳梗塞患者リハビリテーション開始率": "2-4_026_急性脳梗塞患者リハビリテーション開始率レポート.pdf",
"胃がんの患者に対するESDの施行率": "2-5_027_胃がんの患者に対するESDの施行率レポート.pdf",
"胆嚢切除術患者に対する腹腔鏡下手術施行率": "2-6_028_胆嚢切除術患者に対する腹腔鏡下手術施行率レポート.pdf",
"虫垂炎の患者に対する手術施行率": "2-7_029_虫垂炎の患者に対する手術施行率レポート.pdf",
"虫垂切除術患者に対する腹腔鏡下手術施行率": "2-8_030_虫垂切除術患者に対する腹腔鏡下手術施行率レポート.pdf",
"帝王切開における全身麻酔施行率": "2-9_031_帝王切開における全身麻酔施行率レポート.pdf",
"帝王切開における輸血施行率": "2-10_032_帝王切開における輸血施行率レポート.pdf",
"脳卒中地域連携パスの使用率": "3-1_033_脳卒中地域連携パスの使用率レポート.pdf",
"脳卒中地域連携パスの使用率(転院先)": "3-2_034_脳卒中地域連携パスの使用率(転院先)レポート.pdf",
"大腿骨地域連携パスの使用率": "3-3_035_大腿骨地域連携パスの使用率レポート.pdf",
"大腿骨地域連携パスの使用率(転院先)": "3-4_036_大腿骨地域連携パスの使用率(転院先)レポート.pdf",
"急性期病棟における退院調整の実施率": "3-5_037_急性期病棟における退院調整の実施率レポート.pdf",
"介護施設・福祉施設からの入院率": "3-6_038_介護施設・福祉施設からの入院率レポート.pdf",
"他の病院・診療所の病棟からの転院率": "3-7_039_他の病院・診療所の病棟からの転院率レポート.pdf",
"介護施設・福祉施設等への退院率": "3-8_040-041_介護施設・福祉施設等への退院率レポート.pdf",
"他の病院・診療所の病棟への転院率": "3-9_042_他の病院・診療所の病棟への転院率レポート.pdf",
"自宅退院患者における在宅医療を受ける率": "3-10_043_自宅退院患者における在宅医療を受ける率レポート.pdf",
}


#出力フォルダを設定する ※各レポートで個別実装
#引数：target_year：対象年度、target_group：対象グループ、target_period：対象期間
#戻り値：出力先フォルダのパス
#ERROR:-1
def set_output_folder(target_year,target_number):
    
    try:
        #出力先に当該年度のフォルダがあるか確認する
        output_folder = CONST.OUTPUT_FOLDER
        output_folder = os.path.join(output_folder, str(target_year))
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        #output_folder内に「05_経年推移レポート」フォルダがあるか確認する
        transition_folder = os.path.join(output_folder, "04_経年推移レポート")
        if not os.path.exists(transition_folder):
            os.makedirs(transition_folder)
    
        #chronic_folder内に当該グループのフォルダがあるか確認する
        number_folder = os.path.join(transition_folder, str(target_number))
        if not os.path.exists(number_folder):
            os.makedirs(number_folder)

    except:
        print("エラー: 出力先フォルダの設定に失敗しました。年度={}、納品回数={}".format(target_year,target_number))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1
      
    #report_type_folder内のファイル、フォルダを全て削除
    delete_file_count = 0
    for file in os.listdir(number_folder):
        try:
            os.remove(os.path.join(number_folder, file))
            #削除したファイル数をカウント
            delete_file_count += 1
        except:
            print("フォルダ：{}".format(number_folder))
            print(f"エラー: {file} の削除に失敗しました。")
            print("エラー内容：{}".format(sys.exc_info()[1]))
            return -1

    print("削除したファイル数：{}".format(delete_file_count))
    return number_folder


def create_cover_page(target_year,target_number, target_create_date, output_folder):
    print("\n表紙ファイルの作成を開始します。")
    try:
        #表紙ファイルを開く
        over_page_path = CONST.COVER_PAGE_FORMAT_PATH_TRANSITION
        #openpyxl で開く
        workbook = openpyxl.load_workbook(over_page_path)
        if workbook is None:
            print("\nエラー: 表紙ファイルの読み込みに失敗しました。年度={}、納品回数={}".format(target_year,target_number))
            print("ファイルパス：{}".format(over_page_path))
            return -1

        #シート名「00-1_表紙」を開く
        sheet = workbook["00-1_表紙"]
        #セルCD1に指標タイプを入力
        sheet["CD1"] = "経年推移レポート"
        #セルCD2に年度を入力
        sheet["CD2"] = target_year

        #セルCD5に作成日を入力
        sheet["CD5"] = target_create_date
        #セルA4の色を変更
        sheet["A4"].fill = PatternFill(start_color=CONST.REPORT_INDCTR_COLOR[3], end_color=CONST.REPORT_INDCTR_COLOR[3], fill_type="solid")
    
        output_folder = os.path.join(output_folder, "00-1_表紙.xlsx")
        workbook.save(output_folder)
        if workbook is None:
            print("\nエラー: 表紙ファイルの保存に失敗しました。年度={}、納品回数={}".format(target_year,target_number))
            print("ファイルパス：{}".format(output_folder))
            return -1

        #PDFに変換
        # Excel ファイルを開く
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Excel のウィンドウを非表示にする
        wb = excel.Workbooks.Open(output_folder)
        newPath = str(output_folder).replace(".xlsx",".pdf")
        wb.ExportAsFixedFormat(0, newPath)  # 0 は PDF 形式
        wb.Close()
        excel.Quit()

        #pdfファイルの作成が完了したら元のエクセルファイルを削除
        os.remove(output_folder)

        print(f"\n表紙ファイルを作成しました。ファイルパス: {output_folder}\n")
        return 0
    except:
        print("\nエラー: 表紙ファイルの作成に失敗しました。年度={}、納品回数={}".format(target_year,target_number))
        print("ファイルパス：{}".format(output_folder))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1

    #NIT納品レポートを作業用フォルダにコピー
    if(copy_source_report(target_year,target_group,target_report_type, output_folder) == -1):
        return -1


#NIT納品レポートを作業用フォルダにコピー
#引数：target_year：対象年度、target_group：対象グループ、target_period：対象期間、target_report_type：レポート種類
#ERROR:-1
def copy_source_report(target_year,target_number, output_folder):
    print("\nNIT納品レポートのコピーを開始します。")


    #NIT納品レポートのフォルダを取得
    nit_report_folder = CONST.NIT_REPORT_FOLDER
    nit_report_folder = os.path.join(nit_report_folder, str(target_year))
    nit_report_folder = os.path.join(nit_report_folder, "04_経年推移レポート")
    nit_report_folder = os.path.join(nit_report_folder, str(target_number))

    for folder_name in CONST.TRANSITION_REPORT:
        temp_nit_report_folder = os.path.join(nit_report_folder, folder_name)
        print("フォルダパス：{}".format(temp_nit_report_folder))
        if not os.path.exists(temp_nit_report_folder):
            print("\nエラー: NIT納品レポートのフォルダが存在しません。年度={}、納品回数={}".format(target_year,target_number))
            print("フォルダパス：{}".format(temp_nit_report_folder))
            return -1

        #nit_report_folderにファイルが存在しない場合はエラー
        if not os.listdir(temp_nit_report_folder):
            print("\nエラー: NIT納品レポートのフォルダにファイルが存在しません。年度={}、納品回数={}".format(target_year,target_number))
            print("フォルダパス：{}".format(temp_nit_report_folder))
            return -1
        
        try:
            for file in os.listdir(temp_nit_report_folder):
                file_name = os.path.splitext(file)[0]
                #print("  file_name:" + file_name)
                if file_name in DICT_TRANSITION_REPORT_NAMES:
                    new_file_name = DICT_TRANSITION_REPORT_NAMES[file_name]
                    new_file_name = os.path.join(output_folder, new_file_name)
                    shutil.copy(os.path.join(temp_nit_report_folder, file), new_file_name)

                else:
                    print("\nエラー: レポートファイルが名称が一致しません。年度={}、納品回数={}".format(target_year,target_number))
                    print("ファイル名：{}".format(file))
                    return -1
        except:
            print("\nエラー: NIT納品レポートのコピーに失敗しました。年度={}、納品回数={}".format(target_year,target_number))
            print("フォルダパス：{}".format(nit_report_folder))
            print("エラー内容：{}".format(sys.exc_info()[1]))
            return -1
        
    print("\nNIT納品レポートのコピーが完了しました。年度={}、納品回数={}".format(target_year,target_number))
    return 0    


def copy_TEIGI_FILES(target_year, output_folder):
    print("\n指標定義ファイルのコピーを開始します。")
    try:
        TEIGI_FILES_PATH = CONST.TEIGI_FILES_TRANSITION

        #TEIGI_FILES_PATH内のpdfファイルをoutput_folderにコピー
        for file in os.listdir(TEIGI_FILES_PATH):
            if file.endswith(".pdf"):
                shutil.copy(os.path.join(TEIGI_FILES_PATH, file), os.path.join(output_folder, file))

        return 0
    except:
        print("\nエラー: 指標定義ファイルのコピーに失敗しました。年度={}".format(target_year))
        print("ファイルパス：{}".format(TEIGI_FILES_PATH))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1

#def combine_report(target_year, target_number, output_folder):
    print("\nレポートのPDF結合を開始します。")
    try:
        #output_folder内のpdfファイルを結合
        pdf_files = [f for f in os.listdir(output_folder) if f.endswith(".pdf")]
        pdf_files.sort()
        if not pdf_files:
            print("\nエラー: レポートファイルが見つかりません。年度={}、納品回数={}、フォルダパス={}".format(target_year,target_number,output_folder))
            return -1

        report_file_name = str(target_year) + "年度_経年推移レポート_" + str(target_number) + ".pdf"
        #出力先はoutput_folderの親フォルダに作成
        reportoutput_folder = os.path.dirname(output_folder)


        #PDFファイルを結合、ファイル名をしおりに追加
        writer = PdfWriter()
        current_page = 0

        for file in pdf_files:
            reader = PdfReader(os.path.join(output_folder, file))
            num_pages = len(reader.pages)

            # 各ページをwriterに追加
            for page in reader.pages:
                writer.add_page(page)

            # ファイル名をしおりとして追加（拡張子なしで追加もOK）
            title = os.path.splitext(os.path.basename(file))[0]
            
            if "1-0_000_" in title:
                title = title.replace("1-0_000_", "")
            elif "2-0_000_" in title:
                title = title.replace("2-0_000_", "")
            elif "3-0_000_" in title:
                title = title.replace("3-0_000_", "")

            if "レポート" in title:
                current_page += num_pages
                continue
            writer.add_outline_item(title=title, page_number=current_page)
            # 次のしおりのページ位置を把握
            current_page += num_pages
            
        # 保存
        with open(os.path.join(reportoutput_folder, report_file_name), "wb") as f:
            writer.write(f)

        print("\nレポートのPDF結合を完了しました。:{}".format(report_file_name))

        #pdf_filesを結合
        #pdf_merger = PdfMerger()
        #for file in pdf_files:
        #    pdf_merger.append(os.path.join(output_folder, file))
        #pdf_merger.write(os.path.join(reportoutput_folder, report_file_name))

    except:
        print("\nエラー: レポートの結合に失敗しました。年度={}、納品回数={}、フォルダパス={}".format(target_year,target_number,output_folder))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1

def combine_report_pypdf(target_year, target_number, output_folder):
    print("\nレポートのPDF結合を開始します。")
    try:
        pdf_files = sorted(f for f in os.listdir(output_folder) if f.lower().endswith(".pdf"))
        if not pdf_files:
            print(f"\nエラー: レポートファイルが見つかりません。年度={target_year}、納品回数={target_number}、フォルダパス={output_folder}")
            return -1

        report_file_name = f"{target_year}年度_経年推移レポート_{target_number}.pdf"
        reportoutput_folder = os.path.dirname(output_folder)
        os.makedirs(reportoutput_folder, exist_ok=True)

        writer = PdfWriter()
        current_page = 0

        for file in pdf_files:
            full_path = os.path.join(output_folder, file)
            reader = PdfReader(full_path)
            num_pages = len(reader.pages)

            for page in reader.pages:
                writer.add_page(page)

            title = os.path.splitext(file)[0]
            for prefix in ["1-0_000_", "2-0_000_", "3-0_000_"]:
                if title.startswith(prefix):
                    title = title.replace(prefix, "", 1)
                    break

            if "レポート" in title:
                current_page += num_pages
                continue

            writer.add_outline_item(title=title, page_number=current_page)
            current_page += num_pages

        output_path = os.path.join(reportoutput_folder, report_file_name)
        with open(output_path, "wb") as f:
            writer.write(f)

        print(f"\nレポートのPDF結合を完了しました。: {report_file_name}")
        return 0

    except Exception as e:
        print(f"\nエラー: レポートの結合に失敗しました。年度={target_year}、納品回数={target_number}、フォルダパス={output_folder}")
        print(f"エラー内容：{e}")
        return -1


def create_report():
    print("\nアンケート調査レポートを作成します。")

    print("\n年度（西暦4桁、半角）を入力してください。")
    user_input = input(":>>")
    if not user_input.isdigit() or len(user_input) != 4:
        print("\nエラー: 西暦4桁の半角数字を入力してください。")
        return -1

    target_year = int(user_input)
    print(f"\n対象年度: {str(target_year)}\n")

    print("\n納品回数を選んでください。")
    print("1:1回目、2:2回目、3:3回目")
    user_input = input(":>>")
    if user_input not in ["1", "2", "3"]:
        print("\nエラー: 1回目、2回目、3回目のいずれかを入力してください。")
        return -1

    target_number = int(user_input)
    print(f"\n対象納品回数: {str(target_number)}\n")

    print("\nレポートの作成日をyyyy/m/d形式で入力してください。")
    user_input = input(":>>")
    if not re.match(r'^\d{4}/\d{1,2}/\d{1,2}$', user_input):
        print("\nエラー: 日付の形式が不正です。yyyy/m/d形式で入力してください。")
        return -1

    target_create_date = datetime.strptime(user_input, "%Y/%m/%d")
    print(f"\n対象作成日: {target_create_date.strftime('%Y/%m/%d')}\n")

    #出力先フォルダの設定
    output_folder = set_output_folder(target_year, target_number)
    if output_folder is None or output_folder == -1:
        return -1

    #表紙を作成
    if(create_cover_page(target_year,target_number, target_create_date, output_folder) == -1):
        return -1

    #NIT納品レポートを作業用フォルダにコピー
    if(copy_source_report(target_year,target_number, output_folder) == -1):
        return -1

    #指標定義ファイルを作業用フォルダにコピー
    if(copy_TEIGI_FILES(target_year, output_folder) == -1):
        return -1

    #レポートを結合
    if(combine_report_pypdf(target_year, target_number, output_folder) == -1):
        return -1

    return 0