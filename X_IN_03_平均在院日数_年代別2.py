import sqlite3
import openpyxl
import logging
from typing import List, Dict, Optional, Tuple
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

# ログ設定
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def getSQL_疾患別患者数(HOSPITAL_ID: int, _C: str) -> str:
    """疾患別患者数を取得するSQLクエリを生成"""
    return f"""
    SELECT * FROM NUMBER_OF_PATIENTS_BY_DISEASE{_C} 
    WHERE Hospital_HOSPITAL_ID = {HOSPITAL_ID} 
    ORDER BY 年度, 期, Disease_DISEASE_ID
    """

def getSQL_CI_03_平均在院日数(HOSPITAL_ID: int, disease_ID: int, Age_ID: int, _C: str) -> str:
    """平均在院日数を取得するSQLクエリを生成"""
    return f"""
    SELECT 
        CI_03{_C}.年度,
        CI_03{_C}.期,
        CI_03{_C}.平均在院日数
    FROM CI_03{_C}
    WHERE 
        NOT CI_03{_C}.期 = 'TOTAL' AND
        CI_03{_C}.Hospital_HOSPITAL_ID = {HOSPITAL_ID} AND
        CI_03{_C}.Disease_DISEASE_ID = {disease_ID} AND
        CI_03{_C}.Age_AGE_ID = {Age_ID}
    ORDER BY CI_03{_C}.年度, CI_03{_C}.期
    """

def create_key(disease_id: int, age_id: int) -> str:
    """疾患IDと年代IDからキーを生成"""
    return f"{disease_id:02d}{age_id:02d}"

def clear_sheet_data(sheet) -> None:
    """シートのデータをクリア"""
    try:
        for row in sheet.rows:
            for cell in row:
                cell.value = None
    except Exception as e:
        logger.warning(f"シートクリア中にエラー: {e}")

def get_acute_suffix(is_acute: str) -> str:
    """急性期かどうかでサフィックスを決定"""
    return "" if is_acute == "急性期" else "_C"

def format_value_for_display(value: int) -> str:
    """表示用の値をフォーマット"""
    if value == -1:
        return "N/A"
    elif value == -2:
        return "-"
    else:
        return str(value)

def format_value_for_graph(value: int) -> int:
    """グラフ作成用の値をフォーマット"""
    return 0 if value in [-1, -2] else value

def getRepAgeData_年代別(wb, HOSPITAL_ID: int, HOSPITAL_NAME: str, is_acute: str) -> int:
    """平均在院日数の年代別レポートを作成"""
    try:
        logger.info(f"{HOSPITAL_NAME} CI_03 開始")

        _C = get_acute_suffix(is_acute)

        # シートの準備
        if "CI_03" in wb.sheetnames:
            clear_sheet_data(wb["CI_03"])
        else:
            wb.create_sheet("CI_03")
        
        sheet = wb["CI_03"]

        # 疾患別患者数データの取得
        list_疾患別患者数 = X_01.excuteSQL(getSQL_疾患別患者数(HOSPITAL_ID, _C))
        if list_疾患別患者数 is None:
            logger.error("疾患別患者数の取得に失敗しました。")
            return -1

        # 疾患別患者数を辞書に変換
        dict_疾患別患者数 = {}
        for temp疾患別患者数 in list_疾患別患者数:
            key = f"{temp疾患別患者数[1]}{temp疾患別患者数[2]}{temp疾患別患者数[4]}"
            dict_疾患別患者数[key] = temp疾患別患者数[3]

        # 平均在院日数データの取得
        dic_平均在院日数 = {}
        for tempDisease in CONST.MASTA_DISEASE:
            for tempAge in CONST.MASTA_AGE:
                try:
                    temp平均在院日数 = X_01.excuteSQL(
                        getSQL_CI_03_平均在院日数(HOSPITAL_ID, tempDisease[0], tempAge[0], _C)
                    )
                    key = create_key(tempDisease[0], tempAge[0])
                    dic_平均在院日数[key] = temp平均在院日数
                except Exception as e:
                    logger.error(f"データ取得エラー (疾患: {tempDisease[1]}, 年代: {tempAge[1]}): {e}")
                    dic_平均在院日数[key] = []

        # データの書き込み
        targetColCnt = 1
        for tempDisease in CONST.MASTA_DISEASE:
            write_disease_data(sheet, tempDisease, dic_平均在院日数, targetColCnt, HOSPITAL_NAME)
            targetColCnt += len(CONST.MASTA_AGE) + 3

        logger.info(f"{HOSPITAL_NAME} CI_03 終了")
        return 0

    except Exception as e:
        logger.error(f"予期しないエラーが発生しました: {e}")
        return -1

def write_disease_data(sheet, tempDisease: Tuple[int, str], dic_平均在院日数: Dict, 
                      targetColCnt: int, HOSPITAL_NAME: str) -> None:
    """疾患データをシートに書き込み"""
    # 3回繰り返す: 0:グラフ作成用値、1:表示用値、2:比較用値
    for i in range(3):
        if i == 0:
            # ヘッダー情報の書き込み
            sheet.cell(1, targetColCnt).value = HOSPITAL_NAME
            sheet.cell(2, targetColCnt).value = "平均在院日数_疾患別_年代別"
            sheet.cell(3, targetColCnt).value = tempDisease[1]
            sheet.cell(5, targetColCnt).value = "年度"
            sheet.cell(5, targetColCnt + 1).value = "期"

        wroopCnt = 0
        for tempAge in CONST.MASTA_AGE:
            key = create_key(tempDisease[0], tempAge[0])
            tempList = dic_平均在院日数.get(key, [])
            
            # データの書き込み
            rowCnt = 6
            for tempRow in tempList:
                if wroopCnt == 0:
                    sheet.cell(rowCnt, targetColCnt).value = tempRow[0]  # 年度
                    sheet.cell(rowCnt, targetColCnt + 1).value = tempRow[1]  # 期
                
                # ヘッダー
                if i == 0:
                    sheet.cell(5, targetColCnt + 2 + wroopCnt).value = tempAge[1]
                    sheet.cell(rowCnt, targetColCnt + 2 + wroopCnt).value = format_value_for_graph(tempRow[2])
                elif i == 1:
                    sheet.cell(5, targetColCnt + 2 + wroopCnt).value = f"{tempAge[1]}_ラベル"
                    sheet.cell(rowCnt, targetColCnt + 2 + wroopCnt).value = format_value_for_display(tempRow[2])
                elif i == 2:
                    sheet.cell(5, targetColCnt + 2 + wroopCnt).value = f"{tempAge[1]}_比較用"
                    sheet.cell(rowCnt, targetColCnt + 2 + wroopCnt).value = tempRow[2]
                
                rowCnt += 1
            wroopCnt += 1

