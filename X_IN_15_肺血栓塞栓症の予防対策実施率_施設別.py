import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

#
#2024.2Qから指標名変更
#リスクレベルが「中」以上の手術を施行した患者の肺血栓塞栓症の予防対策の実施率
#
# 分母：肺血栓塞栓症予防実施率(%)
# 分子：肺血栓塞栓症予防管理料算定症例数(人)
# 率　：全身麻酔手術症例数(人)
#

def getSQL_疾患別患者数(HOSPITAL_ID, _C):
    return "SELECT * FROM NUMBER_OF_PATIENTS_BY_DISEASE" + _C + " WHERE Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " "  \
            "ORDER BY 年度, 期, Disease_DISEASE_ID"

def getSQL_CI_15_肺血栓塞栓症の予防対策実施率(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_15" + _C + ".年度 ,  " + \
	            "CI_15" + _C + ".期,  " + \
                "CI_15" + _C + ".肺血栓塞栓症予防実施率, " + \
                "CI_15" + _C + ".肺血栓塞栓症予防管理料算定症例数, " + \
                "CI_15" + _C + ".全身麻酔手術症例数 " + \
            "FROM " + \
	            "CI_15" + _C + " " + \
            "WHERE " + \
	            "NOT CI_15" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_15" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_15" + _C + ".年度, CI_15" + _C + ".期 " 


def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(HOSPITAL_NAME + " CI_15 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_15" in wb.sheetnames:
        wb.remove(wb["CI_15"])
    wb.create_sheet("CI_15")
    sheet = wb["CI_15"]    

    # 書込み列の位置
    colCnt = 1
    
    # 書込み列の繰り返し数
    roopColCnt = 12

    #肺血栓塞栓症の予防対策実施率を取得
    temp肺血栓塞栓症の予防対策実施率 = X_01.excuteSQL(getSQL_CI_15_肺血栓塞栓症の予防対策実施率(HOSPITAL_ID, _C))

    # 1行目に病院名
    sheet.cell(1, 1 + (colCnt - 1) * roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, 1 + (colCnt - 1) * roopColCnt).value = "リスクレベルが「中」以上の手術を施行した患者の肺血栓塞栓症の予防対策"
    # 3行目に疾患名
    sheet.cell(3, 1 + (colCnt - 1) * roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, 1 + (colCnt - 1) * roopColCnt).value = ""

    # 5行名からヘッダを出力
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt).value = "年度"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 1).value = "期"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 2).value = "リスクレベルが「中」以上の手術を施行した患者の肺血栓塞栓症の予防対策の実施率"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 3).value = "分母のうち、肺血栓塞栓症の予防対策が実施された患者数"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 4).value = "肺血栓塞栓症発症のリスクレベルが「中」以上の手術を施行した退院患者数"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 5).value = "リスクレベルが「中」以上の手術を施行した患者の肺血栓塞栓症の予防対策の実施率_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 6).value = "分母のうち、肺血栓塞栓症の予防対策が実施された患者数_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 7).value = "全肺血栓塞栓症発症のリスクレベルが「中」以上の手術を施行した退院患者数_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 8).value = "リスクレベルが「中」以上の手術を施行した患者の肺血栓塞栓症の予防対策の実施率_比較用"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 9).value = "分母のうち、肺血栓塞栓症の予防対策が実施された患者数_比較用"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 10).value = "肺血栓塞栓症発症のリスクレベルが「中」以上の手術を施行した退院患者数_比較用"

    # 6行目からデータ入力
    rowCnt = 6
    for tempRow in temp肺血栓塞栓症の予防対策実施率:

        #年度
        sheet.cell(rowCnt, 1 + (colCnt - 1) * roopColCnt).value = tempRow[0]
        #期
        sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 1).value = tempRow[1]
        
        #肺血栓塞栓症予防実施率
        if tempRow[2] == -1 or tempRow[2] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = tempRow[2]  / 100
        #肺血栓塞栓症予防管理料算定症例数
        if tempRow[3] == -1 or tempRow[3] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = tempRow[3] 
        #全身麻酔手術症例数
        if tempRow[4] == -1 or tempRow[4] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 4).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 4).value = tempRow[4] 

        #肺血栓塞栓症予防実施率_ラベル
        if tempRow[2] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "N/A"
        elif tempRow[2] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = tempRow[2] / 100
        #肺血栓塞栓症予防管理料算定症例数_ラベル
        if tempRow[3] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = "N/A"
        elif tempRow[3] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = tempRow[3]
        #全身麻酔手術症例数_ラベル
        if tempRow[4] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = "N/A"
        elif tempRow[4] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = tempRow[4]

        #肺血栓塞栓症予防実施率_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 8).value = tempRow[2] / 100
        #肺血栓塞栓症予防管理料算定症例数_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 9).value = tempRow[3]
        #全身麻酔手術症例数_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 10).value = tempRow[4]

        rowCnt+=1
    colCnt+=1
    #wb.save(wbPath)
    print(HOSPITAL_NAME + " CI_15 終了")



def getRepAgeData_old(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_15 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_15" in wb.sheetnames:
        ws = wb["CI_15"]
        for row in ws.rows:
            for cell in row:
                cell.value = None
        
        #wb.remove(wb["CI_15"])

    wb.create_sheet("CI_15")
    sheet = wb["CI_15"]    

    #当該病院の疾患別患者数を取得
    list_疾患別患者数 = X_01.excuteSQL(getSQL_疾患別患者数(HOSPITAL_ID, _C))
    if list_疾患別患者数 is None:
        print("エラー: 疾患別患者数の取得に失敗しました。")
        return -1

    dict_疾患別患者数 = {}

    #疾患別患者数を年度、期、DISEASE_IDをキーに辞書に追加
    for temp疾患別患者数 in list_疾患別患者数:
        # 年度、期、DISEASE_IDがキー
        key = str(temp疾患別患者数[1]) + temp疾患別患者数[2] + str(temp疾患別患者数[4])
        #print("key:" + key + " value:" + str(temp疾患別患者数[3])) 
        dict_疾患別患者数[key] = temp疾患別患者数[3]

    # 書込み列の位置
    colCnt = 1
    
    # 書込み列の繰り返し数
    roopColCnt = 15

    #肺血栓塞栓症の予防対策実施率を取得
    temp肺血栓塞栓症の予防対策実施率 = X_01.excuteSQL(getSQL_CI_15_肺血栓塞栓症の予防対策実施率(HOSPITAL_ID, _C))

    # 1行目に病院名
    sheet.cell(1, 1 + (colCnt - 1) * roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, 1 + (colCnt - 1) * roopColCnt).value = "リスクレベルが「中」以上の手術を施行した患者の肺血栓塞栓症の予防対策の実施率"
    # 3行目に疾患名
    sheet.cell(3, 1 + (colCnt - 1) * roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, 1 + (colCnt - 1) * roopColCnt).value = ""

    # 5行名からヘッダを出力
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt).value = "年度"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 1).value = "期"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 2).value = "肺血栓塞栓症予防実施率"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 3).value = "肺血栓塞栓症予防管理料算定症例数"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 4).value = "全身麻酔手術症例数"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 5).value = "疾患別患者数"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 6).value = "肺血栓塞栓症予防実施率_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 7).value = "肺血栓塞栓症予防管理料算定症例数_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 8).value = "全身麻酔手術症例数_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 9).value = "疾患別患者数_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 10).value = "肺血栓塞栓症予防実施率_比較用"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 11).value = "肺血栓塞栓症予防管理料算定症例数_比較用"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 12).value = "全身麻酔手術症例数_比較用"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 13).value = "疾患別患者数_比較用"

    if temp肺血栓塞栓症の予防対策実施率 is None:
        print("エラー: 肺血栓塞栓症の予防対策実施率の取得に失敗しました。")
        return -1

    # 6行目からデータ入力
    rowCnt = 6
    for tempRow in temp肺血栓塞栓症の予防対策実施率:
        #辞書から疾患別患者数を取り出すためのキー（疾患IDは25を使用）
        key = str(tempRow[0]) + str(tempRow[1]) + "25"

        #年度
        sheet.cell(rowCnt, 1 + (colCnt - 1) * roopColCnt).value = tempRow[0]
        #期
        sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 1).value = tempRow[1]
        
        #肺血栓塞栓症予防実施率
        if tempRow[2] == -1 or tempRow[2] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = tempRow[2]  / 100
        #肺血栓塞栓症予防管理料算定症例数
        if tempRow[3] == -1 or tempRow[3] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = tempRow[3] 
        #全身麻酔手術症例数
        if tempRow[4] == -1 or tempRow[4] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 4).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 4).value = tempRow[4] 
        #疾患別患者数
        if key in dict_疾患別患者数:
            if dict_疾患別患者数[key] == -1 or dict_疾患別患者数[key] == -2:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = 0
            else:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = dict_疾患別患者数[key]

        #肺血栓塞栓症予防実施率_ラベル
        if tempRow[2] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = "N/A"
        elif tempRow[2] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = tempRow[2] / 100
        #肺血栓塞栓症予防管理料算定症例数_ラベル
        if tempRow[3] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = "N/A"
        elif tempRow[3] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = tempRow[3]
        #全身麻酔手術症例数_ラベル
        if tempRow[4] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 8).value = "N/A"
        elif tempRow[4] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 8).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 8).value = tempRow[4]
        #疾患別患者数_ラベル
        if key in dict_疾患別患者数:
            if dict_疾患別患者数[key] == -2:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 9).value = "-"
            elif dict_疾患別患者数[key] == -1:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 9).value = "N/A"
            else:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 9).value = dict_疾患別患者数[key]

        #肺血栓塞栓症予防実施率_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 10).value = tempRow[2] / 100
        #肺血栓塞栓症予防管理料算定症例数_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 11).value = tempRow[3]
        #全身麻酔手術症例数_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 12).value = tempRow[4]
        #疾患別患者数_比較用
        if key in dict_疾患別患者数:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 13).value = dict_疾患別患者数[key]

        rowCnt+=1
    colCnt+=1
    #wb.save(wbPath)
    print(" " + HOSPITAL_NAME + " CI_15 終了")


