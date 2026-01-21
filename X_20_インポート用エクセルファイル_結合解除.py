import os
import sys
import shutil
import openpyxl
import tkinter as tk
import X_00_CONST as CONST
from tkinter import filedialog
from copy import copy
from openpyxl.cell.cell import MergedCell


def exit_program():
    print("\n\n********診療アウトカム評価事業  インポート用エクセルセル結合解除 終了********\n")
    sys.exit(0)

#出力フォルダを設定する
#引数：target_year：対象年度、target_group：対象グループ、target_period：対象期間
#戻り値：出力先フォルダのパス
#ERROR:-1
def set_output_folder(target_year,target_group,target_period):
    
    try:
        #出力先に当該年度のフォルダがあるか確認する
        output_folder = CONST.IMPORT_EXCEL_FOLDER
        output_folder = os.path.join(output_folder, str(target_year))
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        #output_folder内に「01_急性期指標」フォルダがあるか確認する
        acute_folder = os.path.join(output_folder, "01_急性期指標")
        if not os.path.exists(acute_folder):
            os.makedirs(acute_folder)
    
        #acute_folder内に当該グループのフォルダがあるか確認する
        group_folder = os.path.join(acute_folder, target_group)
        if not os.path.exists(group_folder):
            os.makedirs(group_folder)
    
        #group_folder内に当該期間のフォルダがあるか確認する
        period_folder = os.path.join(group_folder, target_period) 
        if not os.path.exists(period_folder):
            os.makedirs(period_folder)
    
    except:
        print("エラー: 出力先フォルダの設定に失敗しました。年度={}、グループ={}、期間={}、レポート種類={}".format(target_year,target_group,target_period,target_report_type))
        return -1
      
    #period_folder内のファイル、フォルダを全て削除
    delete_file_count = 0
    for file in os.listdir(period_folder):
        try:
            os.remove(os.path.join(period_folder, file))
            #削除したファイル数をカウント
            delete_file_count += 1
        except:
            print("フォルダ：{}".format(period_folder))
            print(f"エラー: {file} の削除に失敗しました。")
            print("エラー内容：{}".format(sys.exc_info()[1]))
            return -1

    print("削除したファイル数：{}".format(delete_file_count))
    return period_folder



#NIT納品レポートを作業用フォルダにコピー
#引数：target_year：対象年度、target_group：対象グループ、target_period：対象期間、target_report_type：レポート種類
#ERROR:-1
def copy_source_report(target_year,target_group,target_period, output_folder):
    print("\nNIT納品レポートのコピーを開始します。")
    #NIT納品レポートのフォルダを取得
    nit_report_folder = CONST.NIT_REPORT_FOLDER
    nit_report_folder = os.path.join(nit_report_folder, str(target_year))
    nit_report_folder = os.path.join(nit_report_folder, "01_急性期指標")
    nit_report_folder = os.path.join(nit_report_folder, target_group)
    nit_report_folder = os.path.join(nit_report_folder, target_period)
    nit_report_folder = os.path.join(nit_report_folder, "02_EXCEL")

    #nit_report_folderが存在しない場合はエラー  
    if not os.path.exists(nit_report_folder):
        print("\nエラー: NIT納品レポートのフォルダが存在しません。年度={}、グループ={}、期間={}".format(target_year,target_group,target_period))
        print("フォルダパス：{}".format(nit_report_folder))
        return -1
    #nit_report_folder内にファイルが存在しない場合はエラー
    if not os.listdir(nit_report_folder):
        print("\nエラー: NIT納品レポートのフォルダにファイルが存在しません。年度={}、グループ={}、期間={}".format(target_year,target_group,target_period))
        print("フォルダパス：{}".format(nit_report_folder))
        return -1

    try:
        #nit_report_folder内のファイルをoutput_folderにコピー
        for file in os.listdir(nit_report_folder):
            #file名の拡張子の前ににtarget_year,target_groupを追加
            new_file_name = file.split(".")[0] + "_" + str(target_year) + "_" + target_group + ".xlsx"
            shutil.copy(os.path.join(nit_report_folder, file), os.path.join(output_folder, new_file_name))

        print("\nNIT納品レポートのコピーに成功しました。年度={}、グループ={}、期間={}".format(target_year,target_group,target_period))
        print("フォルダパス：{}".format(nit_report_folder))
        print("出力フォルダパス：{}".format(output_folder))
        #コピーしたファイルの数を出力
        print("コピーしたファイルの数：{}".format(len(os.listdir(output_folder))))
        return 0

    except:
        print("\nエラー: NIT納品レポートのコピーに失敗しました。年度={}、グループ={}、期間={}".format(target_year,target_group,target_period))
        print("フォルダパス：{}".format(nit_report_folder))
        print("出力フォルダパス：{}".format(output_folder))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1


def getStartCnt(excel_file):
    if excel_file.startswith("05")  or excel_file.startswith("06") \
        or excel_file.startswith("07"):
        return 4
    
    elif excel_file.startswith("12")  or excel_file.startswith("13") \
        or excel_file.startswith("14") or excel_file.startswith("15") \
        or excel_file.startswith("16") or excel_file.startswith("17") \
        or excel_file.startswith("18") or excel_file.startswith("19") \
        or excel_file.startswith("20") or excel_file.startswith("21") \
        or excel_file.startswith("22") or excel_file.startswith("23") \
        or excel_file.startswith("24") or excel_file.startswith("25") \
        or excel_file.startswith("26") or excel_file.startswith("27") \
        or excel_file.startswith("28") or excel_file.startswith("29") \
        or excel_file.startswith("30") or excel_file.startswith("31") \
        or excel_file.startswith("32") or excel_file.startswith("33") \
        or excel_file.startswith("34") or excel_file.startswith("35") \
        or excel_file.startswith("36") or excel_file.startswith("37") \
        or excel_file.startswith("38") or excel_file.startswith("39") \
        or excel_file.startswith("40") or excel_file.startswith("41") \
        or excel_file.startswith("42") or excel_file.startswith("43") \
        or excel_file.startswith("44") or excel_file.startswith("45") \
        or excel_file.startswith("46") or excel_file.startswith("47") \
        or excel_file.startswith("48") or excel_file.startswith("49") \
        or excel_file.startswith("50") or excel_file.startswith("51") \
        or excel_file.startswith("52") or excel_file.startswith("53") \
        or excel_file.startswith("54") or excel_file.startswith("55") \
        or excel_file.startswith("56") or excel_file.startswith("57") \
        or excel_file.startswith("58") :
        return 5    

    else :
        return 6


#
# 予定しない再入院率のシートの結合セルを解除する
#
def formatCollection(fileName, sheet):
    col1Str = ""
    col2Str = ""
    delIndx = 0

    if fileName.startswith("08"):
        col1Str = "C"
        col2Str = "D"
        delIndx = 3
    else:
        col1Str = "D"
        col2Str = "E"
        delIndx = 4

    sheet.column_dimensions[col1Str].hidden = False
	
	# セル範囲が結合されているか確認する
    merged_cells = sheet.merged_cells.ranges

    max_row = sheet.max_row
    #print(f"シートの最大行数: {max_row}")

    for range in merged_cells:
		#print(f"結合されたセル範囲: {range}")
        if col1Str in str(range) and col2Str in str(range):
			# 結合セルを解除
            sheet.unmerge_cells(str(range))
            #print(f"結合解除: {range}")
			#range.coordを「D4:E4」形式で取得
            cell_range = str(range)
			#「:」で分割
            cells = cell_range.split(":")
			# コピー元セルとコピー先セルを指定
            source_cell = sheet[cells[0]]
            destination_cell = sheet[cells[1]]
			# コピー元のセルの値をコピー先のセルに設定
            destination_cell.value = source_cell.value
			# コピー元のセルのスタイルをコピー先のセルに適用
            if source_cell.has_style:
                destination_cell._style = copy(source_cell._style)
	# 列Cまたは列Dを削除
    sheet.delete_cols(delIndx)
    return sheet


def dividExcel(folder_path):
    #folder_path内のエクセルファイルを取得
    excel_files = [f for f in os.listdir(folder_path) if f.endswith(".xlsx")]
    for excel_file in excel_files:
        print(f"  結合解除開始: {excel_file}")
        try:
            wb = openpyxl.load_workbook(os.path.join(folder_path, excel_file))
            #シート「帳票」を取得
            ws = wb["帳票"]
            startCnt = getStartCnt(excel_file)

            if "予定しない再入院率" in excel_file:
                ws = formatCollection(excel_file, ws)

            for i in range(startCnt, ws.max_row + 1):
                if ws.cell(i, 2).value is not None:
                    if i > startCnt:
                        if i - 1 - startCnt == 0:
                            startCnt = i
                        else:
                            merge_range = f"B{startCnt}:B{i - 1}"
                            if ws.cell(i - 1, 3).value is None:
                                #print("$$$ " + merge_range)
                                if startCnt != i - 2:
                                    if merge_range in [str(rng) for rng in ws.merged_cells.ranges]:
                                        ws.unmerge_cells(merge_range)
                                for t in range(startCnt + 1, i - 1):
                                   if not isinstance(ws.cell(t, 2), MergedCell):
                                        ws.cell(t, 2).value = ws.cell(startCnt, 2).value
                            else:
                                #print("&&& " + merge_range)
                                if merge_range in [str(rng) for rng in ws.merged_cells.ranges]:
                                    ws.unmerge_cells(merge_range)
                                for t in range(startCnt + 1, i):
                                    ws.cell(t, 2).value = ws.cell(startCnt, 2).value
                            startCnt = i
                else:
                    if i == ws.max_row and i > startCnt:
                        merge_range = f"B{startCnt}:B{i}"
                        if merge_range in [str(rng) for rng in ws.merged_cells.ranges]:
                            ws.unmerge_cells(merge_range)
                        for t in range(startCnt + 1, i + 1):
                            ws.cell(t, 2).value = ws.cell(startCnt, 2).value


            wb.save(os.path.join(folder_path, excel_file))
            print(f"  結合解除完了: {excel_file}")
            print("")
            
        except Exception as e:
            print("ファイル名" + excel_file + " 結合解除　失敗")
            print(f"エラー: {e}")
            return -1
        
    return 0


if __name__ == "__main__":
    print("\n\n********診療アウトカム評価事業  インポート用エクセルセル結合解除 開始********")

    print("作成レポートを選んでください。")
    print("１：急性期指標レポート、２：アンケート調査レポート、３：慢性期指標レポート。")

    user_input = input(":>>")
    if user_input not in ["1", "2", "3"]:
        print("エラー: 1, 2, 3（半角） のいずれかを入力してください。")
        exit_program()

    if user_input == "1":
        print("\n急性期指標レポートを作成します。")

        print("\n年度（西暦4桁、半角）を入力してください。")
        user_input = input(":>>")
        if not user_input.isdigit() or len(user_input) != 4:
            print("\nエラー: 西暦4桁の半角数字を入力してください。")
            exit_program()

        target_year = int(user_input)
        print(f"\n対象年度: {str(target_year)}\n")

        print("\nグループ（急性期グループ、慢性期グループ）を選んでください。")
        print("1:急性期グループ、2:慢性期グループ")
        user_input = input(":>>")
        if user_input not in ["1", "2"]:
            print("\nエラー: 急性期グループ、慢性期グループのいずれかを入力してください。")
            exit_program()

        target_group = CONST.GROUPS[int(user_input)-1]
        print(f"\n対象グループ: {target_group}\n")

        print("\n1Q、2Q、3Q、4Q、年間のいずれかを選んでください。")
        print("1:1Q、2:2Q、3:3Q、4:4Q、5:年間")
        user_input = input(":>>")
        if user_input not in ["1", "2", "3", "4", "5"]:
            print("\nエラー: 1Q、2Q、3Q、4Q、年間のいずれかを入力してください。")
            exit_program()

        target_period = CONST.PERIODS[int(user_input) -1]
        print(f"\n対象期間: {target_period}\n")

        output_folder = set_output_folder(target_year,target_group,target_period)
        if output_folder == -1:
            exit_program()
    
        if copy_source_report(target_year,target_group,target_period, output_folder) == -1:
            exit_program()

        print("\nエクセルファイルの結合解除を開始します。")
        print("急性期指標、{}年度、{}グループ、{}期間のエクセルファイルを結合解除します。".format(target_year,target_group,target_period))
        if dividExcel(output_folder) == -1:
            exit_program()

        print("\n結合解除が完了しました。")
        exit_program()





    



    #フォルダ内のエクセルファイルを結合解除する