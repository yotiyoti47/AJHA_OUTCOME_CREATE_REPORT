import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

#def getSQL_疾患マスタ():
#    return "SELECT * FROM MASTA_DISEASE ORDER BY ORDER_NUM"

def getSQL_疾患別患者数(HOSPITAL_ID, _C):
    return "SELECT * FROM NUMBER_OF_PATIENTS_BY_DISEASE" + _C + " WHERE Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " "  \
            "ORDER BY 年度, 期, Disease_DISEASE_ID"

def getSQL_CI_01_平均在院日数(HOSPITAL_ID, disease_ID, _C):
    return  "SELECT  " + \
	            "CI_01" + _C + " .年度 ,  " + \
	            "CI_01" + _C + " .期,  " + \
                "CI_01" + _C + " .平均在院日数 " + \
            "FROM " + \
	            "CI_01" + _C + "  " + \
            "WHERE " + \
	            "NOT CI_01" + _C + " .期 = 'TOTAL' AND " + \
	            "CI_01" + _C + " .Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " AND " + \
	            "CI_01" + _C + " .Disease_DISEASE_ID = " + str(disease_ID) + \
            " ORDER BY CI_01" + _C + " .年度, CI_01" + _C + " .期 " 

def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_01" + "  開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(path, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_01" in wb.sheetnames:
        ws = wb["CI_01"]
        for row in ws.rows:
            for cell in row:
                cell.value = None
        
        #wb.remove(wb["CI_01"])

    wb.create_sheet("CI_01")
    sheet = wb["CI_01"]    

    #疾患マスタを取得
    #list_MASTA_DISEASE = y_common.excuteSQL(y_const.dbPath, getSQL_疾患マスタ())
    #当該病院の疾患別患者数を取得
    
    list_疾患別患者数 = X_01.excuteSQL(getSQL_疾患別患者数(HOSPITAL_ID, _C))
    if list_疾患別患者数 is None:
        print("エラー: 疾患別患者数の取得に失敗しました。")
        return -1

    dict_疾患別患者数 = {}


    #疾患別患者数を年度、期、DISEASE_IDをキーに辞書に追加
    for temp疾患別患者数 in list_疾患別患者数:
        # 年度、期、DISEASE_IDがキー
        key = str(temp疾患別患者数[1]) + temp疾患別患者数[2] + str(temp疾患別患者数[4])
        #print("key:" + key + " value:" + str(temp疾患別患者数[3])) 
        dict_疾患別患者数[key] = temp疾患別患者数[3]

    # 書込み列の位置
    colCnt = 1

    # 書込み列の繰り返し数
    roopColCnt = 9

    #疾患の数だけループ
    for tempDisease in CONST.MASTA_DISEASE:
        #当該疾患の平均在院日数を取得
        temp平均在院日数 = X_01.excuteSQL(getSQL_CI_01_平均在院日数(HOSPITAL_ID, tempDisease[0], _C))
        if temp平均在院日数 is None:
            print("エラー: 平均在院日数の取得に失敗しました。")
            return -1

        #print("疾患名 " + tempDisease[1])

        # 1行目に病院名
        sheet.cell(1, 1 + (colCnt - 1) * roopColCnt).value = HOSPITAL_NAME
        # 2行目に指標名
        sheet.cell(2, 1 + (colCnt - 1) * roopColCnt).value = "平均在院日数_疾患別"
        # 3行目に疾患名
        sheet.cell(3, 1 + (colCnt - 1) * roopColCnt).value = tempDisease[1]
        # 4行名に重症度、年代別、性別
    
        # 5行名からヘッダを出力
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt).value = "年度"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 1).value = "期"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 2).value = "平均在院日数"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 3).value = "疾患別患者数"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 4).value = "平均在院日数_ラベル"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 5).value = "疾患別患者数_ラベル"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 6).value = "平均在院日数_比較用"
        sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 7).value = "疾患別患者数_比較用"

        # 6行目からデータ入力
        rowCnt = 6
        for tempRow in temp平均在院日数:
            #辞書から疾患別患者数を取り出すためのキー
            key = str(tempRow[0]) + str(tempRow[1]) + str(tempDisease[0])

            #年度
            sheet.cell(rowCnt, 1 + (colCnt - 1) * roopColCnt).value = tempRow[0]
            #期
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 1).value = tempRow[1]
            #平均在院日数
            if tempRow[2] == -1 or tempRow[2] == -2:
                sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = 0
            else:
                sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = tempRow[2]
            #疾患別患者数
            if key in dict_疾患別患者数:
                if tempDisease[0] == 25 or  dict_疾患別患者数[key] == -1 or dict_疾患別患者数[key] == -2:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 3).value = 0
                else:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 3).value = dict_疾患別患者数[key]                
            #平均在院日数_ラベル
            if tempRow[2] == -1:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 4).value = "N/A"
            elif tempRow[2] == -2:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 4).value = "-"
            else:
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 4).value = tempRow[2]
            #疾患別患者数_ラベル
            if key in dict_疾患別患者数:
                if tempDisease[0] == 25 or dict_疾患別患者数[key] == -2:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "-"
                elif dict_疾患別患者数[key]== -1:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "N/A"            
                else:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = dict_疾患別患者数[key]
            #平均在院日数_比較用
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = tempRow[2]
            #疾患別患者数_比較用
            if key in dict_疾患別患者数:
                if tempDisease[0] == 25:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = 0
                else:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = dict_疾患別患者数[key]
            rowCnt+=1
        colCnt+=1
    #wb.save(wbPath)
    print(" " + HOSPITAL_NAME + " CI_01 終了")

