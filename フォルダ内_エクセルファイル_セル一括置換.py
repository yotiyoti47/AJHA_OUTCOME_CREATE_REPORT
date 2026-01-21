import openpyxl
import os

if __name__ == "__main__":
    print("\n\n********診療アウトカム評価事業　フォルダ内エクセルファイルのセル一括置換 開始********")

    print("\nフォルダパスを入力してください。")
    folder_path = input(":>>")

    print("\n置換前の文字列を入力してください。")
    before_text = input(":>>")

    print("\n置換後の文字列を入力してください。")
    after_text = input(":>>")

    print("\nフォルダ内のエクセルファイルを一括置換を開始します。")
    for file in os.listdir(folder_path):
        if file.endswith(".xlsx"):
            print("\n対象ファイル：{}".format(file))
            wb = openpyxl.load_workbook(os.path.join(folder_path, file))
            #全てのシート取得
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if before_text in str( cell.value):
                            cell.value = str(cell.value).replace(before_text, after_text)

            print("セル一括置換を完了しました。:{}".format(file))
            wb.save(os.path.join(folder_path, file))

    print("\nフォルダ内のエクセルファイルの一括置換を完了しました。")

    print("\n\n********診療アウトカム評価事業　フォルダ内エクセルファイルのセル一括置換 完了********")