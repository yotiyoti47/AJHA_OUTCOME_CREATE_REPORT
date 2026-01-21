import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

def getSQL_CI_07_死亡率_疾患別_性別(disease_ID, sex_ID, _C):
    return  "SELECT  " + \
	            "CI_07" + _C + ".年度 ,  " + \
	            "CI_07" + _C + ".期,  " + \
                "CI_07" + _C + ".死亡率, " + \
                "CI_07" + _C + ".死亡症例数, " + \
                "CI_07" + _C + ".症例数 " + \
            "FROM " + \
	            "CI_07" + _C + " " + \
            "WHERE " + \
	            "NOT CI_07" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_07" + _C + ".Disease_DISEASE_ID = " + str(disease_ID) + " AND " + \
                "CI_07" + _C + ".Sex_SEX_ID = " + str(sex_ID) +  \
            " ORDER BY CI_07" + _C + ".年度, CI_07" + _C + ".期, CI_07" + _C + ".Disease_DISEASE_ID,  CI_07" + _C + ".Sex_SEX_ID " 

def getRepAgeData(wb, is_acute):
    print(" 全体集計" + " CI_07 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_07" in wb.sheetnames:
        wb.remove(wb["CI_07"])
    wb.create_sheet("CI_07")
    sheet = wb["CI_07"]    

    # 書込み列の位置
    colCnt = 1

   # 書込み列の繰り返し数
    roopColCnt = 12

    #疾患の数だけループ
    for tempDisease in CONST.MASTA_DISEASE:
        
        #性別の数だけループ
        for tempSex in CONST.MASTA_SEX:

            #print("疾患名 " + tempDisease[1] + " 疾患名 " + tempSev[1])
            #print(getSQL_CI_05_死亡率_疾患別_重症度別(tempDisease[0], tempSev[0]))
            temp死亡率 = X_01.excuteSQL(getSQL_CI_07_死亡率_疾患別_性別(tempDisease[0], tempSex[0], _C))
            #print("疾患名 " + tempDisease[1])

            # 1行目に病院名
            sheet.cell(1, 1 + (colCnt - 1) * roopColCnt).value = "全体"
            # 2行目に指標名
            sheet.cell(2, 1 + (colCnt - 1) * roopColCnt).value = "死亡率_疾患別_性別"
            # 3行目に疾患名
            sheet.cell(3, 1 + (colCnt - 1) * roopColCnt).value = tempDisease[1]
            # 4行名に重症度、年代別、性別
            sheet.cell(4, 1 + (colCnt - 1) * roopColCnt).value = tempSex[1]

            # 5行名からヘッダを出力
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt).value = "年度"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 1).value = "期"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 2).value = "死亡率"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 3).value = "死亡症例数"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 4).value = "症例数"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 5).value = "死亡率_ラベル"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 6).value = "死亡症例数_ラベル"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 7).value = "症例数_ラベル"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 8).value = "死亡率_比較用"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 9).value = "死亡症例数_比較用"
            sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 10).value = "症例数_比較用"

            if temp死亡率 is None:
                print("エラー: 死亡率の取得に失敗しました。")
                return -1

            # 6行目からデータ入力
            rowCnt = 6
            for tempRow in temp死亡率:

                #年度
                sheet.cell(rowCnt, 1 + (colCnt - 1) * roopColCnt).value = tempRow[0]
                #期
                sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 1).value = tempRow[1]

                #死亡率
                if tempRow[2] == -1 or tempRow[2] == -2:
                    sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = 0
                else:
                    sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = tempRow[2]          
                #死亡症例数
                if tempRow[3] == -1 or tempRow[3] == -2:
                    sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = 0
                else:
                    sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = tempRow[3]   
                #症例数
                if tempRow[4] == -1 or tempRow[4] == -2:
                    sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 4).value = 0
                else:
                    sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 4).value = tempRow[4]   
                                
                #死亡率_ラベル
                if tempRow[2] == -1:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "N/A"
                elif tempRow[2] == -2:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "-"
                else:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = tempRow[2]                
                #死亡症例数_ラベル
                if tempRow[3] == -1:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = "N/A"
                elif tempRow[3] == -2:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = "-"
                else:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = tempRow[3]
                #症例数_ラベル
                if tempRow[4] == -1:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = "N/A"
                elif tempRow[4] == -2:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = "-"
                else:
                    sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = tempRow[4]

                #死亡率_比較用
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 8).value = tempRow[2]
                #死亡症例数_比較用
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 9).value = tempRow[3]
                #症例数_比較用
                sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 10).value = tempRow[4]

                rowCnt+=1
            colCnt+=1
    #wb.save(wbPath)
    print(" 全体集計" + " CI_07 終了")

