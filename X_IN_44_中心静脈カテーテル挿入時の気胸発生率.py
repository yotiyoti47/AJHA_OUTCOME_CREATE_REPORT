import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

def getSQL_CI_44_中心静脈カテーテル挿入時の気胸発生率(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_44" + _C + ".年度 ,  " + \
	            "CI_44" + _C + ".期,  " + \
                "CI_44" + _C + ".中心静脈カテーテル挿入時の気胸発生率, " + \
                "CI_44" + _C + ".分母のうち入院後医原性気胸が発生した患者数, " + \
                "CI_44" + _C + ".中心静脈カテーテルが挿入された患者数 " + \
            "FROM " + \
	            "CI_44" + _C + " " + \
            "WHERE " + \
	            "NOT CI_44" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_44" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_44" + _C + ".年度, CI_44" + _C + ".期 " 

def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_44 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_44" in wb.sheetnames:
        wb.remove(wb["CI_44"])
    wb.create_sheet("CI_44")
    sheet = wb["CI_44"]    

    # 書込み列の位置
    colCnt = 1

    # 書込み列の繰り返し数
    roopColCnt = 12
    
    #中心静脈カテーテル挿入時の気胸発生率
    temp中心静脈カテーテル挿入時の気胸発生率 = X_01.excuteSQL(getSQL_CI_44_中心静脈カテーテル挿入時の気胸発生率(HOSPITAL_ID, _C))

    # 1行目に病院名
    sheet.cell(1, 1 + (colCnt - 1) * roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, 1 + (colCnt - 1) * roopColCnt).value = "中心静脈カテーテル挿入時の気胸発生率"
    # 3行目に疾患名
    sheet.cell(3, 1 + (colCnt - 1) * roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, 1 + (colCnt - 1) * roopColCnt).value = ""

    # 5行名からヘッダを出力
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt).value = "年度"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 1).value = "期"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 2).value = "脳卒中地域連携パスの使用率_転院先"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 3).value = "分母のうち入院後医原性気胸が発生した患者数"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 4).value = "中心静脈カテーテルが挿入された患者数"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 5).value = "脳卒中地域連携パスの使用率_転院先_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 6).value = "分母のうち入院後医原性気胸が発生した患者数_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 7).value = "中心静脈カテーテルが挿入された患者数_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 8).value = "脳卒中地域連携パスの使用率_転院先_比較用"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 9).value = "分母のうち入院後医原性気胸が発生した患者数_比較用"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 10).value = "中心静脈カテーテルが挿入された患者数_比較用"
    
    if temp中心静脈カテーテル挿入時の気胸発生率 is None:
        print("エラー: 中心静脈カテーテル挿入時の気胸発生率の取得に失敗しました。")
        return -1

    # 6行目からデータ入力
    rowCnt = 6
    for tempRow in temp中心静脈カテーテル挿入時の気胸発生率:

        #年度
        sheet.cell(rowCnt, 1 + (colCnt - 1) * roopColCnt).value = tempRow[0]
        #期
        sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 1).value = tempRow[1]
        
        #脳卒中地域連携パスの使用率_転院先
        if tempRow[2] == -1 or tempRow[2] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = tempRow[2] / 100
        #分母のうち入院後医原性気胸が発生した患者数
        if tempRow[3] == -1 or tempRow[3] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = tempRow[3] 
        #中心静脈カテーテルが挿入された患者数
        if tempRow[4] == -1 or tempRow[4] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 4).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 4).value = tempRow[4] 

        #脳卒中地域連携パスの使用率_転院先_ラベル
        if tempRow[2] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "N/A"
        elif tempRow[2] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = tempRow[2] / 100
        #分母のうち入院後医原性気胸が発生した患者数_ラベル
        if tempRow[3] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = "N/A"
        elif tempRow[3] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = tempRow[3]
        #中心静脈カテーテルが挿入された患者数_ラベル
        if tempRow[4] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = "N/A"
        elif tempRow[4] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = tempRow[4]

        #脳卒中地域連携パスの使用率_転院先_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 8).value = tempRow[2] / 100
        #分母のうち入院後医原性気胸が発生した患者数_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 9).value = tempRow[3]
        #中心静脈カテーテルが挿入された患者数_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 10).value = tempRow[4]

        rowCnt+=1
    colCnt+=1
    #wb.save(wbPath)
    print(" " + HOSPITAL_NAME + " CI_44 終了")

