import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

def getSQL_CI_55_血液培養の実施_2セット実施率(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_55" + _C + ".年度 ,  " + \
	            "CI_55" + _C + ".期,  " + \
                "CI_55" + _C + ".血液培養の実施_2セット実施率, " + \
                "CI_55" + _C + ".血液培養オーダが1日に2件以上ある日数, " + \
                "CI_55" + _C + ".血液培養オーダ日数 " + \
            "FROM " + \
	            "CI_55" + _C + " " + \
            "WHERE " + \
	            "NOT CI_55" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_55" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_55" + _C + ".年度, CI_55" + _C + ".期 " 

def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_55 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_55" in wb.sheetnames:
        wb.remove(wb["CI_55"])
    wb.create_sheet("CI_55")
    sheet = wb["CI_55"]    

    # 書込み列の位置
    colCnt = 1

    # 書込み列の繰り返し数
    roopColCnt = 12
    
    #血液培養の実施_2セット実施率
    temp血液培養の実施_2セット実施率 = X_01.excuteSQL(getSQL_CI_55_血液培養の実施_2セット実施率(HOSPITAL_ID, _C))

    # 1行目に病院名
    sheet.cell(1, 1 + (colCnt - 1) * roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, 1 + (colCnt - 1) * roopColCnt).value = "血液培養2セット実施率"
    # 3行目に疾患名
    sheet.cell(3, 1 + (colCnt - 1) * roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, 1 + (colCnt - 1) * roopColCnt).value = ""

    # 5行名からヘッダを出力
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt).value = "年度"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 1).value = "期"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 2).value = "血液培養の実施(2セット実施率)"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 3).value = "血液培養オーダが1日に2件以上ある日数(人日)"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 4).value = "血液培養オーダ日数(人日)"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 5).value = "血液培養の実施(2セット実施率)_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 6).value = "血液培養オーダが1日に2件以上ある日数(人日)_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 7).value = "血液培養オーダ日数(人日)_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 8).value = "血液培養の実施(2セット実施率)_比較用"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 9).value = "血液培養オーダが1日に2件以上ある日数(人日)_比較用"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 10).value = "血液培養オーダ日数(人日)_比較用"

    if temp血液培養の実施_2セット実施率 is None:
        print("エラー: 血液培養の実施(2セット実施率)の取得に失敗しました。")
        return -1

    # 6行目からデータ入力
    rowCnt = 6
    for tempRow in temp血液培養の実施_2セット実施率:

        #年度
        sheet.cell(rowCnt, 1 + (colCnt - 1) * roopColCnt).value = tempRow[0]
        #期
        sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 1).value = tempRow[1]
        
        #血液培養の実施_2セット実施率
        if tempRow[2] == -1 or tempRow[2] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = tempRow[2]  / 100
        #血液培養オーダが1日に2件以上ある日数
        if tempRow[3] == -1 or tempRow[3] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = tempRow[3] 
        #血液培養オーダ日数
        if tempRow[4] == -1 or tempRow[4] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 4).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 4).value = tempRow[4] 

        #血液培養の実施_2セット実施率_ラベル
        if tempRow[2] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "N/A"
        elif tempRow[2] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = tempRow[2] / 100
        #血液培養オーダが1日に2件以上ある日数_ラベル
        if tempRow[3] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = "N/A"
        elif tempRow[3] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = tempRow[3]
        #血液培養オーダ日数_ラベル
        if tempRow[4] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = "N/A"
        elif tempRow[4] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = tempRow[4]

        #血液培養の実施_2セット実施率_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 8).value = tempRow[2] / 100
        #血液培養オーダが1日に2件以上ある日数_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 9).value = tempRow[3]
        #血液培養オーダ日数_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 10).value = tempRow[4]

        rowCnt+=1
    colCnt+=1
    #wb.save(wbPath)
    print(" " + HOSPITAL_NAME + " CI_55 終了")

