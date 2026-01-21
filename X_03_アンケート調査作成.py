import re
from datetime import datetime
import tkinter as tk
from tkinter import filedialog
import X_01_レポート作成 as COMMON
import X_00_CONST as CONST
import os
import sys
import openpyxl
from openpyxl.styles import PatternFill
import win32com.client
import shutil
#from PyPDF2 import PdfReader, PdfWriter
from pypdf import PdfReader, PdfWriter
from openpyxl import load_workbook

def set_output_folder(target_year, target_number, target_report_type):
    try:
        #出力先に当該年度のフォルダがあるか確認する
        output_folder = CONST.OUTPUT_FOLDER
        output_folder = os.path.join(output_folder, str(target_year))
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        #output_folder内に「02_アンケート調査」フォルダがあるか確認する
        acute_folder = os.path.join(output_folder, "02_アンケート調査")
        if not os.path.exists(acute_folder):
            os.makedirs(acute_folder)

        number_folder = os.path.join(acute_folder, str(target_number))
        if not os.path.exists(number_folder):
            os.makedirs(number_folder)

        report_type_folder = os.path.join(number_folder, target_report_type)
        if not os.path.exists(report_type_folder):
            os.makedirs(report_type_folder)

       
    except:
        print("エラー: 出力先フォルダの設定に失敗しました。年度={}、回数={}".format(target_year,target_number))
        return -1

    #report_type_folder内のファイル、フォルダを全て削除
    delete_file_count = 0
    for file in os.listdir(report_type_folder):
        try:
            os.remove(os.path.join(report_type_folder, file))
            #削除したファイル数をカウント
            delete_file_count += 1
        except:
            print("フォルダ：{}".format(report_type_folder))
            print(f"エラー: {file} の削除に失敗しました。")
            print("エラー内容：{}".format(sys.exc_info()[1]))
            return -1

    print("削除したファイル数：{}".format(delete_file_count))
    return report_type_folder


def create_cover_page(target_year, target_number, target_create_date, output_folder):
    print("\n表紙ファイルの作成を開始します。")
    try:
        #表紙ファイルを開く
        over_page_path = CONST.COVER_PAGE_FORMAT_PATH_SURVEY
        #openpyxl で開く
        workbook = openpyxl.load_workbook(over_page_path)
        if workbook is None:
            print("\nエラー: 表紙ファイルの読み込みに失敗しました。年度={}、回数={}".format(target_year,target_number))
            print("ファイルパス：{}".format(over_page_path))
            return -1

        #シート名「00-1_表紙」を開く
        sheet = workbook["00-1_表紙"]
        #セルCD1に指標タイプを入力
        sheet["CD1"] = "アンケート調査"
        #セルCD2に年度を入力
        sheet["CD2"] = target_year
 
        sheet["CD3"] = target_number

        #セルCD5に作成日を入力
        sheet["CD6"] = target_create_date
        #セルA4の色を変更
        sheet["A4"].fill = PatternFill(start_color=CONST.REPORT_INDCTR_COLOR[1], end_color=CONST.REPORT_INDCTR_COLOR[1], fill_type="solid")
        #セルA22の色を変更
     
        output_folder = os.path.join(output_folder, "00-1_表紙.xlsx")
        workbook.save(output_folder)
        if workbook is None:
            print("\nエラー: 表紙ファイルの保存に失敗しました。年度={}、回数={}".format(target_year,target_number))
            print("ファイルパス：{}".format(output_folder))
            return -1

        #PDFに変換
        # Excel ファイルを開く
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Excel のウィンドウを非表示にする
        wb = excel.Workbooks.Open(output_folder)
        newPath = str(output_folder).replace(".xlsx",".pdf")
        wb.ExportAsFixedFormat(0, newPath)  # 0 は PDF 形式
        wb.Close()
        excel.Quit()

        #pdfファイルの作成が完了したら元のエクセルファイルを削除
        os.remove(output_folder)

        print(f"\n表紙ファイルを作成しました。ファイルパス: {output_folder}\n")
        return 0
    except:
        print("\nエラー: 表紙ファイルの作成に失敗しました。年度={}、回数={}".format(target_year,target_number))
        print("ファイルパス：{}".format(output_folder))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1


#NIT納品レポートを作業用フォルダにコピー
#引数：target_year：対象年度、target_group：対象グループ、target_period：対象期間、target_report_type：レポート種類
#ERROR:-1
def copy_source_report(target_year, target_number, target_report_type, output_folder):
    print("\nNIT納品レポートのコピーを開始します。")

    if target_report_type == "01_病院名":
        target_report_type = "01_PDF"
    elif target_report_type == "02_病院番号":
        target_report_type = "02_EXCEL"


    #NIT納品レポートのフォルダを取得
    nit_report_folder = CONST.NIT_REPORT_FOLDER
    nit_report_folder = os.path.join(nit_report_folder, str(target_year))
    nit_report_folder = os.path.join(nit_report_folder, "02_アンケート調査")
    nit_report_folder = os.path.join(nit_report_folder, target_number)
    nit_report_folder = os.path.join(nit_report_folder, target_report_type)

    #nit_report_folderが存在しない場合はエラー  
    if not os.path.exists(nit_report_folder):
        print("\nエラー: NIT納品レポートのフォルダが存在しません。年度={}、回数={}".format(target_year,target_number))
        print("フォルダパス：{}".format(nit_report_folder))
        return -1
    #nit_report_folder内にファイルが存在しない場合はエラー
    if not os.listdir(nit_report_folder):
        print("\nエラー: NIT納品レポートのフォルダにファイルが存在しません。年度={}、回数={}".format(target_year,target_number))
        print("フォルダパス：{}".format(nit_report_folder))
        return -1

    try:

        for file in os.listdir(nit_report_folder):
            #ファイル名が”第１回”または”第2回”から始まらなければエラー
            if not file.startswith("第１回") and not file.startswith("第２回"):
                print("\nエラー: ファイル名が”第１回”または”第２回”から始まりません。※回数は全角  年度={}、回数={}、レポート種類={}".format(target_year,target_number,target_report_type))
                print("ファイル名：{}".format(file))
                return -1
            #output_folderにコピー


            if target_report_type == "01_PDF":
                #ファイル名に「入院、患者満足度」が含まれていたらファイル名を「61_入院　患者満足度調査レポート_第１回」に変更
                if "入院、患者満足度" in file:
                    #ファイル名に「第1回」が含まれている場合は「61_入院　患者満足度調査レポート_第１回」に変更
                    if "第１回" in file:
                        new_file_name = "61_入院　患者満足度調査レポート_第１回.pdf"
                    elif "第２回" in file:
                        new_file_name = "61_入院　患者満足度調査レポート_第２回.pdf"
                elif "入院、病院推奨度" in file:
                    #ファイル名に「第1回」が含まれている場合は「62_入院　病院推奨度調査レポート_第１回」に変更
                    if "第１回" in file:
                        new_file_name = "62_入院　病院推奨度調査レポート_第１回.pdf"
                    elif "第２回" in file:
                        new_file_name = "62_入院　病院推奨度調査レポート_第２回.pdf"
                elif "入院、医師満足度" in file:
                    #ファイル名に「第1回」が含まれている場合は「63_入院　医師満足度調査レポート_第１回」に変更
                    if "第１回" in file:
                        new_file_name = "63_入院　医師満足度調査レポート_第１回.pdf"
                    elif "第２回" in file:
                        new_file_name = "63_入院　医師満足度調査レポート_第２回.pdf"
                elif "外来、患者満足度" in file:
                    #ファイル名に「第1回」が含まれている場合は「64_外来　患者満足度調査レポート_第１回」に変更
                    if "第１回" in file:
                        new_file_name = "64_外来　患者満足度調査レポート_第１回.pdf"
                    elif "第２回" in file:
                        new_file_name = "64_外来　患者満足度調査レポート_第２回.pdf"
                elif "外来、医師満足度" in file:
                    #ファイル名に「第1回」が含まれている場合は「65_外来　医師満足度調査レポート_第１回」に変更
                    if "第１回" in file:
                        new_file_name = "65_外来　医師満足度調査レポート_第１回.pdf"
                    elif "第２回" in file:
                        new_file_name = "65_外来　医師満足度調査レポート_第２回.pdf"
                elif "1か月、百床あたり" in file:
                    new_file_name = "66_インシデント・アクシデント（1か月、百床あたり）レポート.pdf"
                elif "全報告中医師の占める割合" in file:
                    new_file_name = "67_インシデント・アクシデント（全報告中医師の占める割合）レポート.pdf"
                elif "インフルエンザ" in file:
                    new_file_name = "68_職員の予防接種（インフルエンザ）レポート.pdf"
                else:
                    print("\nエラー: ファイル名が不正です。")
                    print("ファイル名：{}".format(file))
                    return -1

                shutil.copy(os.path.join(nit_report_folder, file), os.path.join(output_folder, new_file_name))
                
            else:
                #ファイルをコピー
                shutil.copy(os.path.join(nit_report_folder, file), os.path.join(output_folder, file))
               
               
            
        print("\nNIT納品レポートのコピーに成功しました。年度={}、回数={}、レポート種類={}".format(target_year,target_number,target_report_type))
        print("フォルダパス：{}".format(nit_report_folder))
        print("出力フォルダパス：{}".format(output_folder))
        #コピーしたファイルの数を出力
        print("コピーしたファイルの数：{}".format(len(os.listdir(output_folder))))
        return 0

    except:
        print("\nエラー: NIT納品レポートのコピーに失敗しました。年度={}、回数={}、レポート種類={}".format(target_year,target_number,target_report_type))
        print("フォルダパス：{}".format(nit_report_folder))
        print("出力フォルダパス：{}".format(output_folder))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1


def copy_TEIGI_FILES(target_year, target_number, output_folder):
    print("\n指標定義ファイルのコピーを開始します。")
    TEIGI_FILES_PATH = CONST.TEIGI_FILES_SURVEY
    
    try:
        #TEIGI_FILES_PATH内のpdfファイルをoutput_folderにコピー
        for file in os.listdir(TEIGI_FILES_PATH):
            if(target_number == "第１回"):
                #ファイル名の先頭2文字が61から65までのPDFファイルをコピー
                if(file.startswith("61") or file.startswith("62") or file.startswith("63") or file.startswith("64") or file.startswith("65")):
                    shutil.copy(os.path.join(TEIGI_FILES_PATH, file), os.path.join(output_folder, file))

            elif(target_number == "第２回"):
                #TEIGI_FILES_PATH内のPDFファイルのみコピー
                if(file.endswith(".pdf")):
                    shutil.copy(os.path.join(TEIGI_FILES_PATH, file), os.path.join(output_folder, file))

        return 0
    except:
        print("\nエラー: 指標定義ファイルのコピーに失敗しました。年度={}、回数={}".format(target_year,target_number))
        print("ファイルパス：{}".format(TEIGI_FILES_PATH))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1


#def combine_report(target_year, target_number, target_report_type, output_folder):
    print("\nレポートのPDF結合を開始します。")
    try:
        #output_folder内のpdfファイルを結合
        pdf_files = [f for f in os.listdir(output_folder) if f.endswith(".pdf")]
        pdf_files.sort()
        if not pdf_files:
            print("\nエラー: レポートファイルが見つかりません。年度={}、回数={}、レポート種類={}、フォルダパス={}".format(target_year,target_number,target_report_type,output_folder))
            return -1

        report_file_name = str(target_year) + "年度_アンケート調査レポート_" + target_number + "_" + target_report_type + ".pdf"
        #出力先はoutput_folderの2階層上のフォルダに作成
        reportoutput_folder = os.path.join(os.path.dirname(output_folder), "..")


        #PDFファイルを結合、ファイル名をしおりに追加
        writer = PdfWriter()
        current_page = 0

        for file in pdf_files:
            reader = PdfReader(os.path.join(output_folder, file))
            num_pages = len(reader.pages)

            # 各ページをwriterに追加
            for page in reader.pages:
                writer.add_page(page)

            # ファイル名をしおりとして追加（拡張子なしで追加もOK）
            title = os.path.splitext(os.path.basename(file))[0]
            if "レポート" in title:
                current_page += num_pages
                continue
            writer.add_outline_item(title=title, page_number=current_page)
            # 次のしおりのページ位置を把握
            current_page += num_pages

            
            
        # 保存
        with open(os.path.join(reportoutput_folder, report_file_name), "wb") as f:
            writer.write(f)

        print("\nレポートのPDF結合を完了しました。:{}".format(report_file_name))

        #pdf_filesを結合
        #pdf_merger = PdfMerger()
        #for file in pdf_files:
        #    pdf_merger.append(os.path.join(output_folder, file))
        #pdf_merger.write(os.path.join(reportoutput_folder, report_file_name))

    except:
        print("\nエラー: レポートの結合に失敗しました。年度={}、回数={}、レポート種類={}、フォルダパス={}".format(target_year,target_number,target_report_type,output_folder))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1


def combine_report_pypdf(target_year, target_number, target_report_type, output_folder):
    print("\nレポートのPDF結合を開始します。")
    try:
        # output_folder 内の PDF ファイルを結合
        pdf_files = [f for f in os.listdir(output_folder) if f.lower().endswith(".pdf")]
        pdf_files.sort()

        if not pdf_files:
            print(f"\nエラー: レポートファイルが見つかりません。年度={target_year}、回数={target_number}、レポート種類={target_report_type}、フォルダパス={output_folder}")
            return -1

        report_file_name = f"{target_year}年度_アンケート調査レポート_{target_number}_{target_report_type}.pdf"
        reportoutput_folder = os.path.abspath(os.path.join(output_folder, "..", ".."))  # 2階層上

        writer = PdfWriter()
        current_page = 0

        for file in pdf_files:
            full_path = os.path.join(output_folder, file)
            reader = PdfReader(full_path)
            num_pages = len(reader.pages)

            for page in reader.pages:
                writer.add_page(page)

            title = os.path.splitext(os.path.basename(file))[0]
            if "レポート" not in title:
                writer.add_outline_item(title=title, page_number=current_page)

            current_page += num_pages

        # 出力ディレクトリ作成（なければ）
        os.makedirs(reportoutput_folder, exist_ok=True)
        output_path = os.path.join(reportoutput_folder, report_file_name)

        with open(output_path, "wb") as f:
            writer.write(f)

        print(f"\nレポートのPDF結合を完了しました。: {output_path}")
        return 0

    except Exception as e:
        print(f"\nエラー: レポートの結合に失敗しました。年度={target_year}、回数={target_number}、レポート種類={target_report_type}、フォルダパス={output_folder}")
        print(f"エラー内容：{e}")
        return -1

def change_report_name(excel_files, output_folder):
    print("\nレポートのファイル名を変更します。")
    for file in excel_files:
        try:
            # 61_入院_患者満足度調査レポート_第１回.pdf
            # 61_入院_患者満足度調査レポート_第２回.pdf
            # 62_入院_病院推奨度調査レポート_第１回.pdf
            # 62_入院_病院推奨度調査レポート_第２回.pdf
            # 63_入院_医師満足度調査レポート_第１回.pdf
            # 63_入院_医師満足度調査レポート_第２回.pdf
            # 64_外来_患者満足度調査レポート_第１回.pdf
            # 64_外来_患者満足度調査レポート_第２回.pdf
            # 65_外来_医師満足度調査レポート_第１回.pdf
            # 65_外来_医師満足度調査レポート_第２回.pdf
            # 66_インシデント・アクシデント（1か月、百床あたり）レポート.pdf
            # 67_インシデント・アクシデント（全報告中医師の占める割合）レポート.pdf
            # 68_職員の予防接種（インフルエンザ）レポート.pdf


            if "入院" in str(file):
                #ファイルを61_入院_患者満足度調査_レポート_に変更してコピー
                new_file1 = "61_入院_患者満足度調査レポート_"
                new_file2 = "62_入院_病院推奨度調査レポート_"
                new_file3 = "63_入院_医師満足度調査レポート_"

                if "第１回" in str(file):
                    new_file1 += "第１回"
                    new_file2 += "第１回"
                    new_file3 += "第１回"
                elif "第２回" in str(file):
                    new_file1 += "第２回"
                    new_file2 += "第２回"
                    new_file3 += "第２回"
                new_file1 += ".xlsx"
                new_file2 += ".xlsx"
                new_file3 += ".xlsx"

                shutil.copy(os.path.join(output_folder, file), os.path.join(output_folder, new_file1))
                shutil.copy(os.path.join(output_folder, file), os.path.join(output_folder, new_file2))
                shutil.copy(os.path.join(output_folder, file), os.path.join(output_folder, new_file3))

                #fileを削除
                os.remove(os.path.join(output_folder, file))

            elif "外来" in str(file):

                #ファイルを2_入院_病院推奨度調査_レポート_に変更してコピー
                new_file1 = "64_外来_患者満足度調査レポート_"
                new_file2 = "65_外来_医師満足度調査レポート_"

                if "第１回" in str(file):
                    new_file1 += "第１回"
                    new_file2 += "第１回"
                elif "第２回" in str(file):
                    new_file1 += "第２回"
                    new_file2 += "第２回"
                new_file1 += ".xlsx"
                new_file2 += ".xlsx"

                shutil.copy(os.path.join(output_folder, file), os.path.join(output_folder, new_file1))
                shutil.copy(os.path.join(output_folder, file), os.path.join(output_folder, new_file2))
                                

                #fileを削除
                os.remove(os.path.join(output_folder, file))

            elif "医療安全" in str(file):
                new_file1 = "66_インシデント・アクシデント（1か月、百床あたり）レポート.xlsx"
                new_file2 = "67_インシデント・アクシデント（全報告中医師の占める割合）レポート.xlsx"

                shutil.copy(os.path.join(output_folder, file), os.path.join(output_folder, new_file1))
                shutil.copy(os.path.join(output_folder, file), os.path.join(output_folder, new_file2))
              
                #fileを削除
                os.remove(os.path.join(output_folder, file))

            elif "職員の予防接種" in str(file):
                #ファイル名を68_職員の予防接種（インフルエンザ）レポート.xlsxに変更
                new_file = "68_職員の予防接種（インフルエンザ）レポート.xlsx"
                shutil.copy(os.path.join(output_folder, file), os.path.join(output_folder, new_file))
                
                #fileを削除
                os.remove(os.path.join(output_folder, file))
            
            
        except:
            print("\nエラー: レポートのファイル名を変更に失敗しました。\n ファイル名={}".format(file))
            print("エラー内容：{}".format(sys.exc_info()[1]))
            return -1

    return 0


def tranaHospNameToPublicNO(excel_file_path, dict_hosp_DB, output_folder, sheetName):
    filename = os.path.basename(excel_file_path)

    print("\n病院名を番号変換を開始します。:{}".format(filename))   
    print("シート名:{}".format(sheetName))


    startRow = 5
    if("職員の予防接種" in str(filename)):
        startRow = 4

    try:
        wb = load_workbook(excel_file_path)

        # すべてのシートを非表示にする
        for sheet in wb.worksheets:
            sheet.sheet_state = 'hidden'

        wb[sheetName].sheet_state = 'visible'
        wb.active = wb[sheetName]
        ws = wb[sheetName]
        col_data = [cell.value for cell in ws['B']]
        for i in range(startRow, len(col_data) + 1, 1):
            if(ws.cell(row=i, column=2).value == "合計" or ws.cell(row=i, column=2).value == None):
                break
            #print(str(ws.cell(row=i, column=2) + " i " + str(i)))
            #print("変換前：" + str(ws.cell(row=i, column=2).value) + "   変換後：" + str(dict_hosp_DB[ws.cell(row=i, column=2).value]))
            pNo = dict_hosp_DB[ws.cell(row=i, column=2).value]         
            ws.cell(row=i, column=2, value=str(pNo))         

        wb.save(excel_file_path)  # 保存
        wb.close()

        print("病院名を番号変換を完了しました。:{}".format(filename))
        print("PDFに変換を開始します。")
        #PDFに変換
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Excel のウィンドウを非表示にする
        wb = excel.Workbooks.Open(excel_file_path)
        # 指定シートを取得
        ws = wb.Sheets(sheetName)
        newPath = str(os.path.join(output_folder, filename.replace(".xlsx",".pdf")))

        ws.ExportAsFixedFormat(0, newPath)  # 0 は PDF 形式
        wb.Close()
        excel.Quit()

        print("PDFに変換を完了しました。:{}".format(filename))

    except Exception as e:
        print(f"\n病院名を番号変換時にエラーが発生しました。: {e}")
        return -1












def create_report():
    print("\nアンケート調査レポートを作成します。")

    print("\n年度（西暦4桁、半角）を入力してください。")
    user_input = input(":>>")
    if not user_input.isdigit() or len(user_input) != 4:
        print("\nエラー: 西暦4桁の半角数字を入力してください。")
        return -1

    target_year = int(user_input)
    print(f"\n対象年度: {str(target_year)}\n")

    # 1回目（第１回、患者満足度調査）、２回目（第１回＋第２回患者満足度調査、医療安全、職員インフル）
    print("\n回数を選択してください。")
    print("1. 第1回、2. 第２回")
    user_input = input(":>>")
    if user_input not in ["1", "2"]:
        print("\nエラー: 第1回、第２回のいずれかを入力してください。")
        return -1

    if user_input == "1":
        target_number = "第１回"
    elif user_input == "2":
        target_number = "第２回"

    print(f"\n対象回数: {target_number}\n")

    print("\nレポートの種類を選んでください。")
    print("1:病院名、2:病院番号")
    user_input = input(":>>")
    if user_input not in ["1", "2"]:
        print("\nエラー: 病院名、病院番号のいずれかを入力してください。")
        return -1
    
    target_report_type = CONST.REPORT_TYPES[int(user_input) -1]
    print(f"\n対象レポート種類: {target_report_type}\n")

    print("\nレポートの作成日をyyyy/m/d形式で入力してください。")
    user_input = input(":>>")
    if not re.match(r'^\d{4}/\d{1,2}/\d{1,2}$', user_input):
        print("\nエラー: 日付の形式が不正です。yyyy/m/d形式で入力してください。")
        return -1

    target_create_date = datetime.strptime(user_input, "%Y/%m/%d")
    print(f"\n対象作成日: {target_create_date.strftime('%Y/%m/%d')}\n")

    print("\n参加病院一覧ファイルを指定してください。")
    #ファイルダイアログを開く
    root = tk.Tk()
    #root.withdraw()
    list_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not list_path:
        print("\nエラー: 参加病院一覧ファイルが指定されていません。")
        return -1
    

    #当該年度の平均在院日数（疾患別）から1Qのデータ提出病院のHospital_HOSPITAL_IDを取得し、
    #HOSPITALテーブルから病院ID、病院名を取得する
    print("target_year={}年度第1四半期「平均在院日数（疾患別）」テーブルからデータ提出病院のID、病院名を取得".format(target_year))
    hospital_data = COMMON.getHospIDAndName(target_year)
    if hospital_data is None or hospital_data == -1:
        print("\nエラー: 病院データの取得に失敗しました。年度={}、".format(target_year))
        return -1   
    #cnt = 1
    #for data in hospital_data:
    #    print("No={} data: name={}: no={}".format(cnt,data[1],data[0]))
    #    cnt += 1

    # hospital_dataから「参加病院全体」が含まれている要素を削除する
    hospital_data = [data for data in hospital_data if "参加病院全体" not in data[1]]
    print("\n フィルタリング後の病院データ")
    print("病院データ 取得件数={}、年度={}\n".format(str(len(hospital_data)),target_year, ))

    # 辞書型
    # key:病院名、value:公開用病院番号
    dict_public_no = COMMON.createDictPublicNo(target_year, hospital_data)
    if dict_public_no is None or dict_public_no == -1:
        print("\nエラー: 公開用病院番号の作成に失敗しました。年度={}".format(target_year))
        return -1

    #出力先フォルダの設定
    output_folder = set_output_folder(target_year, target_number, target_report_type)
    if output_folder is None or output_folder == -1:
        return -1

   #表紙を作成
    if(create_cover_page(target_year, target_number, target_create_date, output_folder) == -1):
        return -1

    #病院一覧を作成
    if(COMMON.create_hospitalList(target_year, list_path, output_folder) == -1):
        return -1

    #NIT納品レポートをコピー
    if(copy_source_report(target_year, target_number, target_report_type, output_folder) == -1):
        return -1

    if(copy_TEIGI_FILES(target_year, target_number, output_folder) == -1):
        return -1

    if target_report_type == CONST.REPORT_TYPES[1]:
        #output_folderからエクセルファイルを取得
        excel_files = [f for f in os.listdir(output_folder) if f.endswith(".xlsx")]
        if not excel_files:
            print("\nエラー: エクセルファイルが見つかりません。年度={}、回数={}、レポート種類={}、フォルダパス={}".format(target_year,target_number,target_report_type,output_folder))
            return -1
        
        if(change_report_name(excel_files, output_folder) == -1):
            return -1

        #output_folderからエクセルファイルのパスを取得
        excel_files2 = [f for f in os.listdir(output_folder) if f.endswith(".xlsx")]



        #print("$$$$$  DEBUG  $$$$$")
        for file in excel_files2:
            #fileのファイル名を取得
            print("file:{}".format(file))
            file_path = os.path.join(output_folder, file)
            #print("file_path:{}".format(file_path))
            
            # 61_入院_患者満足度調査
            # 62_入院_病院推奨度調査
            # 63_入院_医師満足度調査
            # 64_外来_患者満足度調査
            # 65_外来_医師満足度調査
            # 66_インシデント・アクシデント（1か月、百床あたり）
            # 67_インシデント・アクシデント（全報告中医師の占める割合）
            # 68_職員の予防接種（インフルエンザ）

            if("入院_患者満足度調査" in str(file)):
                if(tranaHospNameToPublicNO(file_path, dict_public_no, output_folder, "患者満足度") == -1):
                    return -1

            elif("入院_病院推奨度調査" in str(file)):
                if(tranaHospNameToPublicNO(file_path, dict_public_no, output_folder, "病院推奨度") == -1):
                    return -1

            elif("入院_医師満足度調査" in str(file)):
                if(tranaHospNameToPublicNO(file_path, dict_public_no, output_folder, "医師満足度") == -1):
                    return -1
            
            elif("外来_患者満足度調査" in str(file)):
                if(tranaHospNameToPublicNO(file_path, dict_public_no, output_folder, "患者満足度") == -1):
                    return -1
            
            elif("外来_医師満足度調査" in str(file)):
                if(tranaHospNameToPublicNO(file_path, dict_public_no, output_folder, "医師満足度") == -1):
                    return -1
            
            elif("インシデント・アクシデント（1か月、百床あたり）" in str(file)):
                if(tranaHospNameToPublicNO(file_path, dict_public_no, output_folder, "インシデント・アクシデント") == -1):
                    return -1

            elif("インシデント・アクシデント（全報告中医師の占める割合）" in str(file)):
                if(tranaHospNameToPublicNO(file_path, dict_public_no, output_folder, "インシデント・アクシデント医師の占める割合") == -1):
                    return -1
            
            elif("職員の予防接種（インフルエンザ）" in str(file)):
                if(tranaHospNameToPublicNO(file_path, dict_public_no, output_folder, "職員の予防接種") == -1):
                    return -1
          


    if(combine_report_pypdf(target_year, target_number, target_report_type, output_folder) == -1):
        return -1


    return 0