import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

# 2024年度にCI番号変更
# CI_81　→　CI_62
# ただし、DBのテーブルは「CI_81」のまま
def getSQL_CI_62_入院_医師満足度調査_回答率(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_81" + _C + ".年度 ,  " + \
	            "CI_81" + _C + ".期,  " + \
                "CI_81" + _C + ".回答率, " + \
                "CI_81" + _C + ".回答数, " + \
                "CI_81" + _C + ".客体数 " + \
            "FROM " + \
	            "CI_81" + _C + " " + \
            "WHERE " + \
	            "NOT CI_81" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_81" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_81" + _C + ".年度, CI_81" + _C + ".期 " 

def getSQL_CI_62_入院_医師満足度調査_満足度_割合(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_81" + _C + ".年度 ,  " + \
	            "CI_81" + _C + ".期,  " + \
                "CI_81" + _C + ".十分だった_割合, " + \
                "CI_81" + _C + ".まあまあ十分だった_割合, " + \
                "CI_81" + _C + ".あまり十分ではなかった_割合, " + \
                "CI_81" + _C + ".十分ではなかった_割合, " + \
                "CI_81" + _C + ".説明を受けていない_割合, " + \
                "CI_81" + _C + ".無効回答_割合 " + \
            "FROM " + \
	            "CI_81" + _C + " " + \
            "WHERE " + \
	            "NOT CI_81" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_81" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_81" + _C + ".年度, CI_81" + _C + ".期 " 

def getSQL_CI_62_入院_医師満足度調査_満足度_件数(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_81" + _C + ".年度 ,  " + \
	            "CI_81" + _C + ".期,  " + \
                "CI_81" + _C + ".十分だった, " + \
                "CI_81" + _C + ".まあまあ十分だった, " + \
                "CI_81" + _C + ".あまり十分ではなかった, " + \
                "CI_81" + _C + ".十分ではなかった, " + \
                "CI_81" + _C + ".説明を受けていない, " + \
                "CI_81" + _C + ".無効回答 " + \
            "FROM " + \
	            "CI_81" + _C + " " + \
            "WHERE " + \
	            "NOT CI_81" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_81" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_81" + _C + ".年度, CI_81" + _C + ".期 " 

def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_62 開始")

    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_62" in wb.sheetnames:
        wb.remove(wb["CI_62"])
    wb.create_sheet("CI_62")
    sheet = wb["CI_62"]    

    # 書込み列の位置
    colCnt = 1

    # 書込み列の繰り返し数
    roopColCnt = 1
   
    #入院_医師満足度調査_回答率
    temp入院_医師満足度調査_回答率 = X_01.excuteSQL(getSQL_CI_62_入院_医師満足度調査_回答率(HOSPITAL_ID, _C))

    # 1行目に病院名
    sheet.cell(1, 1 ).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, 1 ).value = "入院_医師満足度調査_回答率"
    # 3行目に疾患名
    sheet.cell(3, 1 ).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, 1 ).value = ""

    # 5行名からヘッダを出力
    sheet.cell(5, roopColCnt).value = "年度"
    sheet.cell(5, roopColCnt + 1).value = "期"
    sheet.cell(5, roopColCnt + 2).value = "回答率"
    sheet.cell(5, roopColCnt + 3).value = "回答数"
    sheet.cell(5, roopColCnt + 4).value = "客体数"
    sheet.cell(5, roopColCnt + 5).value = "回答率_ラベル"
    sheet.cell(5, roopColCnt + 6).value = "回答数_ラベル"
    sheet.cell(5, roopColCnt + 7).value = "客体数_ラベル"
    sheet.cell(5, roopColCnt + 8).value = "回答率_比較用"
    sheet.cell(5, roopColCnt + 9).value = "回答数_比較用"
    sheet.cell(5, roopColCnt + 10).value = "客体数_比較用"
    
    if temp入院_医師満足度調査_回答率 is None:
        print("エラー: 入院_医師満足度調査_回答率の取得に失敗しました。")
        return -1
    
    # 6行目からデータ入力
    rowCnt = 6
    for tempRow in temp入院_医師満足度調査_回答率:

        #年度
        sheet.cell(rowCnt, roopColCnt ).value = tempRow[0]
        #期
        sheet.cell(rowCnt, roopColCnt  + 1).value = tempRow[1]
        
        #回答率
        if tempRow[2] == -1 or tempRow[2] == -2:
            sheet.cell(rowCnt, roopColCnt  + 2).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt  + 2).value = tempRow[2] 
        #回答数
        if tempRow[3] == -1 or tempRow[3] == -2:
            sheet.cell(rowCnt, roopColCnt  + 3).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt  + 3).value = tempRow[3] 
        #客体数
        if tempRow[4] == -1 or tempRow[4] == -2:
            sheet.cell(rowCnt, roopColCnt  + 4).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt  + 4).value = tempRow[4] 

        #回答率_ラベル
        if tempRow[2] == -1:
            sheet.cell(rowCnt,  roopColCnt + 5).value = "N/A"
        elif tempRow[2] == -2:
            sheet.cell(rowCnt,  roopColCnt + 5).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 5).value = tempRow[2]
        #回答数_ラベル
        if tempRow[3] == -1:
            sheet.cell(rowCnt,  roopColCnt + 6).value = "N/A"
        elif tempRow[3] == -2:
            sheet.cell(rowCnt,  roopColCnt + 6).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 6).value = tempRow[3]
        #客体数_ラベル
        if tempRow[4] == -1:
            sheet.cell(rowCnt,  roopColCnt + 7).value = "N/A"
        elif tempRow[4] == -2:
            sheet.cell(rowCnt,  roopColCnt + 7).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 7).value = tempRow[4]

        #回答率_比較用
        sheet.cell(rowCnt,  roopColCnt + 8).value = tempRow[2]
        #回答数_比較用
        sheet.cell(rowCnt,  roopColCnt + 9).value = tempRow[3]
        #客体数_比較用
        sheet.cell(rowCnt,  roopColCnt + 10).value = tempRow[4]

        rowCnt+=1
    colCnt+=1


    # 書込み列の繰り返し数
    roopColCnt = 13
   
    #getSQL_CI_81_入院_医師満足度調査_満足度_割合
    temp入院_医師満足度調査_満足度_割合 = X_01.excuteSQL(getSQL_CI_62_入院_医師満足度調査_満足度_割合(HOSPITAL_ID, _C))

    # 1行目に病院名
    sheet.cell(1, roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, roopColCnt).value = "入院_医師満足度調査_満足度_割合"
    # 3行目に疾患名
    sheet.cell(3, roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, roopColCnt).value = ""

    # 5行名からヘッダを出力
    sheet.cell(5, roopColCnt).value = "年度"
    sheet.cell(5, roopColCnt + 1).value = "期"
    sheet.cell(5, roopColCnt + 2).value = "十分だった_割合"
    sheet.cell(5, roopColCnt + 3).value = "まあまあ十分だった_割合"
    sheet.cell(5, roopColCnt + 4).value = "あまり十分ではなかった_割合"
    sheet.cell(5, roopColCnt + 5).value = "十分ではなかった_割合"
    sheet.cell(5, roopColCnt + 6).value = "説明を受けていない_割合"
    sheet.cell(5, roopColCnt + 7).value = "無効回答_割合"
    sheet.cell(5, roopColCnt + 8).value = "十分だった_割合_ラベル"
    sheet.cell(5, roopColCnt + 10).value = "あまり十分ではなかった_割合_ラベル"
    sheet.cell(5, roopColCnt + 11).value = "十分ではなかった_割合_ラベル"
    sheet.cell(5, roopColCnt + 12).value = "説明を受けていない_割合_ラベル"
    sheet.cell(5, roopColCnt + 13).value = "無効回答_割合_ラベル"
    sheet.cell(5, roopColCnt + 14).value = "十分だった_割合_比較用"
    sheet.cell(5, roopColCnt + 15).value = "まあまあ十分だった_割合_比較用"
    sheet.cell(5, roopColCnt + 16).value = "あまり十分ではなかった_割合_比較用"
    sheet.cell(5, roopColCnt + 17).value = "十分ではなかった_割合_比較用"
    sheet.cell(5, roopColCnt + 18).value = "説明を受けていない_割合_比較用"
    sheet.cell(5, roopColCnt + 19).value = "無効回答_割合_比較用"

    if temp入院_医師満足度調査_満足度_割合 is None:
        print("エラー: 入院_医師満足度調査_満足度_割合の取得に失敗しました。")
        return -1

    # 6行目からデータ入力
    rowCnt = 6
    for tempRow in temp入院_医師満足度調査_満足度_割合:

        #年度
        sheet.cell(rowCnt, roopColCnt).value = tempRow[0]
        #期
        sheet.cell(rowCnt, roopColCnt + 1).value = tempRow[1]
        
        #十分だった_割合
        if tempRow[2] == -1 or tempRow[2] == -2:
            sheet.cell(rowCnt, roopColCnt + 2).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 2).value = tempRow[2] 
        #まあまあ十分だった_割合
        if tempRow[3] == -1 or tempRow[3] == -2:
            sheet.cell(rowCnt, roopColCnt + 3).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 3).value = tempRow[3] 
        #あまり十分ではなかった_割合
        if tempRow[4] == -1 or tempRow[4] == -2:
            sheet.cell(rowCnt, roopColCnt + 4).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 4).value = tempRow[4] 
        #十分ではなかった_割合
        if tempRow[5] == -1 or tempRow[5] == -2:
            sheet.cell(rowCnt, roopColCnt + 5).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 5).value = tempRow[5] 
        #説明を受けていない_割合
        if tempRow[6] == -1 or tempRow[6] == -2:
            sheet.cell(rowCnt, roopColCnt + 6).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 6).value = tempRow[6]         
        #無効回答_割合
        if tempRow[7] == -1 or tempRow[7] == -2:
            sheet.cell(rowCnt, roopColCnt + 7).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 7).value = tempRow[7] 

        #十分だった_割合_ラベル
        if tempRow[2] == -1:
            sheet.cell(rowCnt,  roopColCnt + 8).value = "N/A"
        elif tempRow[2] == -2:
            sheet.cell(rowCnt,  roopColCnt + 8).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 8).value = tempRow[2]
        #やまあまあ十分だった_割合_ラベル
        if tempRow[3] == -1:
            sheet.cell(rowCnt,  roopColCnt + 9).value = "N/A"
        elif tempRow[3] == -2:
            sheet.cell(rowCnt,  roopColCnt + 9).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 9).value = tempRow[3]
        #あまり十分ではなかった_割合_ラベル
        if tempRow[4] == -1:
            sheet.cell(rowCnt,  roopColCnt + 10).value = "N/A"
        elif tempRow[4] == -2:
            sheet.cell(rowCnt,  roopColCnt + 10).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 10).value = tempRow[4]
        #十分ではなかった_割合_ラベル
        if tempRow[5] == -1:
            sheet.cell(rowCnt,  roopColCnt + 11).value = "N/A"
        elif tempRow[5] == -2:
            sheet.cell(rowCnt,  roopColCnt + 11).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 11).value = tempRow[5]
        #説明を受けていない_割合_ラベル
        if tempRow[6] == -1:
            sheet.cell(rowCnt,  roopColCnt + 12).value = "N/A"
        elif tempRow[6] == -2:
            sheet.cell(rowCnt,  roopColCnt + 12).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 12).value = tempRow[6]
        #無効回答_割合_ラベル
        if tempRow[7] == -1:
            sheet.cell(rowCnt,  roopColCnt + 13).value = "N/A"
        elif tempRow[7] == -2:
            sheet.cell(rowCnt,  roopColCnt + 13).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 13).value = tempRow[7]

        #十分だった_割合_比較用
        sheet.cell(rowCnt, roopColCnt + 14).value = tempRow[2]
        #まあまあ十分だった_割合_比較用
        sheet.cell(rowCnt, roopColCnt + 15).value = tempRow[3]
        #あまり十分ではなかった_割合_比較用
        sheet.cell(rowCnt, roopColCnt + 16).value = tempRow[4]
        #十分ではなかった_割合_比較用
        sheet.cell(rowCnt, roopColCnt + 17).value = tempRow[5]
        #説明を受けていない_割合_比較用
        sheet.cell(rowCnt, roopColCnt + 18).value = tempRow[6]
        #無効回答_割合_比較用
        sheet.cell(rowCnt, roopColCnt + 19).value = tempRow[7]

        rowCnt+=1
    colCnt+=1


    # 書込み列の繰り返し数
    roopColCnt = 34
   
    #入院_医師満足度調査_満足度_件数
    temp入院_医師満足度調査_満足度_件数 = X_01.excuteSQL(getSQL_CI_62_入院_医師満足度調査_満足度_件数(HOSPITAL_ID, _C))

    # 1行目に病院名
    sheet.cell(1, roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, roopColCnt).value = "入院_医師満足度調査_満足度_件数"
    # 3行目に疾患名
    sheet.cell(3, roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, roopColCnt).value = ""

    # 5行名からヘッダを出力
    sheet.cell(5, roopColCnt).value = "年度"
    sheet.cell(5, roopColCnt + 1).value = "期"
    sheet.cell(5, roopColCnt + 2).value = "十分だった"
    sheet.cell(5, roopColCnt + 3).value = "まあまあ十分だった"
    sheet.cell(5, roopColCnt + 4).value = "あまり十分ではなかった"
    sheet.cell(5, roopColCnt + 5).value = "十分ではなかった"
    sheet.cell(5, roopColCnt + 6).value = "説明を受けていない"
    sheet.cell(5, roopColCnt + 7).value = "無効回答"
    sheet.cell(5, roopColCnt + 8).value = "十分だった_ラベル"
    sheet.cell(5, roopColCnt + 9).value = "まあまあ十分だった_ラベル"
    sheet.cell(5, roopColCnt + 10).value = "あまり十分ではなかった_ラベル"
    sheet.cell(5, roopColCnt + 11).value = "十分ではなかった_ラベル"
    sheet.cell(5, roopColCnt + 12).value = "説明を受けていない_ラベル"
    sheet.cell(5, roopColCnt + 13).value = "無効回答_ラベル"
    sheet.cell(5, roopColCnt + 14).value = "十分だった_比較用"
    sheet.cell(5, roopColCnt + 15).value = "まあまあ十分だった_比較用"
    sheet.cell(5, roopColCnt + 16).value = "あまり十分ではなかった_比較用"
    sheet.cell(5, roopColCnt + 17).value = "十分ではなかった_比較用"
    sheet.cell(5, roopColCnt + 18).value = "説明を受けていない_比較用"
    sheet.cell(5, roopColCnt + 19).value = "無効回答_比較用"

    if temp入院_医師満足度調査_満足度_件数 is None:
        print("エラー: 入院_医師満足度調査_満足度_件数の取得に失敗しました。")
        return -1

    # 6行目からデータ入力
    rowCnt = 6
    for tempRow in temp入院_医師満足度調査_満足度_件数:

        #年度
        sheet.cell(rowCnt, roopColCnt).value = tempRow[0]
        #期
        sheet.cell(rowCnt, roopColCnt + 1).value = tempRow[1]
        
        #十分だった
        if tempRow[2] == -1 or tempRow[2] == -2:
            sheet.cell(rowCnt, roopColCnt + 2).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 2).value = tempRow[2] 
        #まあまあ十分だった
        if tempRow[3] == -1 or tempRow[3] == -2:
            sheet.cell(rowCnt, roopColCnt + 3).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 3).value = tempRow[3] 
        #あまり十分ではなかった
        if tempRow[4] == -1 or tempRow[4] == -2:
            sheet.cell(rowCnt, roopColCnt + 4).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 4).value = tempRow[4] 
        #十分ではなかった
        if tempRow[5] == -1 or tempRow[5] == -2:
            sheet.cell(rowCnt, roopColCnt + 5).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 5).value = tempRow[5] 
        #説明を受けていない
        if tempRow[6] == -1 or tempRow[6] == -2:
            sheet.cell(rowCnt, roopColCnt + 6).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 6).value = tempRow[6]         
        #無効回答
        if tempRow[7] == -1 or tempRow[7] == -2:
            sheet.cell(rowCnt, roopColCnt + 7).value = 0
        else:
            sheet.cell(rowCnt, roopColCnt + 7).value = tempRow[7] 

        #十分だった_ラベル
        if tempRow[2] == -1:
            sheet.cell(rowCnt,  roopColCnt + 8).value = "N/A"
        elif tempRow[2] == -2:
            sheet.cell(rowCnt,  roopColCnt + 8).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 8).value = tempRow[2]
        #やまあまあ十分だった_ラベル
        if tempRow[3] == -1:
            sheet.cell(rowCnt,  roopColCnt + 9).value = "N/A"
        elif tempRow[3] == -2:
            sheet.cell(rowCnt,  roopColCnt + 9).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 9).value = tempRow[3]
        #あまり十分ではなかった_ラベル
        if tempRow[4] == -1:
            sheet.cell(rowCnt,  roopColCnt + 10).value = "N/A"
        elif tempRow[4] == -2:
            sheet.cell(rowCnt,  roopColCnt + 10).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 10).value = tempRow[4]
        #十分ではなかった_ラベル
        if tempRow[5] == -1:
            sheet.cell(rowCnt,  roopColCnt + 11).value = "N/A"
        elif tempRow[5] == -2:
            sheet.cell(rowCnt,  roopColCnt + 11).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 11).value = tempRow[5]
        #説明を受けていない_ラベル
        if tempRow[6] == -1:
            sheet.cell(rowCnt,  roopColCnt + 12).value = "N/A"
        elif tempRow[6] == -2:
            sheet.cell(rowCnt,  roopColCnt + 12).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 12).value = tempRow[6]
        #無効回答_ラベル
        if tempRow[7] == -1:
            sheet.cell(rowCnt,  roopColCnt + 13).value = "N/A"
        elif tempRow[7] == -2:
            sheet.cell(rowCnt,  roopColCnt + 13).value = "-"
        else:
            sheet.cell(rowCnt,  roopColCnt + 13).value = tempRow[7]

        #十分だった_比較用
        sheet.cell(rowCnt,  roopColCnt + 14).value = tempRow[2]
        #まあまあ十分だった_比較用
        sheet.cell(rowCnt,  roopColCnt + 15).value = tempRow[3]
        #あまり十分ではなかった_比較用
        sheet.cell(rowCnt,  roopColCnt + 16).value = tempRow[4]
        #十分ではなかった_比較用
        sheet.cell(rowCnt,  roopColCnt + 17).value = tempRow[5]
        #説明を受けていない_比較用
        sheet.cell(rowCnt,  roopColCnt + 18).value = tempRow[6]
        #無効回答_比較用
        sheet.cell(rowCnt,  roopColCnt + 19).value = tempRow[7]

        rowCnt+=1
    colCnt+=1

    #wb.save(wbPath)
    print(" " + HOSPITAL_NAME + " CI_62 終了")


