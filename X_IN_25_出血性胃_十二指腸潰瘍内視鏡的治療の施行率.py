import sqlite3
import openpyxl
import X_00_CONST as CONST
import X_01_レポート作成 as X_01

def getSQL_CI_25_出血性胃_十二指腸潰瘍内視鏡的治療の施行率(HOSPITAL_ID, _C):
    return  "SELECT  " + \
	            "CI_25" + _C + ".年度 ,  " + \
	            "CI_25" + _C + ".期,  " + \
                "CI_25" + _C + ".出血性胃_十二指腸潰瘍に対する内視鏡的治療_止血術_の施行率, " + \
                "CI_25" + _C + ".うち_内視鏡的消化管止血術施行症例数, " + \
                "CI_25" + _C + ".胃潰瘍あるいは十二指腸潰瘍_急性_出血を伴_に該当する退院症例数 " + \
            "FROM " + \
	            "CI_25" + _C + " " + \
            "WHERE " + \
	            "NOT CI_25" + _C + ".期 = 'TOTAL' AND " + \
	            "CI_25" + _C + ".Hospital_HOSPITAL_ID = " + str(HOSPITAL_ID) + " " + \
            " ORDER BY CI_25" + _C + ".年度, CI_25" + _C + ".期 " 

def getRepAgeData(wb, HOSPITAL_ID, HOSPITAL_NAME, is_acute):
    print(" " + HOSPITAL_NAME + " CI_25 開始")
    if is_acute == "急性期":
        _C = ""
    else:
        _C = "_C"

    #エクセル（マクロ付き）を開く
    #wb = openpyxl.load_workbook(wbPath, keep_vba=True)
    
    # 既存シートが存在する場合は削除
    if "CI_25" in wb.sheetnames:
        wb.remove(wb["CI_25"])
    wb.create_sheet("CI_25")
    sheet = wb["CI_25"]    

    # 書込み列の位置
    colCnt = 1

    # 書込み列の繰り返し数
    roopColCnt = 12

    #出血性胃_十二指腸潰瘍に対する内視鏡的治療_止血術_の施行率
    temp出血性胃_十二指腸潰瘍に対する内視鏡的治療_止血術_の施行率 = X_01.excuteSQL(getSQL_CI_25_出血性胃_十二指腸潰瘍内視鏡的治療の施行率(HOSPITAL_ID, _C))

    # 1行目に病院名
    sheet.cell(1, 1 + (colCnt - 1) * roopColCnt).value = HOSPITAL_NAME
    # 2行目に指標名
    sheet.cell(2, 1 + (colCnt - 1) * roopColCnt).value = "出血性胃・十二指腸潰瘍に対する内視鏡的治療_止血術_の施行率"
    # 3行目に疾患名
    sheet.cell(3, 1 + (colCnt - 1) * roopColCnt).value = ""
    # 4行名に重症度、年代別、性別
    sheet.cell(4, 1 + (colCnt - 1) * roopColCnt).value = ""

    # 5行名からヘッダを出力
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt).value = "年度"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 1).value = "期"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 2).value = "出血性胃_十二指腸潰瘍に対する内視鏡的治療_止血術_の施行率"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 3).value = "うち_内視鏡的消化管止血術施行症例数"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 4).value = "胃潰瘍あるいは十二指腸潰瘍_急性_出血を伴_に該当する退院症例数"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 5).value = "出血性胃_十二指腸潰瘍に対する内視鏡的治療_止血術_の施行率_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 6).value = "うち_内視鏡的消化管止血術施行症例数_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 7).value = "胃潰瘍あるいは十二指腸潰瘍_急性_出血を伴_に該当する退院症例数_ラベル"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 8).value = "出血性胃_十二指腸潰瘍に対する内視鏡的治療_止血術_の施行率_比較用"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 9).value = "うち_内視鏡的消化管止血術施行症例数_比較用"
    sheet.cell(5, 1 + (colCnt - 1) * roopColCnt + 10).value = "胃潰瘍あるいは十二指腸潰瘍_急性_出血を伴_に該当する退院症例数_比較用"

    if temp出血性胃_十二指腸潰瘍に対する内視鏡的治療_止血術_の施行率 is None:
        print("エラー: 出血性胃_十二指腸潰瘍に対する内視鏡的治療_止血術_の施行率の取得に失敗しました。")
        return -1

    # 6行目からデータ入力
    rowCnt = 6
    for tempRow in temp出血性胃_十二指腸潰瘍に対する内視鏡的治療_止血術_の施行率:

        #年度
        sheet.cell(rowCnt, 1 + (colCnt - 1) * roopColCnt).value = tempRow[0]
        #期
        sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 1).value = tempRow[1]
        
        #出血性胃_十二指腸潰瘍に対する内視鏡的治療_止血術_の施行率
        if tempRow[2] == -1 or tempRow[2] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 2).value = tempRow[2]  / 100
        #うち_内視鏡的消化管止血術施行症例数
        if tempRow[3] == -1 or tempRow[3] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 3).value = tempRow[3] 
        #胃潰瘍あるいは十二指腸潰瘍_急性_出血を伴_に該当する退院症例数
        if tempRow[4] == -1 or tempRow[4] == -2:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 4).value = 0
        else:
            sheet.cell(rowCnt, 1 +  (colCnt - 1) * roopColCnt + 4).value = tempRow[4] 

        #出血性胃_十二指腸潰瘍に対する内視鏡的治療_止血術_の施行率_ラベル
        if tempRow[2] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "N/A"
        elif tempRow[2] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 5).value = tempRow[2] / 100
        #うち_内視鏡的消化管止血術施行症例数ラベル
        if tempRow[3] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = "N/A"
        elif tempRow[3] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 6).value = tempRow[3]
        #胃潰瘍あるいは十二指腸潰瘍_急性_出血を伴_に該当する退院症例数_ラベル
        if tempRow[4] == -1:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = "N/A"
        elif tempRow[4] == -2:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = "-"
        else:
            sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 7).value = tempRow[4]

        #出血性胃_十二指腸潰瘍に対する内視鏡的治療_止血術_の施行率_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 8).value = tempRow[2] / 100
        #うち_内視鏡的消化管止血術施行症例数_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 9).value = tempRow[3]
        #胃潰瘍あるいは十二指腸潰瘍_急性_出血を伴_に該当する退院症例数_比較用
        sheet.cell(rowCnt,  1 + (colCnt - 1) * roopColCnt + 10).value = tempRow[4]

        rowCnt+=1
    colCnt+=1
    #wb.save(wbPath)
    print(" " + HOSPITAL_NAME + " CI_25 終了")

