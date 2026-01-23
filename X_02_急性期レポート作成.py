    
#パラメータは年度、グループ（急性期グループ、慢性期グループ）、期間（1Q、2Q、3Q、4Q、年間）、病院名or病院番号

import sys
import sqlite3
import X_00_CONST as CONST
import os
import shutil
import openpyxl
import re
from datetime import datetime
from openpyxl.styles import PatternFill
import win32com.client
import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import Border, Side
#from PyPDF2 import PdfMerger, PdfWriter, PdfReader
from pypdf import PdfReader, PdfWriter
from pathlib import Path
from copy import copy
import X_01_レポート作成 as COMMON


#出力フォルダを設定する ※各レポートで個別実装
#引数：target_year：対象年度、target_group：対象グループ、target_period：対象期間
#戻り値：出力先フォルダのパス
#ERROR:-1
def set_output_folder(target_year,target_group,target_period,target_report_type):
    
    try:
        #出力先に当該年度のフォルダがあるか確認する
        output_folder = CONST.OUTPUT_FOLDER
        output_folder = os.path.join(output_folder, str(target_year))
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        #output_folder内に「01_急性期指標」フォルダがあるか確認する
        acute_folder = os.path.join(output_folder, "01_急性期指標")
        if not os.path.exists(acute_folder):
            os.makedirs(acute_folder)
    
        #acute_folder内に当該グループのフォルダがあるか確認する
        group_folder = os.path.join(acute_folder, target_group)
        if not os.path.exists(group_folder):
            os.makedirs(group_folder)
    
        #group_folder内に当該期間のフォルダがあるか確認する
        period_folder = os.path.join(group_folder, target_period) 
        if not os.path.exists(period_folder):
            os.makedirs(period_folder)

        #period_folder内に当該レポート種類のフォルダがあるか確認する
        report_type_folder = os.path.join(period_folder, target_report_type)
        if not os.path.exists(report_type_folder):
            os.makedirs(report_type_folder)
    except:
        print("エラー: 出力先フォルダの設定に失敗しました。年度={}、グループ={}、期間={}、レポート種類={}".format(target_year,target_group,target_period,target_report_type))
        return -1
      
    #report_type_folder内のファイル、フォルダを全て削除
    delete_file_count = 0
    for file in os.listdir(report_type_folder):
        try:
            os.remove(os.path.join(report_type_folder, file))
            #削除したファイル数をカウント
            delete_file_count += 1
        except:
            print("フォルダ：{}".format(report_type_folder))
            print(f"エラー: {file} の削除に失敗しました。")
            print("エラー内容：{}".format(sys.exc_info()[1]))
            return -1

    print("削除したファイル数：{}".format(delete_file_count))
    return report_type_folder

def create_cover_page(target_year,target_group,target_period,target_report_type, target_create_date, output_folder):
    print("\n表紙ファイルの作成を開始します。")
    try:
        #表紙ファイルを開く
        over_page_path = CONST.COVER_PAGE_FORMAT_PATH_ACUTE
        #openpyxl で開く
        workbook = openpyxl.load_workbook(over_page_path)
        if workbook is None:
            print("\nエラー: 表紙ファイルの読み込みに失敗しました。年度={}、グループ={}、期間={}、レポート種類={}".format(target_year,target_group,target_period,target_report_type))
            print("ファイルパス：{}".format(over_page_path))
            return -1

        #シート名「00-1_表紙」を開く
        sheet = workbook["00-1_表紙"]
        #セルCD1に指標タイプを入力
        sheet["CD1"] = "急性期指標"
        #セルCD2に年度を入力
        sheet["CD2"] = target_year
        #セルCD3にグループを入力
        if(target_group == "01_急性期グループ"):
            sheet["CD3"] = "急性期グループ"
        elif(target_group == "02_慢性期グループ"):
            sheet["CD3"] = "慢性期グループ"
        else:
            print("\nエラー: グループが不正です。年度={}、グループ={}、期間={}、レポート種類={}".format(target_year,target_group,target_period,target_report_type))
            return -1

         #セルCD4に期間を入力
        if(target_period == "1Q"):
            sheet["CD4"] = "第１四半期（４～６月）"
        elif(target_period == "2Q"):
            sheet["CD4"] = "第２四半期（７～９月）"
        elif(target_period == "3Q"):
            sheet["CD4"] = "第３四半期（１０～１２月）"
        elif(target_period == "4Q"):
            sheet["CD4"] = "第４四半期（１～３月）"
        elif(target_period == "年間"):
            sheet["CD4"] = "年間"
        else:
            print("\nエラー: 期間が不正です。年度={}、グループ={}、期間={}、レポート種類={}".format(target_year,target_group,target_period,target_report_type))
            return -1

        #R6診療報酬改定の影響で、2024年1Qの期間が4～5月になったため、期間を変更
        if(target_year == 2024 and target_period == "1Q"):
            sheet["CD4"] = "第１四半期（４～５月）"
        elif(target_year == 2024 and target_period == "2Q"):
            sheet["CD4"] = "第２四半期（６～９月）"

        #セルCD5に作成日を入力
        sheet["CD5"] = target_create_date
        #セルA4の色を変更
        sheet["A4"].fill = PatternFill(start_color=CONST.REPORT_INDCTR_COLOR[0], end_color=CONST.REPORT_INDCTR_COLOR[0], fill_type="solid")
        #セルA22の色を変更
        color_index = 0
        if(target_group == "01_急性期グループ"):
            color_index = 0
        elif(target_group == "02_慢性期グループ"):
            color_index = 1
        sheet["A22"].fill = PatternFill(start_color=CONST.REPORT_GROUP_COLOR[color_index], end_color=CONST.REPORT_GROUP_COLOR[color_index], fill_type="solid")
    
        output_folder = os.path.join(output_folder, "00-1_表紙.xlsx")
        workbook.save(output_folder)
        if workbook is None:
            print("\nエラー: 表紙ファイルの保存に失敗しました。年度={}、グループ={}、期間={}、レポート種類={}".format(target_year,target_group,target_period,target_report_type))
            print("ファイルパス：{}".format(output_folder))
            return -1

        #PDFに変換
        # Excel ファイルを開く
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Excel のウィンドウを非表示にする
        wb = excel.Workbooks.Open(output_folder)
        newPath = str(output_folder).replace(".xlsx",".pdf")
        wb.ExportAsFixedFormat(0, newPath)  # 0 は PDF 形式
        wb.Close()
        excel.Quit()

        #pdfファイルの作成が完了したら元のエクセルファイルを削除
        os.remove(output_folder)

        print(f"\n表紙ファイルを作成しました。ファイルパス: {output_folder}\n")
        return 0
    except:
        print("\nエラー: 表紙ファイルの作成に失敗しました。年度={}、グループ={}、期間={}、レポート種類={}".format(target_year,target_group,target_period,target_report_type))
        print("ファイルパス：{}".format(output_folder))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1

#NIT納品レポートを作業用フォルダにコピー
#引数：target_year：対象年度、target_group：対象グループ、target_period：対象期間、target_report_type：レポート種類
#ERROR:-1
def copy_source_report(target_year,target_group,target_period,target_report_type, output_folder):
    print("\nNIT納品レポートのコピーを開始します。")

    TEMP =""
    if(target_report_type == "01_病院名"):
        TEMP = "01_PDF"
    elif(target_report_type == "02_病院番号"):
        TEMP = "02_EXCEL"
    else:
        print("\nエラー: レポート種類が不正です。年度={}、グループ={}、期間={}、レポート種類={}".format(target_year,target_group,target_period,target_report_type))
        return -1

    #NIT納品レポートのフォルダを取得
    nit_report_folder = CONST.NIT_REPORT_FOLDER
    nit_report_folder = os.path.join(nit_report_folder, str(target_year))
    nit_report_folder = os.path.join(nit_report_folder, "01_急性期指標")
    nit_report_folder = os.path.join(nit_report_folder, target_group)
    nit_report_folder = os.path.join(nit_report_folder, target_period)
    nit_report_folder = os.path.join(nit_report_folder, TEMP)

    #nit_report_folderが存在しない場合はエラー  
    if not os.path.exists(nit_report_folder):
        print("\nエラー: NIT納品レポートのフォルダが存在しません。年度={}、グループ={}、期間={}、レポート種類={}".format(target_year,target_group,target_period,target_report_type))
        print("フォルダパス：{}".format(nit_report_folder))
        return -1
    #nit_report_folder内にファイルが存在しない場合はエラー
    if not os.listdir(nit_report_folder):
        print("\nエラー: NIT納品レポートのフォルダにファイルが存在しません。年度={}、グループ={}、期間={}、レポート種類={}".format(target_year,target_group,target_period,target_report_type))
        print("フォルダパス：{}".format(nit_report_folder))
        return -1

    try:
        TRANS_FILE_NAMES = CONST.DICT_REPORT_FILE_NAMES_ACUTE_2024_1Q
        if(target_year == 2024 and target_period == "1Q"):
            TRANS_FILE_NAMES = CONST.DICT_REPORT_FILE_NAMES_ACUTE_2024_1Q
        elif(target_year == 2024 and target_period not in ["1Q"]):
            TRANS_FILE_NAMES = CONST.DICT_REPORT_FILE_NAMES_ACUTE_FROM_2024_2Q
        elif(target_year > 2024 ):
            TRANS_FILE_NAMES = CONST.DICT_REPORT_FILE_NAMES_ACUTE_FROM_2024_2Q

        #TRANS_FILE_NAMES内のファイルをコピー
        #ファイル名の頭2文字を取得し、TRANS_FILE_NAMES内に存在するか確認
        #一致していたらファイル名を変換してコピー
        for file in os.listdir(nit_report_folder):
            #ファイル名の頭2文字を取得
            file_head2 = file.split("_")[0]
            if file_head2 in TRANS_FILE_NAMES:
                #TRANS_FILE_NAMESからfile_head2をキーにファイル名を取得
                new_file_name = TRANS_FILE_NAMES[file_head2]
                if(target_report_type == "02_病院番号"):
                    new_file_name = new_file_name.replace(".pdf",".xlsx")

                #DEBUG
                #もしPDFファイルならファイル名とページ数を出力
                #if(file.endswith(".pdf")):
                #    reader = PdfReader(os.path.join(nit_report_folder, file))
                #    num_pages = len(reader.pages)
                #    print("ファイル名：{}、ページ数：{}".format(file, num_pages))

                #new_file_nameにファイル名を変更してコピー
                new_file_name = os.path.join(output_folder, new_file_name)
                shutil.copy(os.path.join(nit_report_folder, file), new_file_name)
            else:
                print("\nエラー: レポートファイルが名称が一致しません。年度={}、グループ={}、期間={}、レポート種類={}".format(target_year,target_group,target_period,target_report_type))
                print("ファイル名：{}".format(file))
                return -1


        print("\nNIT納品レポートのコピーに成功しました。年度={}、グループ={}、期間={}、レポート種類={}".format(target_year,target_group,target_period,target_report_type))
        print("フォルダパス：{}".format(nit_report_folder))
        print("出力フォルダパス：{}".format(output_folder))
        #コピーしたファイルの数を出力
        print("コピーしたファイルの数：{}".format(len(os.listdir(output_folder))))
        return 0

    except:
        print("\nエラー: NIT納品レポートのコピーに失敗しました。年度={}、グループ={}、期間={}、レポート種類={}".format(target_year,target_group,target_period,target_report_type))
        print("フォルダパス：{}".format(nit_report_folder))
        print("出力フォルダパス：{}".format(output_folder))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1



def copy_TEIGI_FILES(target_year, target_period, output_folder):
    print("\n指標定義ファイルのコピーを開始します。")
    try:
        TEIGI_FILES_PATH = CONST.TEIGI_FILES_ACUTE_2024_1Q
        if(target_year == 2024 and target_period == "1Q"):
            TEIGI_FILES_PATH = CONST.TEIGI_FILES_ACUTE_2024_1Q
        elif(target_year == 2024 and target_period not in ["1Q"]):
            TEIGI_FILES_PATH = CONST.TEIGI_FILES_ACUTE_FROM_2024_2Q
        elif(target_year > 2024 ):
            TEIGI_FILES_PATH = CONST.TEIGI_FILES_ACUTE_FROM_2024_2Q

        #TEIGI_FILES_PATH内のpdfファイルをoutput_folderにコピー
        for file in os.listdir(TEIGI_FILES_PATH):
            if file.endswith(".pdf"):
                shutil.copy(os.path.join(TEIGI_FILES_PATH, file), os.path.join(output_folder, file))

        return 0
    except:
        print("\nエラー: 指標定義ファイルのコピーに失敗しました。年度={}".format(target_year))
        print("ファイルパス：{}".format(TEIGI_FILES_PATH))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1





#
# 予定しない再入院率のシートの結合セルを解除する
#
def formatCollection(fileName, sheet):
    try:
        col1Str = ""
        col2Str = ""
        delIndx = 0

        if fileName.startswith("08"):
            col1Str = "C"
            col2Str = "D"
            delIndx = 3
        else:
            col1Str = "D"
            col2Str = "E"
            delIndx = 4

        sheet.column_dimensions[col1Str].hidden = False
	
	    # セル範囲が結合されているか確認する
        merged_cells = sheet.merged_cells.ranges

        max_row = sheet.max_row
        #print(f"シートの最大行数: {max_row}")

        for range in merged_cells:
		    #print(f"結合されたセル範囲: {range}")
            if col1Str in str(range) and col2Str in str(range):
			    # 結合セルを解除
                sheet.unmerge_cells(str(range))
                #print(f"結合解除: {range}")
			    #range.coordを「D4:E4」形式で取得
                cell_range = str(range)
			    #「:」で分割
                cells = cell_range.split(":")
			    # コピー元セルとコピー先セルを指定
                source_cell = sheet[cells[0]]
                destination_cell = sheet[cells[1]]
			    # コピー元のセルの値をコピー先のセルに設定
                destination_cell.value = source_cell.value
			    # コピー元のセルのスタイルをコピー先のセルに適用
                if source_cell.has_style:
                    destination_cell._style = copy(source_cell._style)
	    # 列Cまたは列Dを削除
        sheet.delete_cols(delIndx)
        return sheet
    except:
        print("\nエラー: 予定しない再入院率のシートの結合セルの解除に失敗しました。")
        print("ファイル名：{}".format(fileName))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        print("シート名：{}".format(sheet.title))
        print("行数：{}".format(sheet.max_row))
        print("列数：{}".format(sheet.max_column))


def formatCollection_修正案1(fileName, sheet):
    try:
        col1Str = ""
        col2Str = ""
        delIndx = 0

        if fileName.startswith("08"):
            col1Str = "C"
            col2Str = "D"
            delIndx = 3
        else:
            col1Str = "D"
            col2Str = "E"
            delIndx = 4

        sheet.column_dimensions[col1Str].hidden = False
	
	    # セル範囲が結合されているか確認する
        merged_cells = sheet.merged_cells.ranges

        max_row = sheet.max_row
        #print(f"シートの最大行数: {max_row}")

        # 修正案1：リストに変換してから反復処理を行う
        for range in list(merged_cells):
		    #print(f"結合されたセル範囲: {range}")
            if col1Str in str(range) and col2Str in str(range):
			    # 結合セルを解除
                sheet.unmerge_cells(str(range))
                #print(f"結合解除: {range}")
			    #range.coordを「D4:E4」形式で取得
                cell_range = str(range)
			    #「:」で分割
                cells = cell_range.split(":")
			    # コピー元セルとコピー先セルを指定
                source_cell = sheet[cells[0]]
                destination_cell = sheet[cells[1]]
			    # コピー元のセルの値をコピー先のセルに設定
                destination_cell.value = source_cell.value
			    # コピー元のセルのスタイルをコピー先のセルに適用
                if source_cell.has_style:
                    destination_cell._style = copy(source_cell._style)
	    # 列Cまたは列Dを削除
        sheet.delete_cols(delIndx)
        return sheet
    except:
        print("\nエラー: 予定しない再入院率のシートの結合セルの解除に失敗しました。")
        print("ファイル名：{}".format(fileName))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        print("シート名：{}".format(sheet.title))
        print("行数：{}".format(sheet.max_row))
        print("列数：{}".format(sheet.max_column))


def torans_HospNameToHospNumber(target_year, target_group, target_period, target_report_type, dict_public_no, output_folder):
    print("\n病院名を病院番号に変換を開始します。")
    #output_folder内のエクセルファイルを取得
    excel_files = [f for f in os.listdir(output_folder) if f.endswith(".xlsx")]
    if not excel_files:
        print("\nエラー: エクセルファイルが見つかりません。年度={}、グループ={}、期間={}、レポート種類={}、フォルダパス={}".format(target_year,target_group,target_period,target_report_type,output_folder))
        return -1

    try:
        cnt = 1
        for file in excel_files:
            #エクセルファイルを開く
            wb = openpyxl.load_workbook(os.path.join(output_folder, file))
            if wb is None:
                print("\nエラー: エクセルファイルの読み込みに失敗しました。年度={}、グループ={}、期間={}、レポート種類={}、フォルダパス={}".format(target_year,target_group,target_period,target_report_type,output_folder))
                #ファイル名を出力
                print("ファイル名：{}".format(file))
                return -1

            # シートを順に確認
            for sheet in wb.worksheets:
                if sheet.title != "帳票":
                    sheet.sheet_state = 'hidden'  # 非表示にする

            #エクセルファイルのシートを取得
            ws = wb["帳票"]
            if ws is None:
                print("\nエラー: 帳票シートが見つかりません。年度={}、グループ={}、期間={}、レポート種類={}、フォルダパス={}".format(target_year,target_group,target_period,target_report_type,output_folder))
                print("ファイル名：{}".format(file))
                return -1
        
            if "予定しない再入院率" in file:
                #print("予定しない再入院率")
                ws = formatCollection_修正案1(file, ws)
            
            if "死亡率" not in file and "医療費" not in file:
                #fileの先頭３文字を削除
                file_name = file[3:]
                file_name = file_name.replace("レポート.xlsx","")            

                if file_name not in COMMON.getDict_HospRow_StartCol():
                    print("エラー: file_nameがhospital_dataに存在しません")
                    print("ファイル名：{}  修正後のファイル名：{}".format(file,file_name))
                    return -1

                dict_HospRow_StartCol = COMMON.getDict_HospRow_StartCol()[file_name]
                hospRow = dict_HospRow_StartCol[0]
                startCol = dict_HospRow_StartCol[1]
                #print("hospRow:"+str(hospRow))
                #print("startCol:"+str(startCol))

                #病院行を取得
                row_data = [cell.value for cell in ws[hospRow]]
                #print("row_data:"+str(row_data))
            
                for i in range(startCol, len(row_data) + 1, 1):
                    cell_value = ws.cell(row=hospRow, column=i).value
                    if cell_value is None:
                        print(f"空白セルをスキップ: 行={hospRow}, 列={i}")
                        continue

                    if cell_value in dict_public_no:
                        pNo = dict_public_no[cell_value]
                    else:
                        print(f"エラー: 病院名が辞書に存在しません。病院名: {cell_value}")
                        print("row数：{}".format(hospRow))
                        print("col数：{}".format(i))
                        print("ファイル名：{}".format(file))
                        return -1

                    ws.cell(row=hospRow, column=i, value=str(pNo)) 
    
                wb.save(os.path.join(output_folder, file))
                wb.close()

            #PDFに変換
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False  # Excel のウィンドウを非表示にする
            wb = excel.Workbooks.Open(os.path.join(output_folder, file))
            # 指定シートを取得
            ws = wb.Sheets("帳票")
            newPath = str(os.path.join(output_folder, file.replace(".xlsx",".pdf")))

            ws.ExportAsFixedFormat(0, newPath)  # 0 は PDF 形式
            wb.Close()
            excel.Quit()

            print("  No.{} {} 番号変換、PDF化  完了".format(cnt, file))
            cnt += 1
        
        return 0
    except Exception as e:
        print("\nエラー: 病院名を病院番号に変換に失敗しました。年度={}、グループ={}、期間={}、レポート種類={}、フォルダパス={}".format(target_year,target_group,target_period,target_report_type,output_folder))
        
        print("エラー内容：{}".format(sys.exc_info()[1]))
        print(f"エラー内容: {e}")
        print("ファイル名：{}".format(file))
        
        print("行数：{}".format(hospRow))
        print("列数：{}".format(startCol))  
        print("病院番号：{}".format(pNo))
        return -1









#def combine_report(target_year, target_group, target_period, target_report_type, output_folder):
    print("\nレポートのPDF結合を開始します。")
    try:
        #output_folder内のpdfファイルを結合
        pdf_files = [f for f in os.listdir(output_folder) if f.endswith(".pdf")]
        pdf_files.sort()
        if not pdf_files:
            print("\nエラー: レポートファイルが見つかりません。年度={}、グループ={}、期間={}、レポート種類={}、フォルダパス={}".format(target_year,target_group,target_period,target_report_type,output_folder))
            return -1

        report_file_name = str(target_year) + "年度_急性期指標レポート_" + target_group + "_" + target_period + "_" + target_report_type + ".pdf"
        #出力先はoutput_folderの３階層上のフォルダに作成
        reportoutput_folder = os.path.join(os.path.dirname(output_folder), "..", "..")


        #PDFファイルを結合、ファイル名をしおりに追加
        writer = PdfWriter()
        current_page = 0

        for file in pdf_files:
            reader = PdfReader(os.path.join(output_folder, file))
            num_pages = len(reader.pages)

            # 各ページをwriterに追加
            for page in reader.pages:
                writer.add_page(page)

            # ファイル名をしおりとして追加（拡張子なしで追加もOK）
            title = os.path.splitext(os.path.basename(file))[0]
            if "レポート" in title:
                current_page += num_pages
                continue
            writer.add_outline_item(title=title, page_number=current_page)
            # 次のしおりのページ位置を把握
            current_page += num_pages

            
            
        # 保存
        with open(os.path.join(reportoutput_folder, report_file_name), "wb") as f:
            writer.write(f)

        print("\nレポートのPDF結合を完了しました。:{}".format(report_file_name))

        #pdf_filesを結合
        #pdf_merger = PdfMerger()
        #for file in pdf_files:
        #    pdf_merger.append(os.path.join(output_folder, file))
        #pdf_merger.write(os.path.join(reportoutput_folder, report_file_name))

    except:
        print("\nエラー: レポートの結合に失敗しました。年度={}、グループ={}、期間={}、レポート種類={}、フォルダパス={}".format(target_year,target_group,target_period,target_report_type,output_folder))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1

#def combine_report2(target_year, target_group, target_period, target_report_type, output_folder):
    print("\nレポートのPDF結合を開始します。")
    try:
        #output_folder内のpdfファイルを結合
        pdf_files = [f for f in os.listdir(output_folder) if f.endswith(".pdf")]
        pdf_files.sort()
        if not pdf_files:
            print("\nエラー: レポートファイルが見つかりません。年度={}、グループ={}、期間={}、レポート種類={}、フォルダパス={}".format(target_year,target_group,target_period,target_report_type,output_folder))
            return -1

        report_file_name = str(target_year) + "年度_急性期指標レポート_" + target_group + "_" + target_period + "_" + target_report_type + ".pdf"
        #出力先はoutput_folderの３階層上のフォルダに作成
        reportoutput_folder = os.path.join(os.path.dirname(output_folder), "..", "..")


        #PDFファイルを結合、ファイル名をしおりに追加
        writer = PdfWriter()
        current_page = 0

        for file in pdf_files:
            file_path = os.path.join(output_folder, file)
            try:
                reader = PdfReader(file_path)
                num_pages = len(reader.pages)

                # 各ページをwriterに追加
                for page in reader.pages:
                    writer.add_page(page)

                # ファイル名をしおりとして追加（拡張子なしで追加もOK）
                title = os.path.splitext(os.path.basename(file))[0]
                if "レポート" not in title:
                    writer.add_outline_item(title=title, page_number=current_page)

                current_page += num_pages

            except Exception as e:
                print(f"\n警告: PDFファイルの読み込みに失敗しました。スキップします: {file}")
                print(f"理由: {e}")
                return -1

            
            
        # 保存
        with open(os.path.join(reportoutput_folder, report_file_name), "wb") as f:
            writer.write(f)

        print("\nレポートのPDF結合を完了しました。:{}".format(report_file_name))

        #pdf_filesを結合
        #pdf_merger = PdfMerger()
        #for file in pdf_files:
        #    pdf_merger.append(os.path.join(output_folder, file))
        #pdf_merger.write(os.path.join(reportoutput_folder, report_file_name))

    except:
        print("\nエラー: レポートの結合に失敗しました。年度={}、グループ={}、期間={}、レポート種類={}、フォルダパス={}".format(target_year,target_group,target_period,target_report_type,output_folder))
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1



def combine_report_pypdf(target_year, target_group, target_period, target_report_type, output_folder):
    print("\nレポートのPDF結合を開始します。")
    try:
        # output_folder 内の pdf ファイルを取得
        pdf_files = [f for f in os.listdir(output_folder) if f.endswith(".pdf")]
        pdf_files.sort()

        if not pdf_files:
            print("\nエラー: レポートファイルが見つかりません。年度={}、グループ={}、期間={}、レポート種類={}、フォルダパス={}".format(
                target_year, target_group, target_period, target_report_type, output_folder))
            return -1

        # 出力先
        report_file_name = f"{target_year}年度_急性期指標レポート_{target_group}_{target_period}_{target_report_type}.pdf"
        report_output_folder = os.path.abspath(os.path.join(output_folder, "..", ".." , ".."))
        output_path = os.path.join(report_output_folder, report_file_name)

        writer = PdfWriter()
        current_page = 0
        error_files = []

        for file in pdf_files:
            file_path = os.path.join(output_folder, file)
            try:
                reader = PdfReader(file_path)
                num_pages = len(reader.pages)

                # ページ追加
                for page in reader.pages:
                    writer.add_page(page)

                # しおり追加（「レポート」を含むファイルは除外）
                title = os.path.splitext(file)[0]
                if "レポート" not in title:
                    writer.add_outline_item(title=title, page_number=current_page)

                current_page += num_pages

            except Exception as e:
                print(f"\n警告: ファイル '{file}' の読み込みに失敗しました。スキップします。理由: {e}")
                error_files.append(file)

        # 書き出し
        with open(output_path, "wb") as f:
            writer.write(f)

        print(f"\nレポートのPDF結合を完了しました: {report_file_name}")

        if error_files:
            print("\n以下のPDFは読み込めなかったため、除外されました:")
            for ef in error_files:
                print(f" - {ef}")

        return 0

    except Exception as e:
        print("\nエラー: レポートの結合に失敗しました。")
        print(f"エラー内容：{e}")
        return -1


#
#
# レポート作成のメインメソッド
# ユーザーの入力項目を取得
#　１：年度
#　２：グループ
#　３：期間
#　４：レポート種類
#　５：作成日
#　６：参加病院一覧ファイル
#
#
def create_report():
    print("\n急性期指標レポートを作成します。")

    print("\n年度（西暦4桁、半角）を入力してください。")
    user_input = input(":>>")
    if not user_input.isdigit() or len(user_input) != 4:
        print("\nエラー: 西暦4桁の半角数字を入力してください。")
        return -1

    target_year = int(user_input)
    print(f"\n対象年度: {str(target_year)}\n")

    print("\nグループ（急性期グループ、慢性期グループ）を選んでください。")
    print("1:急性期グループ、2:慢性期グループ")
    user_input = input(":>>")
    if user_input not in ["1", "2"]:
        print("\nエラー: 急性期グループ、慢性期グループのいずれかを入力してください。")
        return -1

    target_group = CONST.GROUPS[int(user_input)-1]
    print(f"\n対象グループ: {target_group}\n")

    print("\n1Q、2Q、3Q、4Q、年間のいずれかを選んでください。")
    print("1:1Q、2:2Q、3:3Q、4:4Q、5:年間")
    user_input = input(":>>")
    if user_input not in ["1", "2", "3", "4", "5"]:
        print("\nエラー: 1Q、2Q、3Q、4Q、年間のいずれかを入力してください。")
        return -1

    target_period = CONST.PERIODS[int(user_input) -1]
    print(f"\n対象期間: {target_period}\n")

    print("\nレポートの種類を選んでください。")
    print("1:病院名、2:病院番号")
    user_input = input(":>>")
    if user_input not in ["1", "2"]:
        print("\nエラー: 病院名、病院番号のいずれかを入力してください。")
        return -1
    
    target_report_type = CONST.REPORT_TYPES[int(user_input) -1]
    print(f"\n対象レポート種類: {target_report_type}\n")

    print("\nレポートの作成日をyyyy/m/d形式で入力してください。")
    user_input = input(":>>")
    if not re.match(r'^\d{4}/\d{1,2}/\d{1,2}$', user_input):
        print("\nエラー: 日付の形式が不正です。yyyy/m/d形式で入力してください。")
        return -1

    target_create_date = datetime.strptime(user_input, "%Y/%m/%d")
    print(f"\n対象作成日: {target_create_date.strftime('%Y/%m/%d')}\n")

    print("\n参加病院一覧ファイルを指定してください。")
    #ファイルダイアログを開く
    root = tk.Tk()
    #root.withdraw()
    list_path = filedialog.askopenfilename(initialdir=CONST.HOSP_LIST_PATH,filetypes=[("Excel files", "*.xlsx")])
    if not list_path:
        print("\nエラー: 参加病院一覧ファイルが指定されていません。")
        return -1

    

    #当該年度の平均在院日数（疾患別）から1Qのデータ提出病院のHospital_HOSPITAL_IDを取得し、
    #HOSPITALテーブルから病院ID、病院名を取得する
    print("target_year={}年度第1四半期「平均在院日数（疾患別）」テーブルからデータ提出病院のID、病院名を取得".format(target_year))
    hospital_data = COMMON.getHospIDAndName(target_year)
    if hospital_data is None or hospital_data == -1:
        print("\nエラー: 病院データの取得に失敗しました。年度={}、".format(target_year))
        return -1   
    #cnt = 1
    #for data in hospital_data:
    #    print("No={} data: name={}: no={}".format(cnt,data[1],data[0]))
    #    cnt += 1

    # hospital_dataから「参加病院全体」が含まれている要素を削除する
    hospital_data = [data for data in hospital_data if "参加病院全体" not in data[1]]
    print("\n フィルタリング後の病院データ")
    print("病院データ 取得件数={}、年度={}\n".format(str(len(hospital_data)),target_year, ))

    # 辞書型
    # key:病院名、value:公開用病院番号
    dict_public_no = COMMON.createDictPublicNo(target_year, hospital_data)
    if dict_public_no is None or dict_public_no == -1:
        print("\nエラー: 公開用病院番号の作成に失敗しました。年度={}".format(target_year))
        return -1

 
    #出力先フォルダの設定
    output_folder = set_output_folder(target_year,target_group,target_period,target_report_type)
    if output_folder is None or output_folder == -1:
        return -1

    #
    #テスト用
    #
    #output_folder = r"C:\Users\yoshida\Desktop\アウトカム"

    #表紙を作成
    if(create_cover_page(target_year,target_group,target_period,target_report_type, target_create_date, output_folder) == -1):
        return -1
    
    #病院一覧を作成
    if(COMMON.create_hospitalList(target_year, list_path, output_folder) == -1):
        return -1

    #NIT納品レポートを作業用フォルダにコピー
    if(copy_source_report(target_year,target_group,target_period,target_report_type, output_folder) == -1):
        return -1

    #指標定義ファイルを作業用フォルダにコピー
    if(copy_TEIGI_FILES(target_year, target_period, output_folder) == -1):
        return -1

    #REPORT_TYPEが02_病院番号の場合、エクセルの病院名を病院番号に変換してPDF化する
    if(target_report_type == "02_病院番号"):
        if(torans_HospNameToHospNumber(target_year, target_group, target_period, target_report_type, dict_public_no, output_folder) == -1):
            return -1



    #レポートを結合
    if(combine_report_pypdf(target_year, target_group, target_period, target_report_type, output_folder) == -1):
        return -1

    

    

    
 





 
