import copy
import re
from datetime import datetime
import os
import sys
import X_00_CONST as CONST
import X_01_レポート作成 as X_01
import openpyxl
import shutil
import X_IN_01_平均在院日数_疾患別 as IN_01
import X_IN_02_平均在院日数_重症度別 as IN_02
import X_IN_03_平均在院日数_年代別 as IN_03
import X_IN_04_平均在院日数_性別 as IN_04
import X_IN_05_死亡率_疾患別_重症度別 as IN_05
import X_IN_06_死亡率_疾患別_年代別 as IN_06
import X_IN_07_死亡率_疾患別_性別 as IN_07
import X_IN_08_予定しない再入院率_疾患別 as IN_08
import X_IN_09_予定しない再入院率_疾患別_重症度別 as IN_09
import X_IN_10_予定しない再入院率_疾患別_年代別 as IN_10
import X_IN_11_予定しない再入院率_疾患別_性別 as IN_11
import X_IN_12_医療費_疾患別_重症度別 as IN_12
import X_IN_13_医療費_疾患別_年代別 as IN_13
import X_IN_14_医療費_疾患別_性別 as IN_14
import X_IN_15_肺血栓塞栓症の予防対策実施率_施設別 as IN_15
import X_IN_16_肺血栓塞栓症の発生率_管理料算定有無別_施設別 as IN_16
import X_IN_17_褥瘡の持込率_施設別_年代別 as IN_17
import X_IN_18_褥瘡の発生率_施設別_年代別 as IN_18
import X_IN_19_65歳以上の患者における認知症の保有率_入院経路別 as IN_19
import X_IN_20_急性心筋梗塞における主要手術_治療実施率_重症度別 as IN_20
import X_IN_21_急性心筋梗塞における主要手術_治療実施率_年代別 as IN_21
import X_IN_22_急性心筋梗塞における主要手術_治療実施率_性別 as IN_22
import X_IN_23_急性心筋梗塞における主要手術_治療実施率_Kコード別 as IN_23
import X_IN_24_肺炎に対する入院当日の抗生物質使用率_重症度別 as IN_24
import X_IN_25_出血性胃_十二指腸潰瘍内視鏡的治療の施行率 as IN_25
import X_IN_26_急性脳梗塞患者リハビリテーション開始率 as IN_26
import X_IN_27_胃がんの患者に対するESDの施行率 as IN_27
import X_IN_28_胆嚢切除術患者に対する腹腔鏡下手術施行率 as IN_28
import X_IN_29_虫垂炎の患者に対する手術施行率 as IN_29
import X_IN_30_虫垂切除術患者に対する腹腔鏡下手術施行率 as IN_30
import X_IN_31_帝王切開における全身麻酔施行率 as IN_31
import X_IN_32_帝王切開における輸血施行率 as IN_32
import X_IN_33_脳卒中地域連携パスの使用率 as IN_33
import X_IN_34_脳卒中地域連携パスの使用率_転院先 as IN_34
import X_IN_35_大腿骨地域連携パスの使用率 as IN_35
import X_IN_36_大腿骨頸部骨折連携パスの使用率_転院先 as IN_36
import X_IN_37_急性期病棟における退院調整の実施率 as IN_37
import X_IN_38_介護施設_福祉施設からの入院率_年代別 as IN_38
import X_IN_39_他の病院_診療所の病棟からの転院率_年代別 as IN_39
import X_IN_40_介護施設_福祉施設等への退院率_年代別 as IN_40
import X_IN_41_介護施設_福祉施設等への退院率_退院先別 as IN_41
import X_IN_42_他の病院_診療所の病棟への転院率_年代別 as IN_42
import X_IN_43_自宅退院患者における在宅医療を受ける率_年代別 as IN_43
import X_IN_44_中心静脈カテーテル挿入時の気胸発生率 as IN_44
import X_IN_45_急性心筋梗塞患者におけるアスピリン_早期投与率 as IN_45
import X_IN_46_急性心筋梗塞患者におけるアスピリン_退院時投与 as IN_46
import X_IN_47_Door_to_balloon実施率 as IN_47
import X_IN_48_誤嚥性肺炎患者に対する喉頭ファイバースコピーあるいは嚥下造影検査の実施率 as IN_48
import X_IN_49_術後24時間以内の予防的抗菌薬の投与停止率 as IN_49
import X_IN_50_術後48時間以内の予防的抗菌薬の投与停止率 as IN_50
import X_IN_51_服薬指導_安全管理が必要な薬剤の服薬指導実施率 as IN_51
import X_IN_52_服薬指導_薬剤管理指導実施率 as IN_52
import X_IN_53_栄養指導 as IN_53
import X_IN_54_血液培養の実施_血液培養実施率 as IN_54
import X_IN_55_血液培養の実施_2セット実施率 as IN_55
import X_IN_56_65_歳以上の患者の入院早期の栄養アセスメント実施割合 as IN_56
import X_IN_57_身体的拘束の実施率 as IN_57
import X_IN_58_手術開始前1時間以内の予防的抗菌薬投与率 as IN_58
import X_IN_59_転倒転落発生率 as IN_59
import X_IN_59_2_転倒_転落 as IN_59_2
import X_IN_60_転棟転落によるインシデント影響度分類レベル3b以上の発生率 as IN_60
import X_IN_60_2_転倒_転落_3b以上 as IN_60_2


#出力フォルダを設定する ※各レポートで個別実装
#引数：target_year：対象年度、target_group：対象グループ、target_period：対象期間
#戻り値：出力先フォルダのパス
#ERROR:-1
def set_output_folder(target_year):
    
    try:
        #出力先に当該年度のフォルダがあるか確認する
        output_folder = CONST.OUTPUT_FOLDER
        output_folder = os.path.join(output_folder, str(target_year))
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        #output_folder内に「05_個別レポート」フォルダがあるか確認する
        individual_folder = os.path.join(output_folder, "05_個別レポート")
        if not os.path.exists(individual_folder):
            os.makedirs(individual_folder)
    
        #acute_folder内に当該グループのフォルダがあるか確認する
        acute_folder = os.path.join(individual_folder, "01_急性期指標")
        if not os.path.exists(acute_folder):
            os.makedirs(acute_folder)
    except:
        print("エラー: 出力先フォルダの設定に失敗しました。年度={}".format(target_year))
        return -1
      
    #report_type_folder内のファイル、フォルダを全て削除
    delete_file_count = 0
    for file in os.listdir(acute_folder):
        try:
            os.remove(os.path.join(acute_folder, file))
            #削除したファイル数をカウント
            delete_file_count += 1
        except:
            print("フォルダ：{}".format(acute_folder))
            print(f"エラー: {file} の削除に失敗しました。")
            print("エラー内容：{}".format(sys.exc_info()[1]))
            return -1

    print("削除したファイル数：{}".format(delete_file_count))
    return acute_folder


# レポート作成対象となる病院を取得
# 指標「01_平均在院日数_疾患別」の最新の年度、期に登録されている病院を対象とする
def get_target_hospitals(target_year,is_acute):
    print("\nレポート作成対象となる病院を取得します。")

    hospID_list = []

    if is_acute == "急性期":
        ##急性期グループ
        ##
        sql_getMaxYear_急性期 = "SELECT MAX(年度) FROM TABLE_STATUS" + " WHERE QI_NAME = '01_平均在院日数_疾患別' AND STATUS = '登録'"
        maxYear_急性期 =X_01.excuteSQL(sql_getMaxYear_急性期, None)
        if maxYear_急性期 is None:
            print("エラー: 最新年度の取得に失敗しました。")
            return -1
            #print("maxYear_急性期")
        #   print(maxYear_急性期[0][0])

        sql_getMaxQuater_急性期 = "SELECT MAX(期) FROM TABLE_STATUS" + " WHERE QI_NAME = '01_平均在院日数_疾患別' AND STATUS = '登録'" \
            + " AND 年度 = " + str(maxYear_急性期[0][0] ) + " "
        maxQuater_急性期 =X_01.excuteSQL(sql_getMaxQuater_急性期, None)
        if maxQuater_急性期 is None:
            print("エラー: 最新期の取得に失敗しました。")
            return -1

        CONST.INDIV_TARGET_QUARTER_急性期 = maxQuater_急性期[0][0]  

        sql_getTargetHospitals_急性期  = "SELECT DISTINCT Hospital_HOSPITAL_ID FROM CI_01" + " WHERE 年度 = " \
            + str(maxYear_急性期[0][0]) + " AND 期 = '" + maxQuater_急性期[0][0]  + "' AND Disease_DISEASE_ID = 25 " \
            + " AND 平均在院日数 <> -2  "
        hospID_急性期 =X_01.excuteSQL(sql_getTargetHospitals_急性期, None)
        if hospID_急性期 is None:
            print("エラー: 対象病院の取得に失敗しました。")
            return -1
        
        for a in range(len(hospID_急性期)):
            hospID_list.append(hospID_急性期[a][0])

    elif is_acute == "慢性期":
        ##
        ##慢性期グループ
        ##
        sql_getMaxYear_慢性期 = "SELECT MAX(年度) FROM TABLE_STATUS_C" + " WHERE QI_NAME = '01_平均在院日数_疾患別' AND STATUS = '登録'"
        maxYear_慢性期 =X_01.excuteSQL(sql_getMaxYear_慢性期, None)
        if maxYear_慢性期 is None:
            print("エラー: 最新年度の取得に失敗しました。")
            return -1

        sql_getMaxQuater_慢性期 = "SELECT MAX(期) FROM TABLE_STATUS_C" + " WHERE QI_NAME = '01_平均在院日数_疾患別' AND STATUS = '登録'" \
            + " AND 年度 = " + str(maxYear_慢性期[0][0] ) + " "
        maxQuater_慢性期 =X_01.excuteSQL(sql_getMaxQuater_慢性期, None)
        if maxQuater_慢性期 is None:
            print("エラー: 最新期の取得に失敗しました。")
            return -1

        CONST.INDIV_TARGET_QUARTER_慢性期 = maxQuater_慢性期[0][0]

        sql_getTargetHospitals_慢性期  = "SELECT DISTINCT Hospital_HOSPITAL_ID FROM CI_01_C" + " WHERE 年度 = " \
            + str(maxYear_慢性期[0][0]) + " AND 期 = '" + maxQuater_慢性期[0][0]  + "' AND Disease_DISEASE_ID = 25 " \
            + " AND 平均在院日数 <> -2  "
        hospID_慢性期 =X_01.excuteSQL(sql_getTargetHospitals_慢性期, None)
        if hospID_慢性期 is None:
            print("エラー: 対象病院の取得に失敗しました。")
            return -1

        for a in range(len(hospID_慢性期)):
            hospID_list.append(hospID_慢性期[a][0])


    hosIDs = ",".join(str(hospID) for hospID in hospID_list)

    sql_target_hospName = "SELECT 病院名, HOSPITAL_ID FROM HOSPITAL WHERE HOSPITAL_ID IN (" + hosIDs+ " ) ORDER BY 都道府県コード, 病院名_よみがな"
    temp_hospName_list = X_01.excuteSQL(sql_target_hospName, None)
    if temp_hospName_list is None:
        print("エラー: 対象病院の取得に失敗しました。")
        return -1

    hospName_list = []
    for a in range(len(temp_hospName_list)):
        hospName_list.append((temp_hospName_list[a][0],temp_hospName_list[a][1]))

    #print("hospName_list")
    #print(hospName_list)

    #print("急性期 年度{},期{}の対象病院数:{}".format(maxYear_急性期[0][0], maxQuater_急性期[0][0], len(hospID_list)))
    #print("慢性期 年度{},期{}の対象病院数:{}".format(maxYear_慢性期[0][0], maxQuater_慢性期[0][0], len(hospID_list)))
    print("{} ：対象病院数:{}".format(is_acute,len(hospName_list)))
    return hospName_list

def copyFormatFile(hospName_list, output_folder,is_acute,target_create_date):
    print("\nフォーマットファイルをコピーします。{}".format(is_acute))

    #病院の数だけ繰り返し
    repCnt = 1

    _C = ""
    if is_acute == "慢性期":
        _C = "_C"
        temp_repCnt = 0
        #output_folder内のxlsmファイルを取得
        xlsm_file_list = [f for f in os.listdir(output_folder) if f.endswith('.xlsm')]

        for xlsm_file in xlsm_file_list:
            xlsm_file_name = os.path.splitext(xlsm_file )[0]
            #xlsm_file_name内にある数値２桁を取得
            xlsm_file_name_num = re.search(r'\d{2}', xlsm_file_name)
       
            if xlsm_file_name_num is not None:
                xlsm_file_name_num = xlsm_file_name_num.group()
            else:
                print("エラー: 数値２桁が取得できませんでした。")
                print("ファイル名:{}".format(xlsm_file_name))
                return -1

            xlsm_file_name_num = int(xlsm_file_name_num)

            if xlsm_file_name_num > temp_repCnt:
                temp_repCnt = xlsm_file_name_num

        repCnt = temp_repCnt + 1
          
    try:
        format_file_path = CONST.INDIV_REP_FRMT_ACUTE_PATH
 
        for hospName in hospName_list:
        
            if not os.path.exists(format_file_path):
                print("エラー: フォーマットファイルが存在しません。")
                print("フォーマットファイルのパス:{}".format(format_file_path))
                return -1

            temp_repCnt = ""
            if repCnt <= 9:
                temp_repCnt = "0" + str(repCnt)
            else:
                temp_repCnt = str(repCnt)

            newName = "個別レポート_急性期指標_" + str(temp_repCnt) + "_" + hospName[0] + ".xlsm"
            #format_file_pathをnewNameでコピー
            shutil.copy(format_file_path, os.path.join(output_folder, newName))             
            repCnt += 1

    except:
        print("エラー: フォーマットファイルのコピーに失敗しました。")
        print("エラー内容：{}".format(sys.exc_info()[1]))
        return -1

    print("フォーマットファイルのコピーに成功しました。{}".format(is_acute))
    return 0


def create_indiv_report(target_create_date, hosp_list, output_folder,is_acute):
    print("\n個別レポートを作成します。{}".format(is_acute))

    #output_folder内のxlsmファイルを取得
    xlsm_file_list = [f for f in os.listdir(output_folder) if f.endswith('.xlsm')]
    if len(xlsm_file_list) == 0:
        #存在しない場合もあるので正常終了で返す
        print("エラー: xlsmファイルが存在しません。")
        return 0

    # hospName_list
    # 0 病院名
    # 1 病院ID
    for hosp in hosp_list:

        for xlsm_file in xlsm_file_list:
            if hosp[0] in xlsm_file:

                #
                # debug
                #
                # 大分記念病院、和田病院、鹿児島生協病院、いまきいれ総合病院、浦添総合病院以外はcontinue
                if hosp[0] != "大分記念病院" and hosp[0] != "和田病院" and hosp[0] != "鹿児島生協病院" and hosp[0] != "いまきいれ総合病院" and hosp[0] != "浦添総合病院":
                    continue
                
                print("{} を作成します。".format(xlsm_file))
                print("病院ID:{}".format(hosp[1]))
                print("病院名:{}".format(hosp[0]))
                print("フォーマットファイル名:{}".format(xlsm_file))


                wb = openpyxl.load_workbook(os.path.join(output_folder, xlsm_file), keep_vba=True)

                # 竹口病院は2023は慢性期、2024は急性期、慢性期の両方に提出していたので慢性期からデータを取得
                #if hosp[0] == "竹口病院" :
                #    is_acute = "慢性期"

                ws = wb["レポート_表紙"]
                ws["CF4"] = hosp[0]
                ws["CF5"] = target_create_date.strftime("%Y/%m/%d")

                IN_01.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_02.getRepAgeData_重症度別(wb, hosp[1], hosp[0], is_acute)
                #開発中のみコメントアウト
                IN_03.getRepAgeData_年代別(wb, hosp[1], hosp[0], is_acute)
                IN_04.getRepAgeData_性別(wb, hosp[1], hosp[0], is_acute)
                # 死亡率 全体集計かつ処理が重すぎるのでコメントアウト
                #IN_05.getRepAgeData(wb, is_acute)
                #IN_06.getRepAgeData(wb, is_acute)
                #IN_07.getRepAgeData(wb, is_acute)
                IN_08.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_09.getRepAgeData_重症度別(wb, hosp[1], hosp[0], is_acute)
                #開発中のみコメントアウト
                IN_10.getRepAgeData_年代別(wb, hosp[1], hosp[0], is_acute)
                IN_11.getRepAgeData_性別(wb, hosp[1], hosp[0], is_acute)
                # 医療費 全体集計かつ処理が重すぎるのでコメントアウト
                #IN_12.getRepAgeData(wb, is_acute)
                #IN_13.getRepAgeData(wb, is_acute)
                #IN_14.getRepAgeData(wb, is_acute)
                IN_15.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_16.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_17.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_18.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_19.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_20.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_21.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_22.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_23.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_24.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_25.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_26.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_27.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_28.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_29.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_30.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_31.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_32.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_33.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_34.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_35.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_36.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_37.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_38.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_39.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_40.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_41.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_42.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_43.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_44.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_45.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_46.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_47.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_48.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_49.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_50.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_51.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_52.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_53.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_54.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_55.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_56.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_57.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                IN_58.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                # 転倒_転落 2024以降
                IN_59.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                # 転倒_転落 2023まで
                IN_59_2.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                # 転倒_転落_3b以上 2024以降
                IN_60.getRepAgeData(wb, hosp[1], hosp[0], is_acute)
                # 転倒_転落_3b以上 2023まで
                IN_60_2.getRepAgeData(wb, hosp[1], hosp[0], is_acute)


                wb.save(os.path.join(output_folder, xlsm_file))
                wb.close()

                





def create_report():
    print("\n経年比較  急性期指標レポートを作成します。")

    print("\n年度（西暦4桁、半角）を入力してください。")
    user_input = input(":>>")
    if not user_input.isdigit() or len(user_input) != 4:
        print("\nエラー: 西暦4桁の半角数字を入力してください。")
        return -1

    target_year = int(user_input)

    print("\nレポートの作成日をyyyy/m/d形式で入力してください。")
    user_input = input(":>>")
    if not re.match(r'^\d{4}/\d{1,2}/\d{1,2}$', user_input):
        print("\nエラー: 日付の形式が不正です。yyyy/m/d形式で入力してください。")
        return -1

    target_create_date = datetime.strptime(user_input, "%Y/%m/%d")
    print(f"\n対象作成日: {target_create_date.strftime('%Y/%m/%d')}\n")

    #出力先フォルダの設定
    output_folder = set_output_folder(target_year)
    if output_folder is None or output_folder == -1:
        return -1


    #作成対象となる病院を取得
    #まれに慢性期の方に多くデータを提出している病院がいるので注意！
    # 現状、「竹口病院」のみ　2024年度末時点

    # 0 病院名
    # 1 病院ID
    hospID_list_急性期 = get_target_hospitals(target_year,"急性期")
    if hospID_list_急性期 == -1:
        return -1

    # 0 病院名
    # 1 病院ID
    hospID_list_慢性期 = get_target_hospitals(target_year,"慢性期")
    if hospID_list_慢性期 == -1:
        return -1
    #hospID_list_慢性期からhospID_list_急性期との重複を除く

    print("hospID_list_急性期" + str(hospID_list_急性期))


    hospID_list_慢性期 = list(set(hospID_list_慢性期) - set(hospID_list_急性期))

    print("hospID_list_慢性期" + str(hospID_list_慢性期))

    if copyFormatFile(hospID_list_急性期, output_folder,"急性期",target_create_date) == -1:   
        return -1
    if copyFormatFile(hospID_list_慢性期, output_folder,"慢性期",target_create_date) == -1:
        return -1

    # マスタデータをセット
    X_01.setMastaData()



    create_indiv_report(target_create_date, hospID_list_急性期, output_folder,"急性期")
    create_indiv_report(target_create_date, hospID_list_慢性期, output_folder,"慢性期")




    return 0